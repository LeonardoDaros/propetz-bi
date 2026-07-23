# -*- coding: utf-8 -*-
"""Constrói a nova tabela de pedidos Propetz (versão cliente) a partir do modelo extraído."""
import sys, json, os
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
MODELO = json.load(open(os.path.join(BASE, "_modelo_tabela.json"), encoding="utf-8"))

# FOB (frete a cotar) ou CIF (frete incluso: preços de tabela e promo +3,5%
# — regra Leonardo 22/07). O gerar_tabela_pedido.bat produz OS DOIS.
MODALIDADE = os.environ.get("PEDIDO_MODALIDADE", "FOB").upper()
FATOR = 1.035 if MODALIDADE == "CIF" else 1.0

def _preco(v):
    """Aplica o fator da modalidade (CIF = +3,5%) com arredondamento a centavo."""
    return round(float(v) * FATOR, 2)

SAIDA = os.path.join(BASE, f"Pedido Propetz Distribuição {MODALIDADE} - AGOSTO 2026.xlsx")
SENHA = "propetz2026"

TEAL_ESCURO = "FF085041"
TEAL = "FF0F6E56"
TEAL_CLARO = "FFE1F5EE"
AZUL_INPUT = "FFEFF6FF"
LARANJA_CLARO = "FFFFF2CC"
LARANJA_TXT = "FFC55A11"
CINZA_TXT = "FF64748B"
BRANCO = "FFFFFFFF"

fino = Side(style="thin", color="FFB6D9D9")
borda = Border(left=fino, right=fino, top=fino, bottom=fino)
f_normal = Font(name="Inter", size=10)
f_peq = Font(name="Inter", size=9, color=CINZA_TXT)

produtos = [p for p in MODELO["produtos"]
            if not p["sem_estoque"] and not p["linha_oculta"] and p["preco_tabela"]]
excluidos = [p["sku"] for p in MODELO["produtos"] if p not in produtos]
promos = {k: v for k, v in MODELO["promos_reais"].items() if k in {p["sku"] for p in produtos}}

wb = openpyxl.Workbook()

# ============================ ABA PEDIDO ============================
ws = wb.active
ws.title = "PEDIDO"
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:L1")
c = ws["A1"]
c.value = f"PROPETZ  ·  PEDIDO DISTRIBUIÇÃO {MODALIDADE}  —  AGOSTO 2026"
c.font = Font(name="Inter", size=16, bold=True, color=BRANCO)
c.fill = PatternFill("solid", fgColor=TEAL_ESCURO)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:L2")
c = ws["A2"]
_frete_txt = ("FRETE JÁ INCLUÍDO nos preços (CIF)" if MODALIDADE == "CIF"
              else "Frete a cotar")
c.value = ("Vigência 01–31/08/2026   ·   À VISTA (PIX/transferência): 5% de desconto no total   ·   "
           f"A PRAZO: boleto 21/35/49/63/77/91 (sujeito a análise)   ·   {_frete_txt}   ·   "
           "LINHAS LARANJA = PROMOÇÃO DE AGOSTO — preço promocional JÁ APLICADO nos totais   ·   "
           "Equipamentos, lâminas, adaptadores e tesouras: QUANTIDADE MÍNIMA (acima dela, qualquer qtde) "
           "· Acessórios: CAIXA FECHADA — a planilha só aceita múltiplos da caixa")
c.font = Font(name="Inter", size=9, color="FF1E293B")
c.fill = PatternFill("solid", fgColor=TEAL_CLARO)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 30

PRIM = 7  # primeira linha de produto
ULT = PRIM + len(produtos) - 1
rng = lambda col: f"${col}${PRIM}:${col}${ULT}"

resumo = [
    ("A", "B", "UNIDADES NO PEDIDO", f"=SUM({rng('M')})", "#,##0"),
    ("C", "D", "TOTAL A PRAZO (JÁ C/ IPI)", f"=SUM({rng('K')})", '"R$" #,##0.00'),
    ("E", "F", "TOTAL À VISTA −5% (JÁ C/ IPI)", f"=ROUND(SUM({rng('K')})*0.95,2)", '"R$" #,##0.00'),
    ("G", "H", "SUA ECONOMIA C/ PROMOS",
     f"=ROUND(SUMPRODUCT({rng('M')},ROUND({rng('F')}*(1+{rng('I')}),2))-SUM({rng('K')}),2)",
     '"R$" #,##0.00'),
    ("I", "K", "ITENS DIFERENTES NO PEDIDO",
     f"=SUMPRODUCT(--({rng('M')}>0))", "0"),
    ("L", "L", "ITENS C/ AVISO", f'=COUNTIF({rng("L")},"⚠*")', "0"),
]
for col_ini, col_fim, rotulo, formula, fmt in resumo:
    ws.merge_cells(f"{col_ini}3:{col_fim}3")
    lab = ws[f"{col_ini}3"]
    lab.value = rotulo
    lab.font = Font(name="Inter", size=8, bold=True, color=CINZA_TXT)
    lab.alignment = Alignment(horizontal="center")
    ws.merge_cells(f"{col_ini}4:{col_fim}4")
    val = ws[f"{col_ini}4"]
    val.value = formula
    val.number_format = fmt
    val.font = Font(name="Inter", size=12, bold=True, color="FF085041")
    val.alignment = Alignment(horizontal="center")
ws.row_dimensions[3].height = 14
ws.row_dimensions[4].height = 20
ws["L4"].font = Font(name="Inter", size=12, bold=True, color="FFC00000")

ws["A5"] = ("Preencha só a coluna QTD (células azuis) — digite livre, os totais calculam sozinhos. "
            "Se aparecer ⚠ na coluna AVISO, a quantidade está fora do padrão (abaixo do mínimo ou "
            "caixa aberta): o pedido segue, e o vendedor confirma com você. Filtros no cabeçalho.")
ws["A5"].font = f_peq
ws.merge_cells("A5:L5")

cab = ["CATEGORIA", "CÓDIGO", "CÓD. BARRAS", "PRODUTO", "QTD", "PREÇO TABELA",
       "PROMO AGOSTO", "ECONOMIA", "IPI", "PREÇO FINAL C/ IPI", "TOTAL", "AVISO"]
for i, t in enumerate(cab, 1):
    c = ws.cell(6, i, t)
    c.font = Font(name="Inter", size=9, bold=True, color=BRANCO)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = borda
ws.row_dimensions[6].height = 24

larguras = {"A": 24, "B": 16, "C": 15, "D": 52, "E": 8, "F": 13, "G": 13,
            "H": 10, "I": 7, "J": 15, "K": 15, "L": 34, "M": 5}
for col, w in larguras.items():
    ws.column_dimensions[col].width = w
ws.column_dimensions["M"].hidden = True  # qtd válida (uso interno das somas)

# SEM lista de dropdown na QTD: o autocompletar do Excel 365 "corrige" a
# digitação p/ o 1º item da lista que casa (5→50, 13→132, 2→20 — bug real
# visto pelo Leonardo). Campo livre com MENSAGEM flutuante explicando a
# regra; quem foge do padrão recebe o AVISO vermelho na coluna L.
dvs = {}
def dv_para(titulo, msg):
    chave = (titulo, msg)
    if chave in dvs:
        return dvs[chave]
    dv = DataValidation(type="whole", operator="between", formula1="0",
                        formula2="1000000", allow_blank=True,
                        showInputMessage=True, showErrorMessage=False,
                        promptTitle=titulo, prompt=msg)
    ws.add_data_validation(dv)
    dvs[chave] = dv
    return dv

def _padrao_caixa(lista, fallback):
    """Deriva da lista ORIGINAL da FOB: degrau modal (caixa), menor valor que
    entra na progressão (mínimo) e valores avulsos fora do degrau (ex. Akemi
    vendia 6 e 8 avulso e depois caixas de 12)."""
    pos = sorted({v for v in (lista or []) if v > 0})
    if not pos:
        p = fallback or 1
        return p, p, []
    if len(pos) == 1:
        return pos[0], pos[0], []
    difs = [b - a for a, b in zip(pos, pos[1:])]
    freq = {}
    for d in difs:
        freq[d] = freq.get(d, 0) + 1
    passo = max(freq, key=freq.get)
    na_progressao = [v for v in pos if v % passo == 0]
    tail_min = min(na_progressao) if na_progressao else pos[0]
    excecoes = [v for v in pos if v % passo != 0]
    return passo, tail_min, excecoes

r = PRIM
cat_ant = None
for p in produtos:
    promo = promos.get(p["sku"])
    em_promo = promo is not None
    fill_promo = PatternFill("solid", fgColor=LARANJA_CLARO)
    primeiro_da_cat = p["categoria"] != cat_ant
    cat_ant = p["categoria"]

    # categoria SEMPRE em negrito e escura ("está muito clarinho" — Leonardo 22/07)
    ws.cell(r, 1, p["categoria"]).font = Font(name="Inter", size=9, bold=True,
                                              color="FF0F6E56")
    ws.cell(r, 2, p["sku"]).font = f_normal
    cb = ws.cell(r, 3, p["barras"] or "")
    cb.font = f_peq
    cb.number_format = "@"
    ws.cell(r, 4, p["nome"]).font = Font(name="Inter", size=10, bold=em_promo)
    q = ws.cell(r, 5, 0)
    q.font = Font(name="Inter", size=10, bold=True)
    q.fill = PatternFill("solid", fgColor=AZUL_INPUT)
    q.alignment = Alignment(horizontal="center")
    q.protection = Protection(locked=False)
    regra = p.get("regra_qtd", "caixa")
    qmin = p.get("qtd_min") or 1
    passo, tail_min, excecoes = _padrao_caixa(p.get("qtd_lista"), p["multiplo"])
    # MÚLTIPLOS FIXOS (regra Leonardo 22/07): nestas categorias vale múltiplo
    # exato, sem avulsos nem mínimo diferente (Akemi PERDEU o "6/8 avulso")
    if regra == "caixa":
        _alvo = f"{p['categoria']} {p['nome']}".lower()
        for _chave, _mult in (("tira laço", 24), ("tira laco", 24),
                              ("autolimpante", 6), ("akemi", 6), ("guia", 6)):
            if _chave in _alvo:
                passo, tail_min, excecoes = _mult, _mult, []
                break
    if regra == "min":
        _titulo = f"Quantidade mínima: {qmin}"
        _msg = (f"A partir de {qmin} unidades vale QUALQUER quantidade "
                f"(ex.: {qmin + 1} pode). Digite o número direto.")
    elif excecoes:
        _exc = ", ".join(str(v) for v in excecoes)
        _titulo = f"Padrão: {_exc} avulso ou caixas de {passo}"
        _msg = (f"Aceita {_exc} avulso, ou caixas de {passo} a partir de "
                f"{tail_min} ({tail_min}, {tail_min + passo}...). Outros valores são bloqueados.")
    elif tail_min != passo:
        _titulo = f"Mínimo {tail_min}, em caixas de {passo}"
        _msg = (f"A partir de {tail_min} unidades, em passos de {passo} "
                f"({tail_min}, {tail_min + passo}, {tail_min + 2 * passo}...). Outros valores são bloqueados.")
    else:
        _titulo = f"Caixa fechada: múltiplos de {passo}"
        _msg = (f"Este item sai em caixa fechada de {passo} ({passo}, "
                f"{passo * 2}, {passo * 3}...). Outros valores são bloqueados.")
    if regra == "min":
        dv_para(_titulo, _msg).add(q.coordinate)
    else:
        # BLOQUEIO REAL de caixa fechada (Leonardo 22/07: "não pode fracionar").
        # Fórmula customizada bloqueia SEM lista (lista = autocompletar 5→50,
        # proibido). Colar valores fura a validação do Excel — por isso a
        # coluna AVISO continua existindo como rede de segurança.
        _pdv = [f"$E{r}=0"] + [f"$E{r}={v}" for v in excecoes]
        _pdv.append(f"AND($E{r}>={tail_min},MOD($E{r},{passo})=0)")
        _fdv = (f'=OR($E{r}="",AND(ISNUMBER($E{r}),$E{r}>=0,'
                f'$E{r}=INT($E{r}),OR({",".join(_pdv)})))')
        dvb = DataValidation(type="custom", formula1=_fdv, allow_blank=True,
                             showInputMessage=True, showErrorMessage=True,
                             promptTitle=_titulo, prompt=_msg,
                             errorTitle="Caixa fechada",
                             error=(f"{_titulo}. Use 0, {tail_min}, "
                                    f"{tail_min + passo}, {tail_min + 2 * passo}... "
                                    "(ou Delete p/ esvaziar)."))
        ws.add_data_validation(dvb)
        dvb.add(q.coordinate)

    ctab = ws.cell(r, 6, _preco(p["preco_tabela"]))
    ctab.number_format = '"R$" #,##0.00'
    ctab.font = Font(name="Inter", size=10, strike=em_promo,
                     color=CINZA_TXT if em_promo else "FF000000")
    if em_promo:
        cp = ws.cell(r, 7, _preco(promo["promo"]))
        cp.number_format = '"R$" #,##0.00'
        cp.font = Font(name="Inter", size=10, bold=True, color=LARANJA_TXT)
        ce = ws.cell(r, 8, promo["desc_pct"] / 100)
        ce.number_format = '0.0%'
        ce.font = Font(name="Inter", size=9, bold=True, color=LARANJA_TXT)
    ci = ws.cell(r, 9, p["ipi"])
    # 2 casas SEMPRE: '0%' arredondava 9,75% p/ "10%" e o distribuidor podia
    # achar que estava sendo cobrado a mais (reclamação real, 22/07)
    ci.number_format = '0.00%'
    ci.font = f_peq
    # PREÇO FINAL: promo SEMPRE aplicada (decisão Leonardo 2026-07-21 —
    # sem regra de pedido mínimo p/ ativar promoções no momento)
    cf = ws.cell(r, 10, f'=ROUND(IF($G{r}<>"",$G{r},$F{r})*(1+$I{r}),2)')
    cf.number_format = '"R$" #,##0.00'
    cf.font = Font(name="Inter", size=10, bold=True)
    # col M (oculta): quantidade VÁLIDA — só inteiro >= 0 entra nos totais
    # (auditoria: -4 subtraía R$ do pedido e 4,5 somava fração)
    cm = ws.cell(r, 13, f"=IF(ISNUMBER($E{r}),IF(AND($E{r}>=0,$E{r}=INT($E{r})),$E{r},0),0)")
    cm.font = f_peq
    ct = ws.cell(r, 11, f"=$M{r}*$J{r}")
    ct.number_format = '"R$" #,##0.00'
    ct.font = f_normal
    if regra == "min":
        _chk = f'IF($E{r}<{qmin},"⚠ ABAIXO DO MÍNIMO ({qmin} un.)","")'
    else:
        # padrão fiel à lista original: avulsos permitidos + progressão da caixa
        _partes = [f"$E{r}={v}" for v in excecoes]
        _partes.append(f"AND($E{r}>={tail_min},MOD($E{r},{passo})=0)")
        _valido = "OR(" + ",".join(_partes) + ")" if len(_partes) > 1 else _partes[0]
        if excecoes:
            _rot = f"⚠ FORA DO PADRÃO ({', '.join(str(v) for v in excecoes)} ou caixas de {passo})"
        elif tail_min != passo:
            _rot = f"⚠ FORA DO PADRÃO (mín. {tail_min}, caixas de {passo})"
        else:
            _rot = f"⚠ NÃO FECHA CAIXA (múltiplos de {passo})"
        _chk = f'IF({_valido},"","{_rot}")'
    # $E="" primeiro: Delete na célula NÃO é erro (vazio = zero, sem aviso)
    av = ws.cell(r, 12, f'=IF($E{r}="","",IF(NOT(ISNUMBER($E{r})),"⚠ QUANTIDADE INVÁLIDA",'
                        f'IF(OR($E{r}<0,$E{r}<>INT($E{r})),"⚠ QUANTIDADE INVÁLIDA",'
                        f'IF($E{r}=0,"",{_chk}))))')
    av.font = Font(name="Inter", size=9, bold=True, color="FFC00000")
    for col in range(1, 13):
        cel = ws.cell(r, col)
        cel.border = borda
        if em_promo and col != 5:
            cel.fill = fill_promo
    ws.cell(r, 5).fill = PatternFill("solid", fgColor=AZUL_INPUT)
    r += 1

ws.merge_cells(f"A{r}:D{r}")
ws.cell(r, 1, "TOTAL DO PEDIDO (JÁ C/ IPI)").font = Font(name="Inter", size=11, bold=True)
ws.cell(r, 1).alignment = Alignment(horizontal="right")
ws.cell(r, 5, f"=SUM({rng('M')})").font = Font(name="Inter", size=11, bold=True)
tt = ws.cell(r, 11, f"=SUM({rng('K')})")
tt.number_format = '"R$" #,##0.00'
tt.font = Font(name="Inter", size=12, bold=True, color="FF085041")
for col in range(1, 13):
    ws.cell(r, col).fill = PatternFill("solid", fgColor=TEAL_CLARO)

ws.auto_filter.ref = f"A6:L{ULT}"
ws.freeze_panes = "A7"
ws.protection.sheet = True
ws.protection.password = SENHA
ws.protection.autoFilter = False
ws.protection.sort = False
ws.protection.formatColumns = False

# ============================ ABA SEU LUCRO ============================
wl = wb.create_sheet("SEU LUCRO")
wl.sheet_view.showGridLines = False
wl.merge_cells("A1:H1")
c = wl["A1"]
c.value = "QUANTO VOCÊ GANHA REVENDENDO PROPETZ  —  painel do seu lucro"
c.font = Font(name="Inter", size=14, bold=True, color=BRANCO)
c.fill = PatternFill("solid", fgColor=TEAL_ESCURO)
c.alignment = Alignment(horizontal="center", vertical="center")
wl.row_dimensions[1].height = 28
wl.merge_cells("A2:H2")
wl["A2"] = ("Revenda sugerida = preço do site Propetz −5% (você vende MAIS BARATO que o nosso "
            "site). Mínimo permitido = site −15% (piso p/ não desvalorizar a marca). Preencha o "
            "PEDIDO e este painel calcula o seu lucro sozinho.")
wl["A2"].font = f_peq

# ---- pré-cálculo por linha (alimenta painel + top oportunidades) ----
sugeridos = MODELO["sugeridos"]  # fallback p/ SKU sem preço de site no mestre
linhas_lucro = []
for i, p in enumerate(produtos):
    # REGRA 22/07 (Leonardo): base = preço do SITE Propetz (col AG do mestre)
    _site = p.get("site_propetz")
    if _site:
        sug = {"sugerido": round(_site * 0.95, 2), "minimo": round(_site * 0.85, 2)}
    else:
        sug = sugeridos.get(p["sku"])
    promo = promos.get(p["sku"])
    preco_vig = _preco(promo["promo"] if promo else p["preco_tabela"])
    ganho = (sug["sugerido"] / preco_vig - 1) if (sug and preco_vig) else None
    linhas_lucro.append({"p": p, "sug": sug, "promo": promo, "vig": preco_vig,
                         "ganho": ganho, "linha_pedido": PRIM + i})

CAB_L = 10                    # linha do cabeçalho da tabela
D0 = CAB_L + 1
DF = D0 + len(linhas_lucro) - 1

# ---- PAINEL AO VIVO do pedido (insight pedido do Leonardo 22/07) ----
paineis = [
    ("A", "B", "LUCRO PROJETADO NA REVENDA", f"=SUM($H${D0}:$H${DF})", '"R$" #,##0.00'),
    ("C", "D", "MARGEM MÉDIA DA REVENDA",
     f'=IF(SUMPRODUCT($G${D0}:$G${DF},$D${D0}:$D${DF})=0,"",'
     f'SUM($H${D0}:$H${DF})/SUMPRODUCT($G${D0}:$G${DF},$D${D0}:$D${DF}))', '0.0%'),
    ("E", "F", "ITENS NO SEU PEDIDO", f"=SUMPRODUCT(--($G${D0}:$G${DF}>0))", "0"),
]
for c_i, c_f, rotulo, formula, fmt in paineis:
    wl.merge_cells(f"{c_i}3:{c_f}3")
    lab = wl[f"{c_i}3"]
    lab.value = rotulo
    lab.font = Font(name="Inter", size=8, bold=True, color=CINZA_TXT)
    lab.alignment = Alignment(horizontal="center")
    wl.merge_cells(f"{c_i}4:{c_f}4")
    val = wl[f"{c_i}4"]
    val.value = formula
    val.number_format = fmt
    val.font = Font(name="Inter", size=13, bold=True, color="FF085041")
    val.alignment = Alignment(horizontal="center")

# ---- TOP OPORTUNIDADES com status ao vivo ----
wl.merge_cells("A5:H5")
wl["A5"] = "🔥 MAIORES GANHOS DA TABELA — complete o seu pedido com eles:"
wl["A5"].font = Font(name="Inter", size=10, bold=True, color=LARANJA_TXT)
_top = sorted((l for l in linhas_lucro if l["ganho"] is not None),
              key=lambda x: -x["ganho"])[:4]
for j, t in enumerate(_top):
    lr = 6 + j
    wl.merge_cells(f"A{lr}:H{lr}")
    _txt = (f"{t['p']['sku']} · {t['p']['nome'][:45]} — ganho de "
            f"+{t['ganho'] * 100:.0f}% na revenda").replace('"', "”")
    wl[f"A{lr}"] = (f'="{_txt}"&IF(PEDIDO!$M${t["linha_pedido"]}=0,'
                    f'"   ←  ainda FORA do seu pedido","   ✔ já no seu pedido")')
    wl[f"A{lr}"].font = f_normal

cab2 = ["CÓDIGO", "PRODUTO", "SEU PREÇO\n(c/ promo)", "REVENDA SUGERIDA\n(site −5%)",
        "MÍNIMO PERMITIDO\n(site −15%)", "SEU\nGANHO", "NO SEU PEDIDO\n(unidades)",
        "GANHO PROJETADO\n(R$)"]
for i, t in enumerate(cab2, 1):
    c = wl.cell(CAB_L, i, t)
    c.font = Font(name="Inter", size=9, bold=True, color=BRANCO)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = borda
wl.row_dimensions[CAB_L].height = 40
# colunas largas o bastante p/ ler os cabeçalhos (reclamação 22/07)
for col, w in {"A": 15, "B": 42, "C": 14, "D": 19, "E": 19, "F": 10,
               "G": 15, "H": 18}.items():
    wl.column_dimensions[col].width = w

r = D0
for l in linhas_lucro:
    p, sug, promo = l["p"], l["sug"], l["promo"]
    wl.cell(r, 1, p["sku"]).font = f_normal
    wl.cell(r, 2, p["nome"]).font = f_normal
    cpv = wl.cell(r, 3, round(l["vig"], 2))
    cpv.number_format = '"R$" #,##0.00'
    cpv.font = Font(name="Inter", size=10, bold=bool(promo),
                    color=LARANJA_TXT if promo else "FF000000")
    cs = wl.cell(r, 4, sug["sugerido"] if sug else None)
    cs.number_format = '"R$" #,##0.00'
    cs.font = f_normal
    cm = wl.cell(r, 5, sug["minimo"] if sug else None)
    cm.number_format = '"R$" #,##0.00'
    cm.font = f_peq
    cg = wl.cell(r, 6, round(l["ganho"], 4) if l["ganho"] is not None else None)
    cg.number_format = '"+"0%'
    cg.font = Font(name="Inter", size=10, bold=True, color="FF0F6E56")
    lp = l["linha_pedido"]
    cq = wl.cell(r, 7, f"=PEDIDO!$M${lp}")
    cq.number_format = '0'
    cq.font = Font(name="Inter", size=10, bold=True)
    cq.alignment = Alignment(horizontal="center")
    ch = wl.cell(r, 8, f'=IF(OR($D{r}="",$G{r}=0),"",'
                       f'ROUND($G{r}*($D{r}-PEDIDO!$J${lp}),2))')
    ch.number_format = '"R$" #,##0.00'
    ch.font = Font(name="Inter", size=10, bold=True, color="FF085041")
    for col in range(1, 9):
        wl.cell(r, col).border = borda
        if promo:
            wl.cell(r, col).fill = PatternFill("solid", fgColor=LARANJA_CLARO)
    r += 1
wl.auto_filter.ref = f"A{CAB_L}:H{DF}"
wl.freeze_panes = f"A{D0}"
wl.protection.sheet = True
wl.protection.password = SENHA
wl.protection.autoFilter = False

# ==================== ABA CADASTRO E CONDIÇÕES ====================
wc = wb.create_sheet("CADASTRO E CONDIÇÕES")
wc.sheet_view.showGridLines = False
wc.merge_cells("A1:D1")
c = wc["A1"]
c.value = "CADASTRO DO CLIENTE  +  CONDIÇÕES COMERCIAIS"
c.font = Font(name="Inter", size=14, bold=True, color=BRANCO)
c.fill = PatternFill("solid", fgColor=TEAL_ESCURO)
c.alignment = Alignment(horizontal="center", vertical="center")
wc.row_dimensions[1].height = 28
campos = ["Razão Social", "CNPJ", "I.E. e I.M.", "Endereço", "Bairro", "Cidade e Estado",
          "CEP", "Telefone / Celular", "E-mail", "Instagram", "Horário de entrega",
          "Data de aniversário", "Vendedor(a) Propetz"]
r = 3
for campo in campos:
    wc.cell(r, 1, campo + ":").font = Font(name="Inter", size=10, bold=True)
    # sem merge: o save do openpyxl propagaria o destrave da célula-mestre
    # para o range inteiro — célula única larga dá o mesmo visual
    e = wc.cell(r, 2, "")
    e.fill = PatternFill("solid", fgColor=AZUL_INPUT)
    e.border = borda
    e.protection = Protection(locked=False)
    r += 1
r += 1
wc.cell(r, 1, "CONDIÇÕES DE PAGAMENTO").font = Font(name="Inter", size=11, bold=True, color="FF085041")
r += 1
condicoes = ["À vista: PIX ou transferência — 5% de desconto sobre o total do pedido.",
             "A prazo: boleto 21/35/49/63/77/91 dias (condições sujeitas a análise financeira).",
             "Preços PROMOÇÃO (laranja): já aplicados automaticamente no seu pedido.",
             ("Frete: JÁ INCLUÍDO nos preços (tabela CIF)." if MODALIDADE == "CIF"
              else "Frete: a cotar no fechamento do pedido."),
             "IPI destacado por item na aba PEDIDO; os TOTAIS do pedido já incluem o IPI."]
for t in condicoes:
    wc.merge_cells(f"A{r}:D{r}")
    wc.cell(r, 1, "•  " + t).font = f_normal
    r += 1
r += 1
wc.cell(r, 1, "QUANTIDADE MÍNIMA POR MODELO").font = Font(name="Inter", size=11, bold=True, color="FF085041")
r += 1
for modelo, qtd in MODELO["minimos"]:
    if "avental" in str(modelo).lower() or "jaleco" in str(modelo).lower():
        continue  # não vendemos mais (Leonardo 22/07)
    wc.merge_cells(f"A{r}:C{r}")
    wc.cell(r, 1, modelo).font = f_normal
    wc.cell(r, 1).border = borda
    q = wc.cell(r, 4, qtd)
    q.font = Font(name="Inter", size=10, bold=True)
    q.alignment = Alignment(horizontal="center")
    q.border = borda
    r += 1
wc.merge_cells(f"A{r}:D{r}")
wc.cell(r, 1, "Exceções pelo padrão de embalagem: Tesouras Laser = mín. 12 · "
              "Tesoura Super Curva 7,5\" Super Premium = mín. 4 · "
              "Pentes Akemi = múltiplos de 6 · Tira-Laço = múltiplos de 24").font = f_peq
r += 1
for col, w in {"A": 34, "B": 48, "C": 10, "D": 14}.items():
    wc.column_dimensions[col].width = w
wc.protection.sheet = True
wc.protection.password = SENHA

wb.active = 0
props = wb.properties
props.title = "Pedido Propetz Distribuição — Agosto 2026"
props.creator = "Propetz / Grupo Daros"
try:
    wb.save(SAIDA)
except PermissionError:
    # arquivo aberto no Excel: salva ao lado p/ não perder o trabalho
    SAIDA = SAIDA.replace(".xlsx", " (NOVA - feche e apague a anterior).xlsx")
    wb.save(SAIDA)
    print("ATENÇÃO: o arquivo estava ABERTO no Excel — salvei como versão nova ao lado.")

print(f"OK: {SAIDA}")
print(f"produtos no pedido: {len(produtos)} | promos destacadas: {len(promos)} | excluidos: {len(excluidos)}")
print("excluidos:", excluidos)
