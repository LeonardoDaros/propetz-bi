import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
import yaml
import os
import hashlib
import hmac
import html
import math
import secrets
import json
import base64
import threading
import requests
import uuid
from zoneinfo import ZoneInfo
import agenda_comercial as agenda
import ui_propetz as ui
from collections import defaultdict
from datetime import datetime, timedelta, date

# ============================================================
# CONFIG
# ============================================================
st.set_page_config(
    page_title="Propetz BI",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="auto"
)

# ============================================================
# CUSTOM CSS
# ============================================================
ui.apply_theme()

# ============================================================
# USER DATABASE (stored in YAML - editable)
# ============================================================
USERS_FILE = os.path.join(os.path.dirname(__file__), "users.yaml")

def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    # BREAK-GLASS: só quando o users.yaml some E não dá p/ restaurar do state
    # (não acontece em operação normal — users.yaml vem no clone + branch state).
    # NENHUMA senha em texto puro no código (as antigas nome2026 vazaram). A
    # senha do admin de emergência vem do secret BREAKGLASS_PASS; sem ele, o
    # fallback é INUTILIZÁVEL (senha aleatória) — o app não trava, só exige o
    # users.yaml real. Configure BREAKGLASS_PASS nos Secrets se quiser a rede.
    try:
        bg = st.secrets.get("BREAKGLASS_PASS", "")
    except Exception:
        bg = ""
    senha = bg if bg else secrets.token_hex(24)  # sem secret → login impossível
    return {"users": {"leonardo": {"name": "Leonardo Daros",
                                   "password": hash_password(senha),
                                   "role": "admin", "vendor_filter": None}}}

# Hash de senha: scrypt (stdlib) com salt por usuário + key-stretching.
# Formato: "scrypt$<n>$<r>$<p>$<salt_hex>$<hash_hex>". SHA-256 puro NUNCA mais
# é gerado — só reconhecido para MIGRAR hashes antigos no próximo login.
_SCRYPT_N, _SCRYPT_R, _SCRYPT_P = 16384, 8, 1

def hash_password(pwd):
    salt = secrets.token_bytes(16)
    dk = hashlib.scrypt(pwd.encode(), salt=salt, n=_SCRYPT_N, r=_SCRYPT_R,
                        p=_SCRYPT_P, dklen=32, maxmem=64 * 1024 * 1024)
    return f"scrypt${_SCRYPT_N}${_SCRYPT_R}${_SCRYPT_P}${salt.hex()}${dk.hex()}"

def _verify_password(pwd, stored):
    """Confere a senha em TEMPO CONSTANTE (hmac.compare_digest). Aceita o novo
    formato scrypt e o legado sha256 (só para migração). NUNCA levanta exceção
    (hash corrompido/None/não-ASCII → (False, False)). Retorna (ok, é_legado)."""
    if pwd is None:
        return False, False
    stored = str(stored or "")
    if stored.startswith("scrypt$"):
        try:
            _, n, r, p, salt_hex, h = stored.split("$")
            dk = hashlib.scrypt(str(pwd).encode(), salt=bytes.fromhex(salt_hex),
                                n=int(n), r=int(r), p=int(p), dklen=len(h) // 2,
                                maxmem=64 * 1024 * 1024)
            return hmac.compare_digest(dk.hex(), h), False
        except Exception:
            return False, False
    # legado: sha256 puro (64 hex) — reconhecido só para re-hashear no login
    try:
        legacy = hashlib.sha256(str(pwd).encode()).hexdigest()
        return hmac.compare_digest(legacy, stored), True
    except Exception:
        return False, False

def save_users(users_data):
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        yaml.dump(users_data, f, default_flow_style=False, allow_unicode=True)
    _push_state_file("users.yaml")

def verify_login(username, password):
    users = load_users()
    # Normaliza: celular capitaliza a 1ª letra e pode incluir espaço no autocomplete
    key = str(username).strip().lower()
    user = users["users"].get(key)
    if not user:
        return None
    ok, is_legacy = _verify_password(password, user.get("password", ""))
    if not ok:
        return None
    if is_legacy:
        # MIGRAÇÃO transparente: re-hasheia com scrypt no 1º login bem-sucedido
        try:
            users["users"][key]["password"] = hash_password(password)
            save_users(users)
        except Exception:
            pass
    return user

# ============================================================
# BRUTE FORCE PROTECTION
# ============================================================
LOGIN_ATTEMPTS_FILE = os.path.join(os.path.dirname(__file__), "login_attempts.json")
_BF_MAX = 5          # falhas por USUÁRIO antes de bloquear
_BF_MAX_IP = 30      # falhas por IP (mais alto: escritório inteiro é 1 IP via NAT)
_BF_JANELA = 300     # segundos de bloqueio / janela de contagem
_BF_LOCK = threading.Lock()
_bf_ultimo_push = [0.0]   # debounce da publicação no state (anti-DoS)

def _client_ip():
    """IP do cliente atrás do proxy do Streamlit Cloud, se disponível. O
    X-Forwarded-For é FORJÁVEL pelo cliente, então o limite por IP é só um
    reforço (limiar alto) — a proteção real é a contagem por usuário."""
    try:
        xff = st.context.headers.get("X-Forwarded-For", "")
        return xff.split(",")[0].strip() or None
    except Exception:
        return None

def _load_login_attempts():
    """Lê o arquivo LOCAL (caminho quente do login, SEM tocar o GitHub — evita
    que uma enxurrada de logins falhos vire escrita/leitura remota = DoS). O
    boot restaura este arquivo do branch state (_sync_state_from_github), então
    o bloqueio sobrevive a reinícios; a publicação é assíncrona e com debounce."""
    if os.path.exists(LOGIN_ATTEMPTS_FILE):
        try:
            with open(LOGIN_ATTEMPTS_FILE, "r") as f:
                d = json.load(f)
            return d if isinstance(d, dict) else {}
        except Exception:
            return {}
    return {}

def _save_login_attempts_local(d):
    try:
        with open(LOGIN_ATTEMPTS_FILE, "w") as f:
            json.dump(d, f)
    except Exception:
        pass

def _persist_attempts_debounced():
    """Publica login_attempts no state NO MÁXIMO 1x/60s, assíncrono — sob
    ataque de brute force isso limita a 1 escrita/min em vez de milhares."""
    agora = datetime.now().timestamp()
    if agora - _bf_ultimo_push[0] < 60:
        return
    _bf_ultimo_push[0] = agora
    _push_state_file("login_attempts.json")  # já roda em thread, best-effort

def _limite_da_chave(key):
    return _BF_MAX_IP if key.startswith("ip:") else _BF_MAX

def _bf_keys(username, ip):
    """Chaves de contagem: uma por usuário e uma por IP (limita adivinhação
    distribuída), com limiares diferentes (ver _limite_da_chave)."""
    ks = [f"u:{str(username).lower().strip()}"]
    if ip:
        ks.append(f"ip:{ip}")
    return ks

def check_rate_limit(username, ip=None):
    """(bloqueado, segundos_restantes). Bloqueia por USUÁRIO (>=5) ou por IP
    (>=30). Leitura LOCAL — sem rede no caminho do login."""
    if ip is None:
        ip = _client_ip()
    attempts = _load_login_attempts()
    now = datetime.now().timestamp()
    pior = 0
    for key in _bf_keys(username, ip):
        info = attempts.get(key)
        if not info or now - info.get("last_fail", 0) > _BF_JANELA:
            continue
        if info.get("count", 0) >= _limite_da_chave(key):
            pior = max(pior, int(_BF_JANELA - (now - info.get("last_fail", 0))))
    return (pior > 0), max(pior, 0)

def record_failed_attempt(username, ip=None):
    if ip is None:
        ip = _client_ip()
    now = datetime.now().timestamp()
    keys = _bf_keys(username, ip)
    with _BF_LOCK:
        d = _load_login_attempts()
        for key in keys:
            info = d.get(key)
            if not info or now - info.get("last_fail", 0) > _BF_JANELA:
                d[key] = {"count": 1, "last_fail": now}
            else:
                d[key] = {"count": info.get("count", 0) + 1, "last_fail": now}
        # higiene: descarta entradas velhas p/ o arquivo não crescer
        d = {k: v for k, v in d.items()
             if now - v.get("last_fail", 0) <= _BF_JANELA * 4}
        _save_login_attempts_local(d)
    _persist_attempts_debounced()

def clear_failed_attempts(username, ip=None):
    if ip is None:
        ip = _client_ip()
    keys = set(_bf_keys(username, ip))
    with _BF_LOCK:
        d = _load_login_attempts()
        for k in keys:
            d.pop(k, None)
        _save_login_attempts_local(d)
    _persist_attempts_debounced()

# ============================================================
# ACCESS LOG — registra QUEM entrou, QUANDO e quais PÁGINAS abriu (frequência de uso;
# não mede duração de sessão). Mesma robustez das inativações: lê do GitHub (fonte da
# verdade) e grava por APPEND ATÔMICO — um reinício ruim ou acessos simultâneos não
# apagam o histórico. A gravação roda em segundo plano para não travar a navegação.
# ============================================================
ACCESS_LOG_FILE = os.path.join(os.path.dirname(__file__), "access_log.json")

def _load_access_log():
    """Lê o log do GitHub (fonte da verdade), com fallback no arquivo local."""
    data = _read_state_json("access_log.json", ACCESS_LOG_FILE, [])
    return data if isinstance(data, list) else []

def _append_access_log_entry(entry):
    """Acrescenta UM evento ao log de forma robusta e NÃO bloqueante: numa thread,
    faz read-modify-write atômico no GitHub (lê o log atual, acrescenta, corta nos
    últimos 5000). Sem token: grava só no local (efêmero, melhor esforço)."""
    tok = _gh_token()
    if not tok:
        log = _load_access_log()
        log = (log if isinstance(log, list) else [])[-4999:] + [entry]
        _write_local_json(ACCESS_LOG_FILE, log)
        _STATE_RAW_CACHE.pop("access_log.json", None)
        return

    def _do():
        _gh_mutate_json("access_log.json", ACCESS_LOG_FILE,
            lambda d: ((d if isinstance(d, list) else []) + [entry])[-5000:],
            [], token=tok)

    threading.Thread(target=_do, daemon=True).start()

def log_access(username, user_name, action="login"):
    """Log a user access event."""
    now = datetime.now()
    _append_access_log_entry({
        "user": username,
        "name": user_name,
        "action": action,
        "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
        "date": now.strftime("%Y-%m-%d"),
        "time": now.strftime("%H:%M:%S"),
    })

def log_page_view(username, page_name):
    """Log a page view event."""
    now = datetime.now()
    _append_access_log_entry({
        "user": username,
        "name": st.session_state.get("user_name", username),
        "action": "page_view",
        "page": page_name,
        "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
        "date": now.strftime("%Y-%m-%d"),
        "time": now.strftime("%H:%M:%S"),
    })

def has_full_data_access():
    """Returns True for admin and diretor roles (see all data, no vendor filter)."""
    return st.session_state.get('role') in ('admin', 'diretor')

def _vendor_options(df_clients):
    """Carteiras válidas vêm da base já normalizada pelo carregador."""
    return sorted({str(v).strip() for v in df_clients['vendor'].dropna() if str(v).strip()})

def _access_configuration_error(role, vendor_filter, vendor_options=None):
    """Falha fechada: papel desconhecido ou vendedor sem carteira não vê dados."""
    if role not in ('admin', 'diretor', 'vendedor', 'garantia', 'garantia_master'):
        return "Seu perfil de acesso precisa ser revisado pelo administrador."
    if role == 'vendedor':
        if not isinstance(vendor_filter, str) or not vendor_filter.strip():
            return "Sua carteira não está configurada. Solicite o ajuste ao administrador."
        if vendor_options is not None and vendor_filter.strip() not in vendor_options:
            return "Sua carteira não foi encontrada na base atual. Solicite o ajuste ao administrador."
    return None

def _clients_for_access(df_clients, role, vendor_filter):
    error = _access_configuration_error(role, vendor_filter, _vendor_options(df_clients))
    if error:
        raise ValueError(error)
    if role == 'vendedor':
        return df_clients[df_clients['vendor'].fillna('').astype(str).str.strip() == vendor_filter.strip()].copy()
    return df_clients.copy()

def _refresh_session_access():
    """Revalida o cadastro local a cada interação, inclusive após troca de carteira."""
    try:
        user = load_users()['users'].get(st.session_state.get('username'))
        if not isinstance(user, dict):
            return "Seu cadastro não está mais disponível. Contate o administrador."
        st.session_state['role'] = user.get('role')
        st.session_state['vendor_filter'] = user.get('vendor_filter')
        st.session_state['user_name'] = user.get('name') or st.session_state.get('username', '')
    except (OSError, yaml.YAMLError, KeyError, TypeError, AttributeError):
        return "Não foi possível validar seu cadastro. Tente novamente ou contate o administrador."
    return _access_configuration_error(user.get('role'), user.get('vendor_filter'))

def _show_access_denied(message):
    st.error(message)
    if st.button("Sair e usar outra conta", key="access_denied_logout"):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        _strip_stale_auth_params()
        st.rerun()

def _new_user_error(users, username, name, password, role, vendor_filter, vendor_options):
    if not username.strip() or not name.strip() or not password:
        return "Preencha usuário, nome e senha."
    if username.strip().lower() in {str(key).strip().lower() for key in users['users']}:
        return "Esse usuário já existe. Use as opções de alteração abaixo."
    if len(password) < 12:
        return "Use uma senha com pelo menos 12 caracteres."
    if role not in ('vendedor', 'admin'):
        return "Selecione um papel válido."
    return _access_configuration_error(role, vendor_filter, vendor_options)

def can_approve_inactivations():
    """Somente admin aprova/inativa/reativa direto. Diretor e vendedor SOLICITAM
    (a diretora conhece a carteira de todos, então sugere; o admin decide)."""
    return st.session_state.get('role') == 'admin'

# ============================================================
# INACTIVE CLIENTS DATABASE (stored in JSON - persists across sessions)
# ============================================================
INACTIVE_FILE = os.path.join(os.path.dirname(__file__), "inactive_clients.json")

# Memo por script-run: lê o GitHub no máximo 1x por arquivo por rerun (esvaziado
# no início do main() e após cada gravação). Evita N chamadas de rede por render.
_STATE_RAW_CACHE = {}

def _read_state_raw(remote_name, local_path):
    """Conteúdo bruto do arquivo de estado, tendo o branch 'state' do GitHub como
    FONTE DA VERDADE (quando há token). Cai para o arquivo local só se não houver
    token ou a leitura remota falhar. É isto que impede um boot ruim de levar a uma
    gravação destrutiva: sempre lemos o estado bom do GitHub antes de modificar."""
    if remote_name in _STATE_RAW_CACHE:
        return _STATE_RAW_CACHE[remote_name]
    raw = None
    tok = _gh_token()
    if tok:
        try:
            content, _ = _gh_get_file(remote_name, _GH_STATE_BRANCH, token=tok)
            if content is not None:
                raw = content.decode('utf-8')
        except Exception:
            raw = None
    if raw is None and os.path.exists(local_path):
        try:
            with open(local_path, 'r', encoding='utf-8') as f:
                raw = f.read()
        except Exception:
            raw = None
    _STATE_RAW_CACHE[remote_name] = raw
    return raw

def _read_state_json(remote_name, local_path, default):
    raw = _read_state_raw(remote_name, local_path)
    if raw:
        try:
            return json.loads(raw)  # objeto novo a cada chamada (callers podem mutar)
        except Exception:
            pass
    return default

def load_inactive_clients():
    """Conjunto de IDs inativos — lido do GitHub (fonte da verdade) com fallback local."""
    data = _read_state_json("inactive_clients.json", INACTIVE_FILE, {"inactive_ids": []})
    try:
        return set(data.get("inactive_ids", []))
    except Exception:
        return set()

def save_inactive_clients(inactive_set):
    """Grava o conjunto de IDs inativos (local + GitHub SÍNCRONO, com confirmação)."""
    with open(INACTIVE_FILE, 'w', encoding='utf-8') as f:
        json.dump({"inactive_ids": sorted(list(inactive_set))}, f, ensure_ascii=False, indent=2)
    _STATE_RAW_CACHE.pop("inactive_clients.json", None)
    _push_state_file("inactive_clients.json", sync=True)
    return None

# ============================================================
# SOLICITAÇÕES DE INATIVAÇÃO (vendedor solicita → admin aprova)
# ============================================================
INACTIVE_REQUESTS_FILE = os.path.join(os.path.dirname(__file__), "inactive_requests.json")

def load_inactive_requests():
    """Solicitações/decisões de inativação — GitHub como fonte da verdade, fallback local."""
    data = _read_state_json("inactive_requests.json", INACTIVE_REQUESTS_FILE, {"requests": []})
    try:
        return data.get("requests", [])
    except Exception:
        return []

def save_inactive_requests(reqs):
    with open(INACTIVE_REQUESTS_FILE, 'w', encoding='utf-8') as f:
        json.dump({"requests": reqs[-500:]}, f, ensure_ascii=False, indent=1)
    _STATE_RAW_CACHE.pop("inactive_requests.json", None)
    _push_state_file("inactive_requests.json", sync=True)

# Motivos padronizados (geram um banco analisável; "Outro" + observação cobre o resto)
MOTIVOS_INATIVACAO = [
    "Fechou / encerrou atividades",
    "Não existe mais / cadastro duplicado",
    "Trocou por concorrente / outro fornecedor",
    "Inadimplente",
    "Parou de comprar / sem demanda",
    "Mudou de canal (virou varejo/consumidor)",
    "Outro",
]

def inactivate_clients(cids):
    """Adiciona IDs ao conjunto de inativos de forma ATÔMICA (não apaga os demais,
    mesmo com inativações simultâneas). Retorna True se persistiu de fato."""
    cset = {str(c).strip() for c in cids if str(c).strip()}
    if not cset:
        return True
    _, ok = _gh_mutate_json("inactive_clients.json", INACTIVE_FILE,
        lambda d: {"inactive_ids": sorted(set(d.get("inactive_ids", [])) | cset)},
        {"inactive_ids": []})
    return ok

def reactivate_clients(cids):
    """Remove IDs do conjunto de inativos de forma ATÔMICA. Retorna True se persistiu."""
    cset = {str(c).strip() for c in cids if str(c).strip()}
    if not cset:
        return True
    _, ok = _gh_mutate_json("inactive_clients.json", INACTIVE_FILE,
        lambda d: {"inactive_ids": sorted(set(d.get("inactive_ids", [])) - cset)},
        {"inactive_ids": []})
    return ok

def add_inactivation_request(client_id, client_name, vendor, motivo="", observacao="", direct_approve=False):
    """Registra a inativação com motivo (gravação ATÔMICA). Vendedor/diretora →
    'pendente' (admin aprova). Admin (direct_approve) → 'aprovado' + inativa na hora.
    Retorna True se a ação foi aplicada E persistida; False se duplicada ou se a
    gravação no GitHub falhou (aí o chamador avisa o usuário)."""
    cid = str(client_id).strip()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    user = st.session_state.get("username", "")
    uname = st.session_state.get("user_name", "")
    flags = {}

    def apply(d):
        flags["dup"] = False  # reseta a CADA tentativa (apply roda de novo em retry/409)
        reqs = list(d.get("requests", []))
        existing = next((r for r in reqs
                         if r.get("client_id") == cid and r.get("status") == "pendente"), None)
        if not direct_approve and existing:
            flags["dup"] = True
            return {"requests": reqs[-500:]}  # vendedor: já há pendente, nada muda
        if direct_approve and existing:
            existing["status"] = "aprovado"
            existing["decidido_em"] = now
            existing["decidido_por"] = uname
            return {"requests": reqs[-500:]}
        rec = {
            "client_id": cid, "client_name": str(client_name), "vendor": str(vendor or ""),
            "motivo": str(motivo or "").strip(), "observacao": str(observacao or "").strip(),
            "requested_by": user, "requested_by_name": uname, "date": now,
            "status": "aprovado" if direct_approve else "pendente",
        }
        if direct_approve:
            rec["decidido_em"] = now
            rec["decidido_por"] = uname
        reqs.append(rec)
        return {"requests": reqs[-500:]}

    ok_client = inactivate_clients([cid]) if direct_approve else True
    _, ok_req = _gh_mutate_json("inactive_requests.json", INACTIVE_REQUESTS_FILE, apply, {"requests": []})
    if flags.get("dup"):
        return False  # já existia pendente (não é falha de gravação)
    return ok_req and ok_client

def decide_inactivation_request(client_id, approve, decided_by):
    """Aprova/rejeita uma solicitação pendente de forma ATÔMICA. Se aprova, inativa
    o cliente (também atômico). Retorna True se persistiu."""
    cid = str(client_id).strip()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    status = "aprovado" if approve else "rejeitado"

    def apply(d):
        reqs = list(d.get("requests", []))
        for r in reqs:
            if r.get("client_id") == cid and r.get("status") == "pendente":
                r["status"] = status
                r["decidido_em"] = now
                r["decidido_por"] = decided_by
                break
        return {"requests": reqs[-500:]}

    ok_client = inactivate_clients([cid]) if approve else True
    _, ok_req = _gh_mutate_json("inactive_requests.json", INACTIVE_REQUESTS_FILE, apply, {"requests": []})
    return ok_req and ok_client

def pending_inactivation_requests():
    return [r for r in load_inactive_requests() if r.get("status") == "pendente"]

def _fmt_motivo(r):
    """Combina motivo + observação de um registro num texto curto (ou '')."""
    mot = (r.get('motivo') or '').strip()
    obs = (r.get('observacao') or '').strip()
    if mot and obs:
        return f"{mot} — {obs}"
    return mot or obs

# ============================================================
# GARANTIAS / ASSISTÊNCIA TÉCNICA
# Registro estruturado do que a NF não conta: defeito, causa, peças, custo.
# Persistência ATÔMICA no GitHub (mesmo motor das inativações).
# ============================================================
GARANTIAS_FILE = os.path.join(os.path.dirname(__file__), "garantias.json")

STATUS_GARANTIA = ["Aguardando chegada", "Em bancada", "Aguardando peça",
                   "Confirmado — aguardando R$ frete", "Concluída", "Cancelada"]
STATUS_ATIVOS = ["Aguardando chegada", "Em bancada", "Aguardando peça",
                 "Confirmado — aguardando R$ frete"]
STATUS_FINALIZADOS = ["Concluída", "Cancelada"]
# nomes antigos de registros já gravados -> nomes novos (migração transparente)
_STATUS_LEGADO = {"Aberta": "Aguardando chegada", "Devolvida ao cliente": "Concluída"}
CANAIS_GARANTIA = ["Distribuição", "Varejo", "Feira", "Outro"]
EMPRESAS_NF = ["Matriz", "Filial", "Filial Foz", "TradeCorp"]
DEFEITOS_GARANTIA = ["Não liga", "Motor", "Bateria / não carrega", "Carregador / fonte", "Lâmina / corte",
                     "Afiação tesoura", "Botão / interruptor", "Carcaça quebrada", "Ruído / vibração",
                     "Esquenta demais", "Display / luz", "Dano de transporte", "Outro"]
CAUSAS_GARANTIA = ["Defeito de fabricação", "Mau uso do cliente", "Desgaste natural",
                   "Dano de transporte", "Instalação/voltagem errada", "Sem defeito constatado", "Outra"]
PRIORIDADES_GARANTIA = ["Normal", "Alta", "Urgente"]
_PRIO_ICONE = {"Alta": "🟠 ", "Urgente": "🔴 "}
RESULTADOS_GARANTIA = ["Consertada", "Trocada por produto novo", "Reembolso", "Recusada (fora de garantia)",
                       "Devolvida sem conserto"]

def can_manage_garantias():
    return st.session_state.get('role') in ('admin', 'diretor', 'garantia', 'garantia_master')

def can_edit_garantia_fechada():
    """Garantia finalizada (Concluída/Devolvida/Cancelada) só pode ser reaberta,
    corrigida ou cancelada pelo MASTER da garantia (Jackson) ou pelo admin.
    Marcos/Pedro operam o dia a dia; correção pós-fechamento é controlada."""
    return st.session_state.get('role') in ('admin', 'garantia_master')

def _garantias_visiveis(registros, role):
    """Um único recorte para fila, reincidência, indicadores e exportação."""
    if role not in ('admin', 'diretor', 'garantia', 'garantia_master'):
        return []
    ve_canceladas = role in ('admin', 'diretor', 'garantia_master')
    return [g for g in registros if isinstance(g, dict)
            and (ve_canceladas or g.get("status") != "Cancelada")]


def _garantia_data(valor):
    try:
        return datetime.strptime(str(valor)[:10], "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def _garantia_tempo_info(g, hoje=None):
    """Uma definição de duração para card e painel, sem criar datas de reabertura."""
    hoje = hoje or date.today()
    abertura = _garantia_data(g.get("criado_em"))
    chegada = _garantia_data(g.get("data_chegada"))
    envio = _garantia_data(g.get("data_envio"))
    confirmacao = _garantia_data(g.get("concluido_em"))
    encerrado = g.get("status") in STATUS_FINALIZADOS + ["Confirmado — aguardando R$ frete"]
    inicio = chegada or abertura
    if encerrado:
        fim = envio or confirmacao
        if not fim:
            return {"dias": None, "base": None,
                    "rotulo": "duração encerrada — data final não informada"}
        if chegada and envio:
            rotulo = "na empresa (chegada→envio)"
            base = "chegada_envio"
        elif envio:
            rotulo = "da abertura ao envio"
            base = "abertura_envio"
        elif chegada:
            rotulo = "da chegada à confirmação"
            base = "chegada_confirmacao"
        else:
            rotulo = "da abertura à confirmação"
            base = "abertura_confirmacao"
    else:
        # Reabrir preserva as datas anteriores no registro. O status atual
        # manda: envio/confirmação antigos não podem congelar um caso ativo.
        # Sem data própria do novo ciclo, explicita a data inicial registrada.
        fim = hoje
        rotulo = "na empresa" if chegada else "desde a abertura"
        base = "chegada_hoje" if chegada else "abertura_hoje"
        if envio or confirmacao:
            rotulo = ("desde a chegada registrada" if chegada else "desde a abertura") + " (caso ativo)"
    if not inicio:
        return {"dias": None, "base": None,
                "rotulo": "duração indisponível — data inicial não informada"}
    if fim < inicio or inicio > hoje or fim > hoje:
        return {"dias": None, "base": None,
                "rotulo": "duração indisponível — confira as datas"}
    return {"dias": (fim - inicio).days, "base": base, "rotulo": rotulo}


def _garantia_tempo_rotulo(g, hoje=None):
    info = _garantia_tempo_info(g, hoje)
    return f"{info['dias']}d {info['rotulo']}" if info["dias"] is not None else info["rotulo"]


def _garantia_periodo_vendas(meta):
    """Faixa mensal declarada pela Base Mãe, limitada à publicação da referência.

    A publicação limita a comparação, mas não prova a cobertura da coleta.
    Sem metadados válidos, não inventa uma janela de 12 meses ancorada em hoje.
    """
    try:
        limites = str(meta.get("periodo", "")).strip().lower().split(" a ")
        if len(limites) != 2:
            return None
        meses = []
        for limite in limites:
            mes, ano = limite.strip().split("/")
            if len(ano) == 4:
                if not ano.isdigit() or not 2000 <= int(ano) <= 2099:
                    return None
                ano = ano[-2:]
            if len(ano) != 2 or not ano.isdigit():
                return None
            # O parser compartilhado aceita rótulos com ano de dois dígitos.
            ym = _parse_label_ym(f"{mes}/{ano}")
            if ym is None:
                return None
            meses.append(ym)
        inicio = datetime(*meses[0], 1)
        ano_fim, mes_fim = meses[1]
        fim = datetime(ano_fim + (mes_fim == 12), mes_fim % 12 + 1, 1)
        publicado = datetime.strptime(str(meta.get("gerado_em", "")), "%Y-%m-%d %H:%M")
        # Registros e publicação têm precisão de minuto; inclui esse minuto.
        fim = min(fim, publicado + timedelta(minutes=1))
        return (inicio, fim) if inicio < fim else None
    except (AttributeError, TypeError, ValueError, OverflowError):
        return None


def _garantias_no_periodo_vendas(registros, periodo):
    if not periodo:
        return []
    inicio, fim = periodo
    resultado = []
    for g in registros:
        if not isinstance(g, dict) or g.get("status") == "Cancelada":
            continue
        try:
            criado = datetime.fromisoformat(str(g.get("criado_em", "")))
            if inicio <= criado < fim:
                resultado.append(g)
        except (TypeError, ValueError):
            continue
    return resultado


def _garantia_relacao_vendas(casos, unidades):
    """Relação descritiva; denominador ausente/inválido não vira zero por cento."""
    try:
        unidades = float(unidades)
        return casos / unidades if math.isfinite(unidades) and unidades > 0 else None
    except (TypeError, ValueError, OverflowError):
        return None


def load_garantias():
    data = _read_state_json("garantias.json", GARANTIAS_FILE, {"garantias": []})
    try:
        gs = data.get("garantias", [])
    except Exception:
        return []
    for g in gs:  # migração transparente de status antigos
        if g.get("status") in _STATUS_LEGADO:
            g["status"] = _STATUS_LEGADO[g["status"]]
    return gs

def add_garantia(reg):
    """Cria um registro de garantia (id sequencial G-0001...) de forma atômica.
    Retorna (id, ok)."""
    quem = st.session_state.get("user_name", "")
    agora = datetime.now().strftime("%Y-%m-%d %H:%M")
    out = {}

    def apply(d):
        gs = list(d.get("garantias", []))
        seq = 1 + max([int(g["id"].split("-")[1]) for g in gs if str(g.get("id", "")).startswith("G-")] or [0])
        gid = f"G-{seq:04d}"
        novo = dict(reg)
        novo.update({"id": gid, "criado_em": agora, "criado_por": quem, "status": "Aguardando chegada",
                     "historico": [{"em": agora, "por": quem, "acao": "Registro criado"}]})
        gs.append(novo)
        out["id"] = gid
        return {"garantias": gs}

    _, ok = _gh_mutate_json("garantias.json", GARANTIAS_FILE, apply, {"garantias": []})
    return out.get("id"), ok

def update_garantia(gid, updates, acao):
    """Atualiza um registro (merge) de forma atômica + linha de histórico."""
    quem = st.session_state.get("user_name", "")
    agora = datetime.now().strftime("%Y-%m-%d %H:%M")

    def apply(d):
        gs = list(d.get("garantias", []))
        for g in gs:
            if g.get("id") == gid:
                g.update(updates)
                g.setdefault("historico", []).append({"em": agora, "por": quem, "acao": acao})
                g["historico"] = g["historico"][-300:]  # cap: histórico não infla o json sem limite
                break
        return {"garantias": gs}

    _, ok = _gh_mutate_json("garantias.json", GARANTIAS_FILE, apply, {"garantias": []})
    return ok

def delete_garantia(gid):
    """EXCLUSÃO REAL (some da base, sem rastro) — SÓ o admin, para limpar
    simulações/testes. Caso real nunca se exclui: usa o status Cancelada
    (exclusão lógica com histórico)."""
    def apply(d):
        return {"garantias": [g for g in d.get("garantias", []) if g.get("id") != gid]}
    _, ok = _gh_mutate_json("garantias.json", GARANTIAS_FILE, apply, {"garantias": []})
    return ok

SILVER_DIST_FILE = os.path.join(os.path.dirname(__file__), "silver_distribuicao.json")

def load_silver_distribuicao():
    """FASE 2 (aprovada 22/07): última compra REAL por código de cliente, vinda
    do banco silver e publicada de hora em hora no branch state pela rotina
    local. Sem arquivo/token → {} e o app segue 100% na planilha (comportamento
    pré-Fase 2, degradação graciosa obrigatória)."""
    data = _read_state_json("silver_distribuicao.json", SILVER_DIST_FILE, {})
    return data if isinstance(data, dict) else {}

SILVER_MES_VIVO_FILE = os.path.join(os.path.dirname(__file__), "silver_mes_vivo.json")

def load_silver_mes_vivo():
    """Página Mês ao Vivo: receita do mês corrente do canal Distribuição, direto
    do banco silver, publicada de hora em hora no branch state pela rotina local
    (silver_mes_vivo.py). Sem arquivo/token → {} e a página avisa, sem quebrar."""
    data = _read_state_json("silver_mes_vivo.json", SILVER_MES_VIVO_FILE, {})
    return data if isinstance(data, dict) else {}

from util_comum import parse_label_ym as _parse_label_ym  # reuso (regra global nº 6)

def _rotulo_outro(valor, detalhe):
    """'Outro' vira 'Outro (detalhe)' quando o usuário especificou."""
    if valor == "Outro" and str(detalhe or "").strip():
        return f"Outro ({str(detalhe).strip()})"
    return valor or ""

def _link_rastreio(cod):
    """Código de rastreio como link clicável para o rastreamento dos Correios."""
    cod = str(cod or "").strip()
    if not cod:
        return "—"
    return f"[{cod}](https://rastreamento.correios.com.br/app/index.php?objeto={cod})"

def _garantia_custo_total(g, custo_map):
    """Custo real do caso: peças + fretes (vinda e volta) + extra + produto
    inteiro se trocado por novo."""
    total = 0.0
    for p in g.get("pecas", []):
        c = p.get("custo") or custo_map.get(str(p.get("sku", "")).strip(), 0) or 0
        total += (p.get("qtd", 1) or 1) * c
    total += g.get("frete_vinda", 0) or 0
    total += g.get("frete_volta", 0) or 0
    total += g.get("custo_extra", 0) or 0   # legado (registros antigos)
    if g.get("resultado") == "Trocada por produto novo":
        total += custo_map.get(str(g.get("produto_sku", "")).strip(), 0) or 0
    return round(total, 2)

# ============================================================
# PERSISTÊNCIA REMOTA (GitHub) — o disco do Streamlit Cloud é
# temporário: tudo que não está no repositório some quando o
# container reinicia. Arquivos de ESTADO (usuários, inativados,
# log de acesso) são salvos no branch 'state' do repo via API;
# a planilha enviada pelo Admin é commitada no branch 'main'.
# Requer GITHUB_TOKEN nos secrets do Streamlit Cloud. Sem token,
# tudo continua funcionando apenas localmente (sem persistência).
# ============================================================
_GH_API = "https://api.github.com"
_GH_REPO = "LeonardoDaros/propetz-bi"
_GH_STATE_BRANCH = "state"
_STATE_FILES = ["users.yaml", "inactive_clients.json", "access_log.json", "inactive_requests.json", "garantias.json", "login_attempts.json"]
_GH_WRITE_LOCK = threading.Lock()  # serializa escritas de estado: evita corrida read-SHA/PUT entre threads

def _gh_token():
    """Lê o GITHUB_TOKEN dos secrets, limpando espaços e aspas acidentais
    (erro de colagem comum). Token real não tem aspas/espaços, então isso é
    seguro e evita 401 por causa de um caractere sobrando."""
    try:
        raw = st.secrets.get("GITHUB_TOKEN", None)
    except Exception:
        return None
    if not raw:
        return None
    t = str(raw).strip().strip('"').strip("'").strip()
    return t or None

def _gh_headers(token):
    return {"Authorization": f"Bearer {token}", "Accept": "application/vnd.github+json"}

# Validade do token, capturada de GRAÇA do header que o GitHub devolve em toda
# resposta (github-authentication-token-expiration) — evita depender da memória
# do Leonardo p/ renovar antes de vencer.
_TOKEN_EXPIRA = [None]

def _registra_validade_token(resp):
    try:
        v = resp.headers.get("github-authentication-token-expiration")
        if v:
            _TOKEN_EXPIRA[0] = str(v)[:10]  # "AAAA-MM-DD"
    except Exception:
        pass

def _dias_p_expirar_token():
    """Dias até o token expirar (None se desconhecido ou sem validade)."""
    if not _TOKEN_EXPIRA[0]:
        return None
    try:
        return (datetime.strptime(_TOKEN_EXPIRA[0], "%Y-%m-%d").date() - date.today()).days
    except Exception:
        return None

def _gh_get_file(path, branch, token=None, so_sha=False):
    """Lê um arquivo do repo. Retorna (bytes, sha) ou (None, None).
    Arquivos de 1 a 100 MB: a API contents responde 200 com content=\"\" e
    encoding=\"none\" — nesse caso refaz o GET pedindo o corpo bruto (raw).
    Sem esse tratamento, um estado >1MB seria lido como VAZIO com sha válido
    (e a próxima gravação sobrescreveria tudo). so_sha=True pula o download
    do corpo nesse caso (para quem só precisa do sha, ex.: PUT)."""
    token = token or _gh_token()
    if not token:
        return None, None
    try:
        url = f"{_GH_API}/repos/{_GH_REPO}/contents/{path}"
        r = requests.get(url, params={"ref": branch},
                         headers=_gh_headers(token), timeout=15)
        _registra_validade_token(r)   # de graça: header da própria resposta
        if r.status_code == 200:
            data = r.json()
            if data.get("encoding") == "none":
                if so_sha:
                    return b"", data["sha"]
                hdr = _gh_headers(token)
                hdr["Accept"] = "application/vnd.github.raw+json"
                r2 = requests.get(url, params={"ref": branch}, headers=hdr, timeout=60)
                if r2.status_code == 200:
                    return r2.content, data["sha"]
                return None, None
            return base64.b64decode(data["content"]), data["sha"]
    except Exception:
        pass
    return None, None

def _gh_ensure_state_branch(token=None):
    """Cria o branch 'state' a partir do main, se ainda não existir."""
    token = token or _gh_token()
    if not token:
        return False
    try:
        r = requests.get(f"{_GH_API}/repos/{_GH_REPO}/git/ref/heads/{_GH_STATE_BRANCH}",
                         headers=_gh_headers(token), timeout=15)
        if r.status_code == 200:
            return True
        r_main = requests.get(f"{_GH_API}/repos/{_GH_REPO}/git/ref/heads/main",
                              headers=_gh_headers(token), timeout=15)
        if r_main.status_code != 200:
            return False
        sha = r_main.json()["object"]["sha"]
        r_new = requests.post(f"{_GH_API}/repos/{_GH_REPO}/git/refs",
                              json={"ref": f"refs/heads/{_GH_STATE_BRANCH}", "sha": sha},
                              headers=_gh_headers(token), timeout=15)
        return r_new.status_code in (200, 201)
    except Exception:
        return False

def _gh_put_file(path, content_bytes, message, branch, token=None):
    """Cria/atualiza um arquivo no repo. Retorna True se salvou.
    409/422 = o sha ficou velho entre o GET e o PUT (outra gravação no branch)
    ou dois PUTs criaram o mesmo arquivo ao mesmo tempo — relê o sha e tenta
    mais uma vez em vez de devolver falha espúria."""
    token = token or _gh_token()
    if not token:
        return False
    try:
        if branch == _GH_STATE_BRANCH and not _gh_ensure_state_branch(token=token):
            return False
        for _ in range(2):
            _, sha = _gh_get_file(path, branch, token=token, so_sha=True)
            payload = {"message": message,
                       "content": base64.b64encode(content_bytes).decode(),
                       "branch": branch}
            if sha:
                payload["sha"] = sha
            r = requests.put(f"{_GH_API}/repos/{_GH_REPO}/contents/{path}",
                             json=payload, headers=_gh_headers(token), timeout=60)
            if r.status_code in (200, 201):
                return True
            if r.status_code not in (409, 422):
                return False
        return False
    except Exception:
        return False

def _gh_put_file_status(path, content_bytes, message, branch, sha, token):
    """PUT com SHA explícito (concorrência otimista do GitHub). Retorna (ok, status).
    Status 409 = o arquivo mudou desde o SHA lido (outra gravação entrou no meio)."""
    try:
        payload = {"message": message, "content": base64.b64encode(content_bytes).decode(), "branch": branch}
        if sha:
            payload["sha"] = sha
        r = requests.put(f"{_GH_API}/repos/{_GH_REPO}/contents/{path}",
                         json=payload, headers=_gh_headers(token), timeout=30)
        return (r.status_code in (200, 201), r.status_code)
    except Exception:
        return (False, 0)

def _write_local_json(local_path, data):
    try:
        with open(local_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=1)
    except Exception:
        pass

def _gh_mutate_json(remote_name, local_path, apply_fn, default, token=None):
    """READ-MODIFY-WRITE ATÔMICO do estado no branch 'state'. Retorna (dados, ok):
    - lê o remoto FRESCO (conteúdo + SHA), aplica apply_fn(dados)->novos_dados, grava
      com aquele SHA. Se outra gravação entrou no meio (409), RE-LÊ e repete (até 5x) —
      duas inativações simultâneas NÃO se sobrescrevem.
    - ok=True só se REALMENTE persistiu (ou foi no-op). Se falhou (GitHub fora/token
      inválido), ok=False e NÃO grava local — para não dar falsa sensação de salvo.
    - sem token: modo local (ok=True, melhor esforço offline).
    Serializa no processo via _GH_WRITE_LOCK; o retry cobre concorrência entre containers.
    token: capturado na thread principal e passado adiante quando rodando em background."""
    tok = token or _gh_token()
    if not tok:
        cur = default
        if os.path.exists(local_path):
            try:
                with open(local_path, encoding='utf-8') as f:
                    cur = json.load(f)
            except Exception:
                cur = default
        new = apply_fn(cur)
        _write_local_json(local_path, new)
        _STATE_RAW_CACHE.pop(remote_name, None)
        return new, True
    with _GH_WRITE_LOCK:
        new, ok = None, False
        for _ in range(5):
            content, sha = _gh_get_file(remote_name, _GH_STATE_BRANCH, token=tok)
            if content:
                try:
                    cur = json.loads(content.decode('utf-8'))
                except Exception:
                    break  # remoto EXISTE mas está ilegível: abortar (ok=False) —
                           # jamais aplicar mudanças sobre default segurando um sha
                           # válido, isso sobrescreveria a base inteira em silêncio
            elif content is not None:
                cur = default  # arquivo existe porém vazio: seguro partir do default
            else:
                _gh_ensure_state_branch(token=tok)  # arquivo/branch ainda não existe
                cur, sha = default, None
            # snapshot ANTES do apply: o apply pode mutar registros no lugar (ex.: mudar
            # status para 'aprovado'), e aí comparar new==cur falharia em detectar a mudança
            before = json.dumps(cur, sort_keys=True, ensure_ascii=False) if content is not None else None
            new = apply_fn(cur)
            if before is not None and json.dumps(new, sort_keys=True, ensure_ascii=False) == before:
                ok = True  # nada mudou de fato (ex.: já inativo, ou duplicata) → já persistido
                break
            # compacto (sem indentação): metade do tamanho do indent=1 — afasta
            # o limite de 1 MB da API contents do GitHub
            body = json.dumps(new, ensure_ascii=False, separators=(",", ":")).encode('utf-8')
            done, status = _gh_put_file_status(remote_name, body, f"Estado: {remote_name}",
                                                _GH_STATE_BRANCH, sha, tok)
            if done:
                ok = True
                break
            if status != 409:
                break  # erro que não é conflito (token/rede): não adianta repetir
        if ok and new is not None:
            _write_local_json(local_path, new)  # só grava local se REALMENTE persistiu
        _STATE_RAW_CACHE.pop(remote_name, None)  # força reler o remoto real no próximo load
        return new, ok

def _push_state_file(filename, sync=False):
    """Envia um arquivo de estado para o branch 'state'. Por padrão em thread
    (não trava a UI); com sync=True grava de forma SÍNCRONA e confirmada — usado
    para dados críticos (inativações), onde 'melhor esforço' não basta.
    O token é capturado AQUI (thread principal) e passado adiante — ler st.secrets
    de dentro da thread pode falhar fora do contexto Streamlit."""
    token = _gh_token()
    if not token:
        return
    local = os.path.join(os.path.dirname(__file__), filename)
    if not os.path.exists(local):
        return
    with open(local, "rb") as f:
        content = f.read()

    def _send():
        # Lock serializa as escritas: sem ele, dois logins simultâneos podem ler o
        # mesmo SHA e a 2ª gravação é rejeitada (409) e perdida em silêncio.
        with _GH_WRITE_LOCK:
            _gh_put_file(filename, content, f"Estado: {filename}", _GH_STATE_BRANCH, token=token)

    if sync:
        _send()
    else:
        threading.Thread(target=_send, daemon=True).start()

def _gh_diagnose():
    """Diagnóstico SÍNCRONO da persistência. Retorna lista de
    (ok: bool, título: str, detalhe: str) para exibir na página Admin.
    Para no primeiro erro fatal — o último item indica onde quebrou."""
    out = []
    # 1) Ler secrets
    try:
        raw = st.secrets.get("GITHUB_TOKEN", None)
    except Exception as e:
        out.append((False, "Ler o campo Secrets",
                    f"Falhou ao ler os secrets ({e}). Quase sempre é erro de formato (TOML): "
                    "use aspas RETAS (\") e confira se não há outra linha quebrada no campo."))
        return out
    if not raw:
        out.append((False, "Token presente nos Secrets",
                    "GITHUB_TOKEN não foi encontrado. Confira: o nome é exatamente GITHUB_TOKEN; "
                    "o valor está entre aspas retas (\"); e nenhuma outra linha do campo Secrets "
                    "tem erro de formato (um erro em qualquer linha derruba TODOS os secrets)."))
        return out
    token = _gh_token()  # versão limpa (sem aspas/espaços)
    raw_str = str(raw)
    avisos = []
    if raw_str != raw_str.strip():
        avisos.append("tinha espaços/quebras de linha sobrando (removidos)")
    if raw_str.strip() != token:
        avisos.append("tinha aspas dentro do valor (removidas)")
    forma = "ghp_/github_pat_" if token.startswith(("ghp_", "github_pat_")) else f"'{token[:4]}…'"
    out.append((True, "Token presente",
                f"{len(token)} caracteres, formato {forma}." +
                (" ⚠️ Corrigido: " + "; ".join(avisos) if avisos else "")))
    # 2) Token válido?
    try:
        r = requests.get(f"{_GH_API}/user", headers=_gh_headers(token), timeout=15)
    except Exception as e:
        out.append((False, "Conexão com o GitHub", f"Não consegui falar com a API do GitHub: {e}"))
        return out
    if r.status_code == 401:
        out.append((False, "Token válido",
                    "GitHub recusou (401): token inválido, expirado ou copiado pela metade. "
                    "Gere um novo em github.com/settings/tokens e cole de novo."))
        return out
    if r.status_code == 200:
        out.append((True, "Token válido", f"Autenticado como '{r.json().get('login', '?')}'."))
    else:
        # Token FINE-GRAINED pode não responder /user (sem permissão de conta) e
        # mesmo assim gravar normalmente. NÃO é motivo para parar: o que vale é
        # o teste de escrita lá embaixo.
        out.append((True, "Token válido",
                    f"O GitHub não identificou a conta por aqui (HTTP {r.status_code}) — normal "
                    "em token fine-grained. Vale o teste de escrita abaixo."))
    # 3) Acesso ao repositório
    r = requests.get(f"{_GH_API}/repos/{_GH_REPO}", headers=_gh_headers(token), timeout=15)
    if r.status_code == 404:
        out.append((False, "Acesso ao repositório",
                    f"{_GH_REPO} não encontrado por este token. Em token clássico, marque o escopo "
                    "'repo'. Em token fine-grained, dê acesso a ESTE repositório com permissão "
                    "Contents: Read and write."))
        return out
    if r.status_code == 403:
        out.append((False, "Acesso ao repositório",
                    "Proibido (403): no token clássico falta o escopo 'repo'; no fine-grained, "
                    "confira se ESTE repositório está selecionado e com Contents: Read and write."))
        return out
    if r.status_code != 200:
        out.append((False, "Acesso ao repositório", f"Resposta {r.status_code}: {r.text[:120]}"))
        return out
    out.append((True, "Acesso ao repositório", f"{_GH_REPO} acessível."))
    # 4) Branch 'state'
    ok_branch = _gh_ensure_state_branch(token=token)
    out.append((ok_branch, "Branch 'state'",
                "Pronto." if ok_branch else "Não foi possível criar/verificar o branch 'state'."))
    if not ok_branch:
        return out
    # 5) Escrita real no branch 'state' (MESMO caminho do salvamento de inativações/
    #    usuários/log). Captura o status HTTP para dizer a causa exata se falhar.
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        _, sha = _gh_get_file("_diagnostico.txt", _GH_STATE_BRANCH, token=token)
        payload = {"message": "Teste de persistência",
                   "content": base64.b64encode(f"teste de escrita {stamp}".encode()).decode(),
                   "branch": _GH_STATE_BRANCH}
        if sha:
            payload["sha"] = sha
        rw = requests.put(f"{_GH_API}/repos/{_GH_REPO}/contents/_diagnostico.txt",
                          json=payload, headers=_gh_headers(token), timeout=30)
        if rw.status_code in (200, 201):
            _registra_validade_token(rw)
            _d = _dias_p_expirar_token()
            _val = (f" Validade do token: {_TOKEN_EXPIRA[0]} ({_d} dias)." if _d is not None
                    else " (Este token não informa data de expiração.)")
            out.append((True, "Escrita de teste (branch 'state')",
                        "Gravado com sucesso. ✅ A persistência de inativações, usuários e log "
                        "de acesso está FUNCIONANDO. (A planilha é salva no branch 'main' com o "
                        "mesmo token e deve funcionar igual.)" + _val))
        else:
            out.append((False, "Escrita de teste (branch 'state')",
                        f"O token é válido e enxerga o repo, mas a GRAVAÇÃO falhou: "
                        f"HTTP {rw.status_code} — {rw.text[:160]}"))
    except Exception as e:
        out.append((False, "Escrita de teste (branch 'state')", f"Falha na gravação: {e}"))
    return out

@st.cache_resource
def _sync_state_from_github():
    """Roda 1x por processo (boot do container): restaura os arquivos de
    estado a partir do branch 'state', se existirem lá."""
    for filename in _STATE_FILES:
        content, _ = _gh_get_file(filename, _GH_STATE_BRANCH)
        if content:
            try:
                with open(os.path.join(os.path.dirname(__file__), filename), "wb") as f:
                    f.write(content)
            except Exception:
                pass
    return True

# ============================================================
# AUTENTICAÇÃO — SÓ via sessão do servidor (st.session_state).
#
# HISTÓRICO (2026-07-23): havia login automático por parâmetros de URL
# (?u=&t=), com token = sha256("usuario:hash_senha:propetz")[:16]. Falha grave:
# o token derivava do hash da senha + palavra fixa NO CÓDIGO, então NUNCA
# expirava, valia em qualquer navegador e — como os hashes viviam no repo — era
# forjável para qualquer usuário. REMOVIDO. Login agora é só pelo formulário; a
# sessão vive no servidor (st.session_state), não trafega na URL e não é
# compartilhável. Efeito colateral aceito: recarregar a página (F5) pede login
# de novo. NUNCA reintroduzir auth por query param.
# ============================================================
def _strip_stale_auth_params():
    """Remove ?u=/?t= que um link antigo ainda possa carregar — não concede
    NENHUM acesso (só limpa a URL para o token não ficar visível/registrado)."""
    try:
        mudou = False
        for k in ("u", "t"):
            if k in st.query_params:
                del st.query_params[k]
                mudou = True
        return mudou
    except Exception:
        return False

# Expiração de sessão: por INATIVIDADE (3h) e por TEMPO MÁXIMO de vida (12h).
_SESSION_INATIVIDADE = 3 * 3600
_SESSION_MAX = 12 * 3600

def _touch_session():
    """Marca início (1x) e última atividade (a cada rerun) da sessão."""
    agora = datetime.now().timestamp()
    st.session_state.setdefault("_login_ts", agora)
    st.session_state["_last_seen"] = agora

def _session_expired():
    agora = datetime.now().timestamp()
    ini = st.session_state.get("_login_ts", agora)
    visto = st.session_state.get("_last_seen", agora)
    return (agora - visto > _SESSION_INATIVIDADE) or (agora - ini > _SESSION_MAX)

# ============================================================
# AUTHENTICATION
# ============================================================
def login_page():
    # st.form: digitar + clicar Entrar (ou pressionar Enter) vira UMA ação só.
    # Sem o form, o 1º clique apenas confirmava o campo de senha e o usuário
    # achava que o login não funcionava.
    _, login_column, _ = st.columns([1, 1.25, 1])
    with login_column:
        ui.login_header()
        with st.form("login_form"):
            username = st.text_input("Usuário", key="login_user")
            password = st.text_input("Senha", type="password", key="login_pass")
            submitted = st.form_submit_button("Entrar", use_container_width=True, type="primary")

    if submitted:
        username = str(username).strip().lower()
        ip = _client_ip()
        if not username or not password:
            st.error("Preencha usuário e senha.")
        else:
            # Proteção contra força bruta (por usuário E por IP, persistida)
            is_blocked, seconds_left = check_rate_limit(username, ip)
            if is_blocked:
                minutes = seconds_left // 60
                secs = seconds_left % 60
                st.error(f"🔒 Bloqueado por muitas tentativas incorretas. Tente novamente em {minutes}m{secs}s.")
            else:
                user = verify_login(username, password)
                if user:
                    clear_failed_attempts(username, ip)
                    st.session_state["authenticated"] = True
                    st.session_state["username"] = username
                    st.session_state["user_name"] = user["name"]
                    st.session_state["role"] = user["role"]
                    st.session_state["vendor_filter"] = user.get("vendor_filter")
                    _touch_session()
                    log_access(username, user["name"], "login")
                    st.rerun()
                else:
                    record_failed_attempt(username, ip)
                    attempts = _load_login_attempts()
                    info = attempts.get(f"u:{username}", {})
                    remaining = _BF_MAX - info.get("count", 0)
                    if remaining > 0:
                        st.error(f"Usuário ou senha incorretos. ({remaining} tentativas restantes)")
                    else:
                        st.error(f"🔒 Conta bloqueada por {_BF_JANELA // 60} minutos após muitas tentativas incorretas.")

# ============================================================
# DATA LOADING
# ============================================================
@st.cache_data(ttl=3600)
def load_data():
    """Load and process data from the Excel file."""
    # Try to find the Excel file
    # Search for any .xlsx file in the app directory
    app_dir = os.path.dirname(__file__)
    possible_names = [
        "RELATORIOS ESTADO-CLIENTES - ATUALIZADO.xlsx",
        "Relatorio Distribuidores Mensal.xlsx",
    ]
    possible_paths = []
    for name in possible_names:
        possible_paths.append(os.path.join(app_dir, name))
        possible_paths.append(os.path.join(app_dir, "..", name))

    # Also search for any xlsx in app_dir
    try:
        for f in os.listdir(app_dir):
            if f.endswith('.xlsx') and not f.startswith('~'):
                possible_paths.append(os.path.join(app_dir, f))
    except:
        pass

    # Usa o primeiro candidato que realmente é a planilha-fonte (tem a aba 'IA');
    # evita quebrar se outro .xlsx qualquer estiver na pasta.
    xlsx_path = None
    for p in possible_paths:
        if os.path.exists(p) and os.path.getsize(p) > 0:
            try:
                wb_test = openpyxl.load_workbook(p, read_only=True)
                has_ia = 'IA' in wb_test.sheetnames
                wb_test.close()
            except Exception:
                has_ia = False
            if has_ia:
                xlsx_path = p
                break

    if not xlsx_path:
        st.error("Arquivo Excel não encontrado. Faça upload na barra lateral.")
        return None, None, None, None, None, None

    return process_excel(xlsx_path)

def _handle_planilha_upload(uploaded):
    """Salva a planilha enviada localmente e tenta commitá-la no branch main
    do GitHub (persistência real — sobrevive a reinícios do Streamlit Cloud).
    Retorna True se o commit remoto funcionou."""
    data = bytes(uploaded.getbuffer())
    app_dir = os.path.dirname(__file__)
    with open(os.path.join(app_dir, "Relatorio Distribuidores Mensal.xlsx"), "wb") as f:
        f.write(data)
    # Remove o nome antigo para não sombrear a planilha nova na ordem de busca
    old = os.path.join(app_dir, "RELATORIOS ESTADO-CLIENTES - ATUALIZADO.xlsx")
    try:
        if os.path.exists(old):
            os.remove(old)
    except Exception:
        pass
    pushed = _gh_put_file("Relatorio Distribuidores Mensal.xlsx", data,
                          "Atualização da planilha via app", "main")
    st.cache_data.clear()
    return pushed

# ============================================================
# CURVA ABC POR VALOR — abc_valor.json traz o faturamento por SKU do canal
# Distribuição (últimos 12 meses, extraído da Base Mãe pelo script
# atualizar_abc_valor.py, que o deploy.bat roda automaticamente).
# ============================================================
ABC_VALOR_FILE = os.path.join(os.path.dirname(__file__), "abc_valor.json")

def load_abc_valor():
    """Retorna o dict do abc_valor.json ou None se não existir/estiver inválido."""
    if os.path.exists(ABC_VALOR_FILE):
        try:
            with open(ABC_VALOR_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if data.get("faturamento"):
                return data
        except Exception:
            pass
    return None

def apply_abc_by_value(products):
    """Recalcula a curva ABC dos produtos por FATURAMENTO (Pareto: A = 80%
    acumulado, B = até 95%, C = resto). Produtos sem venda no período = C.
    Se o abc_valor.json não existir, mantém a curva original da planilha."""
    abc_data = load_abc_valor()
    if not abc_data:
        for p in products:
            p['valor_12m'] = 0
            p['qty_12m'] = 0
        return products
    fat = abc_data["faturamento"]
    qts = abc_data.get("quantidade", {})
    for p in products:
        try:
            p['valor_12m'] = float(fat.get(p['code'], 0))
            p['qty_12m'] = float(qts.get(p['code'], 0))
        except Exception:
            p['valor_12m'] = 0
            p['qty_12m'] = 0
    total_val = sum(p['valor_12m'] for p in products)
    if total_val <= 0:
        return products
    cum = 0
    for p in sorted(products, key=lambda x: -x['valor_12m']):
        if p['valor_12m'] <= 0:
            p['abc'] = 'C'
            continue
        cum += p['valor_12m']
        p['abc'] = 'A' if cum <= 0.80 * total_val else ('B' if cum <= 0.95 * total_val else 'C')
    return products

# Mapeamento de vendedores (unificação de carteiras) — fonte única no
# util_comum (o coletor do Mês ao Vivo usa o MESMO mapa; editar lá)
from util_comum import VENDOR_MERGE, normalize_vendor  # noqa: F401

def process_excel(xlsx_path):
    """Process the Excel file and return structured data."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ---- MONTH LABELS from IA sheet ----
    ws_ia = wb['IA']
    month_labels = []
    for c in range(9, 73):
        v = ws_ia.cell(3, c).value
        if v:
            month_labels.append(str(v))
        else:
            break

    # Trim to actual data (find last month with data)
    valid_states = {'SP','PR','SC','MG','RS','RJ','CE','GO','ES','DF','BA','PB','RN','PE','MT','MS','RO','EX','PA','SE','MA','AM','PI','AP','AL','TO','RR','AC'}

    # ---- CLIENTS from IA sheet ----
    clients = []
    for r in range(4, ws_ia.max_row + 1):
        name = ws_ia.cell(r, 4).value
        state = ws_ia.cell(r, 5).value
        client_id = ws_ia.cell(r, 6).value
        vendor = ws_ia.cell(r, 7).value
        status = ws_ia.cell(r, 8).value

        if not name or not state or str(state).strip() not in valid_states:
            continue

        monthly = []
        for c in range(9, 9 + len(month_labels)):
            v = ws_ia.cell(r, c).value
            try:
                monthly.append(round(float(v), 2) if v else 0)
            except:
                monthly.append(0)

        status_str = str(status).strip().lower() if status else ''
        if 'inadimplente' in status_str:
            norm_status = 'Inadimplente'
        elif 'ativo' in status_str and 'ina' not in status_str:
            norm_status = 'Ativo'
        elif 'inativo' in status_str or 'ina' in status_str:
            norm_status = 'Inativo'
        elif 'permuta' in status_str:
            norm_status = 'Permuta'
        else:
            norm_status = status_str.title() if status_str else 'Desconhecido'

        clients.append({
            'name': str(name).strip(),
            'state': str(state).strip(),
            'id': str(client_id).strip() if client_id else '',
            'vendor': normalize_vendor(vendor),
            'status': norm_status,
            'monthly': monthly
        })

    # ---- ANALISES (optional – may not exist in lighter spreadsheets) ----
    analises = {}
    if 'Analises' in wb.sheetnames:
        ws_an = wb['Analises']
        for r in range(4, ws_an.max_row + 1):
            name = ws_an.cell(r, 3).value
            if not name:
                continue
            name = str(name).strip()
            credit = ws_an.cell(r, 7).value
            try:
                credit = float(credit) if credit else 0
            except:
                credit = 0

            totals = {}
            for idx, year in enumerate(['2021','2022','2023','2024','2025','2026']):
                v = ws_an.cell(r, 10 + idx).value
                try:
                    totals[year] = round(float(v), 2) if v else 0
                except:
                    totals[year] = 0

            mb = {}
            for idx, year in enumerate(['2021','2022','2023','2024','2025']):
                v = ws_an.cell(r, 17 + idx).value
                try:
                    mb[year] = int(v) if v else 0
                except:
                    mb[year] = 0

            am = {}
            for idx, year in enumerate(['2021','2022','2023','2024','2025']):
                v = ws_an.cell(r, 22 + idx).value
                try:
                    am[year] = round(float(v), 2) if v else 0
                except:
                    am[year] = 0

            analises[name] = {'credit': credit, 'totals': totals, 'months_bought': mb, 'avg_month': am}

    # ---- RECUPERAÇÃO (optional – may not exist in lighter spreadsheets) ----
    recuperacao = {}
    if 'Recuperação' in wb.sheetnames:
        ws_rec = wb['Recuperação']
        for r in range(5, ws_rec.max_row + 1):
            name = ws_rec.cell(r, 4).value
            if not name:
                continue
            name = str(name).strip()
            rec = ws_rec.cell(r, 9).value
            atencao = ws_rec.cell(r, 10).value
            if rec and str(rec).strip():
                recuperacao[name] = 'Recuperação'
            elif atencao and str(atencao).strip():
                recuperacao[name] = 'Atenção'
            else:
                recuperacao[name] = 'Saudável'

    # ---- PRODUTOS ABC (optional) ----
    products = []
    if 'Dados Demanda' in wb.sheetnames:
        ws_dem = wb['Dados Demanda']
        for r in range(6, ws_dem.max_row + 1):
            cod = ws_dem.cell(r, 3).value
            name = ws_dem.cell(r, 4).value
            cat = ws_dem.cell(r, 5).value
            total = ws_dem.cell(r, 30).value
            abc = ws_dem.cell(r, 32).value
            if not cod:
                continue
            try:
                total_val = int(float(total)) if total else 0
            except:
                total_val = 0
            cat_str = str(cat).strip() if cat else ''
            if cat_str.isdigit():
                cat_str = ''  # planilha às vezes traz o NCM no lugar da categoria
            products.append({
                'code': str(cod).strip(),
                'name': str(name).strip() if name else '',
                'category': cat_str,
                'total_qty': total_val,
                'abc': str(abc).strip() if abc else 'C'
            })
        products.sort(key=lambda x: -x['total_qty'])
        # Curva por VALOR substitui a curva por quantidade vinda da planilha
        products = apply_abc_by_value(products)

    # ---- CLIENT × PRODUCT DATA (Base de DadosProdutos, RIGHT SIDE cols 31+) ----
    client_products = []
    if 'Base de DadosProdutos' in wb.sheetnames:
        ws_bp = wb['Base de DadosProdutos']
        # Right side: C32=product_name, C34=product_code, C35=total_qty, C37=client_name, C38=client_code
        for r in range(5, ws_bp.max_row + 1):
            product_name = ws_bp.cell(r, 32).value
            product_code = ws_bp.cell(r, 34).value
            client_name_bp = ws_bp.cell(r, 37).value
            client_code_bp = ws_bp.cell(r, 38).value
            if not client_name_bp or not product_code:
                continue
            raw_qty = ws_bp.cell(r, 35).value
            try:
                total_qty = int(float(raw_qty)) if raw_qty else 0
            except:
                total_qty = 0
            if total_qty > 0:
                client_products.append({
                    'client_id': str(client_code_bp).strip() if client_code_bp else '',
                    'client_name': str(client_name_bp).strip(),
                    'product_code': str(product_code).strip(),
                    'product_name': str(product_name).strip() if product_name else '',
                    'total_qty': total_qty
                })
    df_client_products = pd.DataFrame(client_products) if client_products else pd.DataFrame()

    # ---- TRIM MONTHS TO ACTUAL DATA (must happen BEFORE risk calc) ----
    last_data_idx = 0
    for mi in range(len(month_labels)-1, -1, -1):
        total = sum(c['monthly'][mi] for c in clients)
        if total > 0:
            last_data_idx = mi
            break

    month_labels = month_labels[:last_data_idx+1]
    for c in clients:
        c['monthly'] = c['monthly'][:last_data_idx+1]

    # ---- ENRICH CLIENTS ----
    # 2026 vai até (52,64): o range antigo (52,54) parava em fev/26 e o gráfico
    # "Receita por Ano" subcontava o ano (achado da auditoria da Fase 2)
    year_ranges = {'2021':(0,4),'2022':(4,16),'2023':(16,28),'2024':(28,40),'2025':(40,52),'2026':(52,64)}

    for client in clients:
        cn = client['name']
        a = analises.get(cn)
        if not a:
            for ak, av in analises.items():
                if cn[:20].upper() == ak[:20].upper():
                    a = av
                    break

        if a:
            client['credit_limit'] = a['credit']
            client['yearly_totals'] = a['totals']
            client['months_bought'] = a['months_bought']
            client['avg_month'] = a['avg_month']
        else:
            client['credit_limit'] = 0
            yt = {}
            mb = {}
            am = {}
            for year, (start, end) in year_ranges.items():
                vals = client['monthly'][start:end]
                yt[year] = round(sum(vals), 2)
                bought = sum(1 for v in vals if v > 0)
                mb[year] = bought
                am[year] = round(yt[year] / bought, 2) if bought > 0 else 0
            client['yearly_totals'] = yt
            client['months_bought'] = mb
            client['avg_month'] = am

        client['total_geral'] = round(sum(client['monthly']), 2)

        # Risk — based on months since last purchase (using trimmed data)
        risk = recuperacao.get(cn)
        if not risk:
            for rk, rv in recuperacao.items():
                if cn[:20].upper() == rk[:20].upper():
                    risk = rv
                    break

        if not risk:
            last_idx = -1
            for i in range(len(client['monthly'])-1, -1, -1):
                if client['monthly'][i] > 0:
                    last_idx = i
                    break
            months_since = (len(month_labels) - 1 - last_idx) if last_idx >= 0 else 999
            risk = 'Recuperação' if months_since >= 6 else ('Atenção' if months_since >= 3 else 'Saudável')

        client['risk'] = risk

        last_idx = -1
        for i in range(len(client['monthly'])-1, -1, -1):
            if client['monthly'][i] > 0:
                last_idx = i
                break
        client['last_purchase'] = month_labels[last_idx] if last_idx >= 0 else 'Nunca'
        client['months_since'] = (len(month_labels) - 1 - last_idx) if last_idx >= 0 else 999

    # ---- FASE 2 (22/07): última compra REAL do banco silver, por CÓDIGO ----
    # Reconciliação validada pela sombra (3 auditorias): o banco é CEGO antes
    # de 2026 e a planilha ENVELHECE após o upload — vale a data mais recente
    # das duas fontes, com a régua ancorada em HOJE (não no último mês carregado).
    silver_dist = load_silver_distribuicao()
    # BLINDAGEM (auditoria Fase 2): estrutura inesperada no json NUNCA pode
    # derrubar o load_data — qualquer coisa fora do formato vira "sem silver"
    silver_cli = silver_dist.get("clientes", {}) if isinstance(silver_dist, dict) else {}
    if not isinstance(silver_cli, dict):
        silver_cli = {}
    if silver_cli and month_labels:
        _anc = _parse_label_ym(month_labels[-1])
        _hoje = date.today()
        for client in clients:
            try:
                info = silver_cli.get(client['id'])
                if not isinstance(info, dict):
                    continue
                _u = datetime.strptime(str(info.get("ultima_compra_real", ""))[:10],
                                       "%Y-%m-%d").date()
                pl_ym = None
                if _anc and client['months_since'] < 999:
                    _tot = _anc[0] * 12 + (_anc[1] - 1) - client['months_since']
                    pl_ym = (_tot // 12, _tot % 12 + 1)
                bk_ym = (_u.year, _u.month)
                ef_ym = max(x for x in (pl_ym, bk_ym) if x)
                meses = max(0, (_hoje.year * 12 + _hoje.month) - (ef_ym[0] * 12 + ef_ym[1]))
            except Exception:
                continue  # cliente com dado estranho fica 100% na planilha
            client['months_since'] = meses
            client['risk'] = ('Recuperação' if meses >= 6
                              else ('Atenção' if meses >= 3 else 'Saudável'))
            if bk_ym >= (pl_ym or (0, 0)):
                # banco vence (ou empata no mês): mostra a DATA REAL da NF
                client['last_purchase'] = _u.strftime("%d/%m/%Y")

    # Convert to DataFrame
    df_clients = pd.DataFrame(clients)
    df_products = pd.DataFrame(products)

    # ---- SKU QUANTITY DATA ----
    df_sku = pd.DataFrame()
    sku_sheet_name = None
    for sn in wb.sheetnames:
        if 'qtd' in sn.lower() and 'cliente' in sn.lower():
            sku_sheet_name = sn
            break
    if sku_sheet_name:
        ws_sku = wb[sku_sheet_name]
        sku_data = []

        # Auto-detect month column positions from row 1.
        # CUIDADO (aprendido em jul/26): o Excel converte cabeçalhos digitados como
        # "MAR/26" em DATA de verdade (datetime). O parser precisa aceitar os dois
        # formatos, senão os blocos novos são pulados em silêncio e o mix congela.
        _MES_ABREV = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN',
                      'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']
        month_cols = []
        month_headers = []
        for c in range(1, ws_sku.max_column + 1):
            v = ws_sku.cell(1, c).value
            if v is None:
                continue
            if isinstance(v, (datetime, date)):
                month_cols.append(c)
                month_headers.append(f"{_MES_ABREV[v.month - 1]}/{str(v.year)[-2:]}")
            elif '/' in str(v):
                month_cols.append(c)
                month_headers.append(str(v).strip().upper())

        # Read data rows starting from row 3 (0-indexed row 2)
        for r in range(3, ws_sku.max_row + 1):
            # For each month block
            for month_idx, base_col in enumerate(month_cols):
                if month_idx >= len(month_headers):
                    break
                mes = month_headers[month_idx]

                # Columns within each block: Produto, SKU, Quantidade, Vendedor, Cliente, Código Cliente
                # Offsets from base_col (1-indexed): 0, 1, 2, 3, 4, 5
                produto = ws_sku.cell(r, base_col).value
                sku = ws_sku.cell(r, base_col + 1).value
                quantidade_raw = ws_sku.cell(r, base_col + 2).value
                vendedor = ws_sku.cell(r, base_col + 3).value
                cliente = ws_sku.cell(r, base_col + 4).value
                cod_cliente = ws_sku.cell(r, base_col + 5).value

                # Only add if we have the key fields
                if produto and sku and cod_cliente:
                    try:
                        quantidade = int(float(quantidade_raw)) if quantidade_raw else 0
                    except:
                        quantidade = 0

                    if quantidade > 0:
                        sku_data.append({
                            'mes': mes,
                            'produto': str(produto).strip(),
                            'sku': str(sku).strip(),
                            'quantidade': quantidade,
                            'vendedor': str(vendedor).strip() if vendedor else '',
                            'cliente': str(cliente).strip() if cliente else '',
                            'cod_cliente': str(cod_cliente).strip()
                        })

        if sku_data:
            df_sku = pd.DataFrame(sku_data)

    # If no Base de DadosProdutos but we have SKU data, build client×product from SKU
    if len(df_client_products) == 0 and len(df_sku) > 0:
        cp_from_sku = df_sku.groupby(['cod_cliente', 'sku', 'produto']).agg(
            total_qty=('quantidade', 'sum')
        ).reset_index()
        # Try to match client names from df_clients
        id_to_name = dict(zip(df_clients['id'].astype(str).str.strip(), df_clients['name']))
        cp_from_sku['client_id'] = cp_from_sku['cod_cliente'].astype(str).str.strip()
        cp_from_sku['client_name'] = cp_from_sku['client_id'].map(id_to_name).fillna('')
        cp_from_sku['product_code'] = cp_from_sku['sku']
        cp_from_sku['product_name'] = cp_from_sku['produto']
        df_client_products = cp_from_sku[['client_id','client_name','product_code','product_name','total_qty']].copy()

    return df_clients, df_products, df_client_products, month_labels, year_ranges, df_sku

# ============================================================
# HELPER FUNCTIONS
# ============================================================
def fmt_brl(v):
    if v >= 1e6:
        return f"R$ {v/1e6:.1f}M"
    if v >= 1e3:
        return f"R$ {v/1e3:.1f}k"
    return f"R$ {v:.0f}"

def fmt_brl_full(v):
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def _year_of_label(lbl):
    """Extrai o ano (ex: '2025') de um rótulo de mês tipo 'jan/25'."""
    parts = str(lbl).replace('-', '/').split('/')
    y = parts[-1].strip() if len(parts) >= 2 else ''
    return f"20{y}" if len(y) == 2 else y

def compute_preset_indices(preset, months):
    """Converte um preset de período ('Últimos 12 meses' etc.) no set de
    índices correspondentes da lista de meses."""
    n = len(months)
    years = []
    for lbl in months:
        y = _year_of_label(lbl)
        if y and y not in years:
            years.append(y)
    cur = years[-1] if years else ''
    prev = years[-2] if len(years) >= 2 else ''

    if preset == "Este mês":
        return set([n - 1]) if n else set()
    if preset == "Últimos 3 meses":
        return set(range(max(0, n - 3), n))
    if preset == "Últimos 6 meses":
        return set(range(max(0, n - 6), n))
    if preset == "Últimos 12 meses":
        return set(range(max(0, n - 12), n))
    if preset.startswith("Este ano") and cur:
        return {i for i, lbl in enumerate(months) if _year_of_label(lbl) == cur}
    if preset.startswith("Ano passado") and prev:
        return {i for i, lbl in enumerate(months) if _year_of_label(lbl) == prev}
    return set(range(n))  # "Tudo" (desde o início)

def show_money_table(df_disp, money_cols, **kwargs):
    """Tabela com colunas em R$ formatadas SEM virar texto — assim a ordenação
    por clique no cabeçalho continua numérica (e não alfabética).
    Retorna o resultado de st.dataframe (evento de seleção, quando usado)."""
    fmt_map = {c: fmt_brl_full for c in money_cols if c in df_disp.columns}
    try:
        return st.dataframe(df_disp.style.format(fmt_map), **kwargs)
    except Exception:
        d = df_disp.copy()
        for c in fmt_map:
            d[c] = d[c].apply(fmt_brl_full)
        return st.dataframe(d, **kwargs)

def _sku_stats(df_sku):
    """Por SKU: (qtd típica/mês = mediana entre compradores, nº de compradores,
    qtd/mês por cliente). Base para todas as estimativas de oportunidade."""
    if len(df_sku) == 0:
        return {}, {}, pd.DataFrame()
    per = df_sku.groupby(['sku', 'cod_cliente']).agg(q=('quantidade', 'sum'), m=('mes', 'nunique')).reset_index()
    per['pm'] = per['q'] / per['m'].clip(lower=1)
    per['cod_cliente'] = per['cod_cliente'].astype(str).str.strip()
    typical = per.groupby('sku')['pm'].median().to_dict()
    buyers = per.groupby('sku')['cod_cliente'].nunique().to_dict()
    return typical, buyers, per

def _preco_medio_map(products_df):
    """Preço médio real por SKU (faturamento 12m ÷ quantidade 12m, Base Mãe)."""
    out = {}
    if 'valor_12m' in products_df.columns and 'qty_12m' in products_df.columns:
        for code, v, q in zip(products_df['code'], products_df['valor_12m'], products_df['qty_12m']):
            if q and q > 0 and v and v > 0:
                out[code] = v / q
    return out

def annual_value_estimate(monthly):
    """Valor anual estimado do cliente: ticket médio dos meses COM compra
    nos últimos 12 meses × 12. Se não comprou nos últimos 12 meses, usa o
    ticket médio histórico. (Substitui o cálculo antigo travado em 2024/2023.)"""
    if not isinstance(monthly, (list, tuple)) or len(monthly) == 0:
        return 0
    last12 = [v for v in monthly[-12:] if v > 0]
    if last12:
        return sum(last12) / len(last12) * 12
    hist = [v for v in monthly if v > 0]
    if hist:
        return sum(hist) / len(hist) * 12
    return 0

def esc(v):
    """Escapa valor de origem EXTERNA (planilha, cadastro, nome de usuário)
    antes de injetar em HTML via unsafe_allow_html — barra XSS."""
    return html.escape(str(v if v is not None else ""))

def risk_badge(risk):
    if risk == 'Recuperação':
        return '<span class="badge badge-red">Recuperação</span>'
    elif risk == 'Atenção':
        return '<span class="badge badge-yellow">Atenção</span>'
    return '<span class="badge badge-green">Saudável</span>'

def status_badge(status):
    if status == 'Ativo':
        return '<span class="badge badge-green">Ativo</span>'
    elif status == 'Inativo':
        return '<span class="badge badge-red">Inativo</span>'
    return f'<span class="badge badge-blue">{esc(status)}</span>'

def insight_html(type_, label, text, action):
    # label/text/action podem carregar dado externo (ex.: nome de vendedor) —
    # escapados por segurança; são texto puro, então escapar não muda o visual.
    css_class = f"insight-{type_}" if type_ in ('danger','warning','success') else ''
    return f"""
    <div class="insight-card {css_class}">
        <div class="insight-type">{esc(label)}</div>
        <div class="insight-text">{esc(text)}</div>
        <div class="insight-action">{esc(action)}</div>
    </div>
    """

# ============================================================
# PAGE: GARANTIAS
# ============================================================
def page_garantias(products_df, df_clients):
    st.header("🔧 Garantias e Assistência")

    meta = load_abc_valor() or {}
    custo_map = {str(k).strip(): v for k, v in (meta.get("custo_unitario") or {}).items()}
    vendas_map = {str(k).strip(): v for k, v in (meta.get("vendas_12m_todos_canais") or {}).items()}

    prod_opts = [f"{r['code']} — {r['name']}" for _, r in products_df.iterrows()]
    cli_dist = sorted(df_clients['name'].unique().tolist()) if len(df_clients) else []

    def _sku_de(opt):
        return opt.split(" — ")[0].strip() if opt else ""

    # O recorte antecede qualquer uso: nem indicadores, alertas ou CSV podem
    # revelar um registro que o papel não pode consultar na fila.
    garantias = _garantias_visiveis(load_garantias(), st.session_state.get("role"))
    tab_novo, tab_bancada, tab_painel = st.tabs(["📥 Nova Garantia", "🔨 Bancada / Fila", "📊 Painel"])

    # ---------------- NOVA GARANTIA ----------------
    with tab_novo:
        st.caption("Registro de entrada — 2 minutos por caso. O que a NF não conta fica registrado aqui.")
        if st.session_state.pop("gar_flash", None):
            st.success(st.session_state.pop("gar_flash_msg", "✅ Registrado."))
        # Campos FORA de st.form de propósito: marcar "Outro" mostra o campo
        # de especificação NA HORA (form fechado não reage antes do submit).
        _NK = ("gn_canal", "gn_canal_outro", "gn_clidist", "gn_clitxt", "gn_prod",
               "gn_dtcompra", "gn_rastreio", "gn_cf_nome", "gn_cf_nf", "gn_cf_chave",
               "gn_def", "gn_def_outro", "gn_obs", "gn_prio")

        def _autofill_cliente():
            # escolher o distribuidor preenche o campo Cliente sozinho — evita o
            # funcionário digitar ali, por engano, o nome do cliente FINAL
            _sel = st.session_state.get("gn_clidist")
            if _sel:
                st.session_state["gn_clitxt"] = _sel

        c1, c2 = st.columns(2)
        canal = c1.selectbox("Canal *", CANAIS_GARANTIA, index=None,
                             placeholder="De onde vem o cliente...", key="gn_canal")
        canal_outro = ""
        if canal == "Outro":
            canal_outro = c1.text_input("Qual canal? *", key="gn_canal_outro",
                                        placeholder="Ex.: marketplace, representante...")
        cliente_dist = c2.selectbox("Cliente da Distribuição (se for)", cli_dist, index=None,
                                    placeholder="Buscar na base...", key="gn_clidist",
                                    on_change=_autofill_cliente)
        cliente_txt = st.text_input("Cliente (nome/razão — preenche sozinho ao escolher o distribuidor)",
                                    key="gn_clitxt")
        _cli_chk = cliente_txt.strip().lower()
        if len(_cli_chk) >= 3:
            _reinc = [x for x in garantias if _cli_chk in str(x.get("cliente", "")).lower()
                      or _cli_chk in str(x.get("cliente_final", "")).lower()]
            if _reinc:
                st.warning(f"⚠️ **Cliente reincidente:** já existem {len(_reinc)} garantia(s) dele "
                           f"({', '.join(x['id'] for x in _reinc[-5:])}). "
                           "Avalie marcar a PRIORIDADE como Alta/Urgente ali embaixo.")
        c3, c4 = st.columns([2, 1])
        produto = c3.selectbox("Produto *", prod_opts, index=None,
                               placeholder="Buscar por código ou nome...", key="gn_prod")
        data_compra = c4.date_input("Data da compra do reclamante (se souber)", value=None,
                                    format="DD/MM/YYYY", key="gn_dtcompra")
        rastreio_entrada = st.text_input("📬 Código de rastreamento da VINDA (sai junto com o protocolo)",
                                         key="gn_rastreio",
                                         placeholder="Ex.: AA123456789BR — na Bancada vira link p/ os Correios...")
        cf_nome, cf_nf, cf_chave = "", "", ""
        if canal == "Distribuição":
            st.markdown("**Venda ao cliente final** — ⚠️ sem NF da venda do distribuidor "
                        "ao cliente, a garantia NÃO é aceita:")
            cf1, cf2 = st.columns(2)
            cf_nome = cf1.text_input("Cliente do distribuidor (consumidor final) *", key="gn_cf_nome",
                                     placeholder="Quem comprou do distribuidor e está reclamando...")
            cf_nf = cf2.text_input("NF da venda ao cliente final *", key="gn_cf_nf",
                                   placeholder="Nota distribuidor → cliente (obrigatória)...")
            cf_chave = st.text_input("Chave de acesso da NF (44 dígitos)",
                                     key="gn_cf_chave",
                                     placeholder="Digite aqui no registro — a Bancada só consulta, não edita...")
        c7, _ = st.columns([1, 1])
        defeito = c7.selectbox("Defeito relatado *", DEFEITOS_GARANTIA, index=None,
                               placeholder="Categoria do problema...", key="gn_def")
        defeito_outro = ""
        if defeito == "Outro":
            defeito_outro = c7.text_input("Qual defeito? *", key="gn_def_outro",
                                          placeholder="Descreva em poucas palavras...")
        co1, co2 = st.columns([3, 1])
        defeito_obs = co1.text_area("Relato do cliente / observações", height=80, key="gn_obs",
                                    placeholder="O que o cliente disse? Quando começou? Acessórios recebidos junto...")
        prioridade = co2.selectbox("Prioridade", PRIORIDADES_GARANTIA, index=0, key="gn_prio",
                                   help="🔴 Urgente / 🟠 Alta furam a fila da bancada — use p/ cliente "
                                        "reincidente ou caso já desgastado.")
        if st.button("📥 Registrar entrada", type="primary", key="gn_enviar"):
            # o campo de TEXTO manda: é o que o funcionário está vendo na tela
            # (o autofill já o preenche ao escolher o distribuidor; se ele editar
            # depois — 'Loja 2', razão social corrigida — a edição vale)
            cliente = cliente_txt.strip() or (cliente_dist or "").strip()
            faltas = [n for n, v in [("canal", canal), ("cliente", cliente), ("produto", produto),
                                     ("defeito relatado", defeito)] if not v]
            if canal == "Outro" and not canal_outro.strip():
                faltas.append("qual canal (marcou 'Outro')")
            if defeito == "Outro" and not defeito_outro.strip():
                faltas.append("qual defeito (marcou 'Outro')")
            if canal == "Distribuição":
                if not cf_nome.strip():
                    faltas.append("o CLIENTE FINAL (quem comprou do distribuidor)")
                if not cf_nf.strip():
                    faltas.append("a NF da venda ao cliente final (sem nota, não aceitamos a garantia)")
            if faltas:
                st.error("⚠️ Preencha: " + ", ".join(faltas) + ". Nada foi registrado.")
            else:
                sku = _sku_de(produto)
                gid, ok = add_garantia({
                    "canal": canal, "canal_outro": canal_outro.strip(),
                    "cliente": cliente, "produto_sku": sku,
                    "produto_nome": produto.split(" — ", 1)[1] if " — " in produto else produto,
                    "empresa_nf": "",   # definida na Bancada, quando o produto chega
                    "data_compra": data_compra.strftime("%Y-%m-%d") if data_compra else "",
                    "nf_entrada": "",   # idem
                    "rastreio_entrada": rastreio_entrada.strip(),
                    "cliente_final": cf_nome.strip(), "cliente_final_nf": cf_nf.strip(),
                    "cliente_final_nf_chave": "".join(ch for ch in (cf_chave or "") if ch.isdigit()),
                    "defeito": defeito, "defeito_outro": defeito_outro.strip(),
                    "defeito_obs": defeito_obs.strip(), "prioridade": prioridade,
                    "pecas": [], "custo_extra": 0, "diagnostico_causa": "", "diagnostico_obs": "",
                    "resultado": "", "nf_saida": "", "custo_total": 0,
                })
                if ok:
                    for k in _NK:
                        st.session_state.pop(k, None)  # limpa o formulário p/ o próximo caso
                    st.session_state["gar_flash"] = True
                    st.session_state["gar_flash_msg"] = f"✅ Garantia **{gid}** registrada — já está na fila da Bancada."
                    st.rerun()
                else:
                    st.error("Não consegui salvar no GitHub agora. Tente de novo em instantes.")

    # ---------------- BANCADA / FILA (sub-abas por status) ----------------
    with tab_bancada:
        _fld = st.session_state.pop("gar_flash_del", None)
        if _fld:
            st.success(_fld)
        cb1, cb2, cb3 = st.columns([2, 1, 1])
        busca = cb1.text_input("🔍 Buscar (id, cliente, cliente final, produto, NF, empresa)", key="gar_busca")
        dt_de = cb2.date_input("Registradas DE", value=None, format="DD/MM/YYYY", key="gar_dtde")
        dt_ate = cb3.date_input("ATÉ", value=None, format="DD/MM/YYYY", key="gar_dtate")

        def _filtra(status_lista):
            out = [g for g in garantias if g.get("status") in status_lista]
            if busca.strip():
                s = busca.strip().lower()
                out = [g for g in out if any(s in str(g.get(k, "")).lower() for k in
                       ("id", "cliente", "cliente_final", "produto_nome", "produto_sku",
                        "empresa_nf", "nf_entrada", "nf_saida", "cliente_final_nf"))]
            if dt_de or dt_ate:
                def _dreg(g):
                    try:
                        return datetime.strptime(str(g.get("criado_em", ""))[:10], "%Y-%m-%d").date()
                    except Exception:
                        return None
                out = [g for g in out if (_d := _dreg(g)) is not None
                       and (not dt_de or _d >= dt_de) and (not dt_ate or _d <= dt_ate)]
            # urgentes primeiro, depois altas; empate = mais antiga primeiro
            _rank = {"Urgente": 0, "Alta": 1}
            out.sort(key=lambda g: (_rank.get(g.get("prioridade", "Normal"), 2),
                                    g.get("criado_em", "")))
            return out

        def _render_fila(lista, tk):
            # A seleção usa o ID, nunca a posição: filtros, ordenação e mudança
            # de status não podem abrir o formulário de outro caso por engano.
            chave_selecao = f"gar_atendimento_{tk}"
            if not lista:
                st.session_state.pop(chave_selecao, None)
                st.info("Nenhuma garantia aqui.")
                return
            por_id = {g["id"]: g for g in lista}
            ids = list(por_id)
            atual = st.session_state.get(chave_selecao)
            # Reatribuir antes do widget preserva o ID quando a lista de opções
            # muda no rerun; ID removido pela busca/fila volta ao primeiro válido.
            st.session_state[chave_selecao] = atual if atual in por_id else ids[0]
            resumo = [{
                "ID": g["id"],
                "Prioridade": f"{_PRIO_ICONE.get(g.get('prioridade', ''), '')}{g.get('prioridade') or 'Normal'}",
                "Cliente final / cliente": g.get("cliente_final") or g.get("cliente") or "—",
                "Produto": g.get("produto_nome") or g.get("produto_sku") or "—",
                "Situação": g.get("status") or "—",
                "Tempo": _garantia_tempo_rotulo(g),
            } for g in lista]
            st.dataframe(pd.DataFrame(resumo), use_container_width=True, hide_index=True,
                         height=min(315, 35 * len(resumo) + 38))
            st.caption("Escolha um atendimento abaixo para consultar a ficha ou registrar uma atualização.")

            def _rotulo_atendimento(gid):
                g = por_id[gid]
                cliente = g.get("cliente_final") or g.get("cliente") or "Sem cliente"
                produto = g.get("produto_nome") or g.get("produto_sku") or "Sem produto"
                prioridade = g.get("prioridade") or "Normal"
                return f"{gid} · {cliente} · {produto} · {prioridade}"

            escolhido = st.selectbox("Abrir atendimento", ids, key=chave_selecao,
                                     format_func=_rotulo_atendimento,
                                     help="Busque pelo protocolo, cliente ou produto. A ficha aberta é a do ID selecionado.")
            if escolhido not in por_id:
                st.info("Selecione um atendimento disponível nesta fila.")
                return
            # Somente um caso gera widgets. As keys dos formulários continuam
            # contendo fila + ID, impedindo reaproveitar um submit de outro caso.
            lista = [por_id[escolhido]]
            for g in lista:
                dias = f" | {_garantia_tempo_rotulo(g)}"
                icone = {"Aguardando chegada": "📬", "Em bancada": "🔧", "Aguardando peça": "📦",
                         "Confirmado — aguardando R$ frete": "🚚",
                         "Concluída": "✅", "Cancelada": "🚫"}.get(g.get("status"), "•")
                _pri = _PRIO_ICONE.get(g.get("prioridade", ""), "")
                with st.expander(f"{_pri}{icone} {g['id']} — {g.get('produto_nome','')[:40]} — {g.get('cliente','')[:30]} "
                                 f"[{g.get('status')}]{dias}", expanded=True):
                    _dtc = g.get("data_compra") or ""
                    if _dtc:
                        try:
                            _dtc = datetime.strptime(_dtc, "%Y-%m-%d").strftime("%d/%m/%Y")
                        except Exception:
                            pass
                    _linha_meta = (f"**Canal:** {_rotulo_outro(g.get('canal',''), g.get('canal_outro'))} | "
                                   f"**Prioridade:** {_pri}{g.get('prioridade') or 'Normal'} | "
                                   f"**Empresa NF:** {g.get('empresa_nf') or '—'} | "
                                   f"**NF entrada:** {g.get('nf_entrada') or '—'} | "
                                   f"**Rastreio entrada:** {_link_rastreio(g.get('rastreio_entrada'))} | "
                                   f"**Compra:** {_dtc or '—'} | **Registro:** {g.get('criado_em','')} "
                                   f"por {g.get('criado_por','')}")
                    if g.get("canal") == "Distribuição":
                        # cliente final, NF e chave se definem na NOVA GARANTIA e aqui
                        # são só leitura — editar na bancada geraria divergência
                        _chv = g.get("cliente_final_nf_chave") or ""
                        _linha_meta += (f"  \n🏷️ **Cliente final (comprou do distribuidor):** "
                                        f"{g.get('cliente_final') or '—'} | "
                                        f"**NF da venda ao cliente final:** {g.get('cliente_final_nf') or '—'} | "
                                        f"**Chave:** {f'`{_chv}`' if _chv else '—'}")
                    st.markdown(_linha_meta)
                    st.markdown(f"**Defeito relatado:** {_rotulo_outro(g.get('defeito',''), g.get('defeito_outro'))} "
                                f"— {g.get('defeito_obs') or 'sem obs.'}")

                    _fechada = g.get("status") in STATUS_FINALIZADOS

                    if _fechada and not can_edit_garantia_fechada():
                        # Marcos/Pedro: finalizada = somente leitura
                        _fretes = f"vinda {fmt_brl_full(g.get('frete_vinda', 0) or 0)} / " \
                                  f"volta {fmt_brl_full(g.get('frete_volta', 0) or 0)}"
                        if g.get("frete_obs"):
                            _fretes += f" ({g['frete_obs']})"
                        def _fdt(s):
                            try:
                                return datetime.strptime(str(s)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
                            except Exception:
                                return "—"
                        _rast = ""
                        if g.get("rastreio_entrada") or g.get("rastreio_saida"):
                            _rast = (f" | **Rastreio:** entrada {g.get('rastreio_entrada') or '—'} / "
                                     f"saída {g.get('rastreio_saida') or '—'}")
                        st.markdown(f"**Causa:** {g.get('diagnostico_causa') or '—'} | "
                                    f"**Serviço:** {g.get('diagnostico_obs') or '—'} | "
                                    f"**Resultado:** {g.get('resultado') or '—'} | "
                                    f"**NF saída:** {g.get('nf_saida') or '—'} | "
                                    f"**Chegada:** {_fdt(g.get('data_chegada'))} | "
                                    f"**Envio:** {_fdt(g.get('data_envio'))}{_rast} | "
                                    f"**Fretes:** {_fretes} | "
                                    f"**Custo do caso:** {fmt_brl_full(g.get('custo_total', 0) or 0)}")
                        if g.get("pecas"):
                            st.markdown("**Peças:** " + "; ".join(
                                f"{p.get('qtd',1)}x {p.get('nome','')}" for p in g["pecas"]))
                        st.info("🔒 Garantia finalizada. Correções, reabertura ou cancelamento: "
                                "somente o master da garantia (Jacson) ou o admin.")
                        if g.get("historico"):
                            st.caption(" ➤ " + " | ".join(f"{h['em']} {h['por']}: {h['acao']}"
                                                          for h in g["historico"][-4:]))
                        continue

                    if _fechada and can_edit_garantia_fechada():
                        st.warning("🔓 Modo master: esta garantia está finalizada — alterações aqui "
                                   "reabrem/corrigem o registro e ficam no histórico.")
                    # "Cancelada" só aparece para master/admin (cancelar = exclusão lógica)
                    _status_opcoes = STATUS_GARANTIA if can_edit_garantia_fechada() \
                        else [s for s in STATUS_GARANTIA if s != "Cancelada"]
                    with st.form(f"gar_upd_{tk}_{g['id']}"):
                        c1, c2 = st.columns(2)
                        novo_status = c1.selectbox("Status", _status_opcoes,
                                                   index=_status_opcoes.index(g.get("status", "Aberta"))
                                                   if g.get("status", "Aberta") in _status_opcoes else 0,
                                                   key=f"st_{tk}_{g['id']}")
                        causa = c2.selectbox("Causa (diagnóstico)", CAUSAS_GARANTIA,
                                             index=CAUSAS_GARANTIA.index(g["diagnostico_causa"])
                                             if g.get("diagnostico_causa") in CAUSAS_GARANTIA else None,
                                             placeholder="Qual foi a causa real...", key=f"ca_{tk}_{g['id']}")
                        diag_obs = st.text_area("O que foi feito (diagnóstico/serviço) *",
                                                value=g.get("diagnostico_obs", ""),
                                                height=70, key=f"do_{tk}_{g['id']}",
                                                help="Obrigatório para salvar a partir do status "
                                                     "'Aguardando peça' (se já mexeu, conta o que fez).")
                        ce1, ce2 = st.columns(2)
                        empresa_nf = ce1.selectbox("Empresa da NF de entrada (ao chegar)",
                                                   EMPRESAS_NF,
                                                   index=EMPRESAS_NF.index(g["empresa_nf"])
                                                   if g.get("empresa_nf") in EMPRESAS_NF else None,
                                                   placeholder="Quem emitiu a entrada...",
                                                   key=f"emp_{tk}_{g['id']}")
                        nf_entrada = ce2.text_input("NF de entrada (ao chegar)",
                                                    value=g.get("nf_entrada", ""),
                                                    key=f"nfe_{tk}_{g['id']}")
                        def _pdate(s):
                            try:
                                return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
                            except Exception:
                                return None
                        cd1, cd2 = st.columns(2)
                        data_chegada = cd1.date_input("📬 Data de CHEGADA na empresa",
                                                      value=_pdate(g.get("data_chegada")),
                                                      format="DD/MM/YYYY", key=f"dc_{tk}_{g['id']}")
                        data_envio = cd2.date_input("🚚 Data de ENVIO (volta ao cliente)",
                                                    value=_pdate(g.get("data_envio")),
                                                    format="DD/MM/YYYY", key=f"de_{tk}_{g['id']}")
                        cr1, cr2 = st.columns(2)
                        prioridade_b = cr1.selectbox("Prioridade", PRIORIDADES_GARANTIA,
                                                     index=PRIORIDADES_GARANTIA.index(g["prioridade"])
                                                     if g.get("prioridade") in PRIORIDADES_GARANTIA else 0,
                                                     key=f"pr_{tk}_{g['id']}",
                                                     help="🔴 Urgente / 🟠 Alta furam a fila.")
                        rastreio_saida = cr2.text_input("🚚 Rastreamento de SAÍDA (volta ao cliente)",
                                                        value=g.get("rastreio_saida", ""),
                                                        key=f"rgs_{tk}_{g['id']}",
                                                        placeholder="Código do envio de volta...")
                        st.markdown("**Peças trocadas / serviços** (até 3 — peça puxa custo da Base Mãe; "
                                    "serviço usa o R$ digitado, deixe 0 se feito em casa):")
                        _serv_map = {"🛠️ SERVIÇO — Afiação": ("SERV-AFIACAO", "Afiação (serviço)"),
                                     "🛠️ SERVIÇO — Mão de obra": ("SERV-MAOOBRA", "Mão de obra (serviço)")}
                        _serv_por_sku = {v[0]: k for k, v in _serv_map.items()}
                        slot_opts = list(_serv_map.keys()) + prod_opts
                        pecas_atuais = g.get("pecas", [])
                        pecas_novas = []
                        for slot in range(3):
                            pc1, pc2, pc3 = st.columns([3, 1, 1])
                            atual = pecas_atuais[slot] if slot < len(pecas_atuais) else None
                            atual_opt = None
                            _custo_ini = 0.0
                            if atual:
                                _sku_alvo = str(atual.get("sku", "")).strip()
                                if _sku_alvo in _serv_por_sku:
                                    atual_opt = _serv_por_sku[_sku_alvo]
                                    _custo_ini = float(atual.get("custo", 0) or 0)
                                else:
                                    # casamento por SKU EXATO (prefixo colidiria em códigos parecidos)
                                    atual_opt = next((o for o in prod_opts if _sku_de(o) == _sku_alvo), None)
                            psel = pc1.selectbox(f"Peça/serviço {slot+1}", slot_opts,
                                                 index=slot_opts.index(atual_opt) if atual_opt else None,
                                                 placeholder="—", key=f"p{slot}_{tk}_{g['id']}")
                            pqtd = pc2.number_input("Qtd", 1, 99,
                                                    value=int(atual.get("qtd", 1)) if atual else 1,
                                                    key=f"q{slot}_{tk}_{g['id']}")
                            pcusto = pc3.number_input("R$ (serviço)", 0.0, 99999.0, value=_custo_ini,
                                                      key=f"pc{slot}_{tk}_{g['id']}",
                                                      help="Só vale para Afiação/Mão de obra. 0 = feito em casa. "
                                                           "Peça de catálogo usa o custo da Base Mãe.")
                            if psel in _serv_map:
                                _ssku, _snome = _serv_map[psel]
                                pecas_novas.append({"sku": _ssku, "nome": _snome,
                                                    "qtd": int(pqtd), "custo": float(pcusto)})
                            elif psel:
                                psku = _sku_de(psel)
                                pecas_novas.append({"sku": psku,
                                                    "nome": psel.split(" — ", 1)[1] if " — " in psel else psel,
                                                    "qtd": int(pqtd),
                                                    "custo": custo_map.get(psku, 0)})
                        c3, c4 = st.columns(2)
                        resultado = c3.selectbox("Resultado", RESULTADOS_GARANTIA,
                                                 index=RESULTADOS_GARANTIA.index(g["resultado"])
                                                 if g.get("resultado") in RESULTADOS_GARANTIA else None,
                                                 placeholder="Ao concluir...", key=f"re_{tk}_{g['id']}")
                        nf_saida = c4.text_input("NF de saída", value=g.get("nf_saida", ""),
                                                 key=f"nf_{tk}_{g['id']}")
                        c5, c6 = st.columns(2)
                        frete_vinda = c5.number_input("Frete VINDA R$ (exigido só p/ Concluir)", 0.0, 99999.0,
                                                      value=float(g.get("frete_vinda", 0) or 0),
                                                      key=f"fv_{tk}_{g['id']}")
                        frete_volta = c6.number_input("Frete VOLTA R$ (exigido só p/ Concluir)", 0.0, 99999.0,
                                                      value=float(g.get("frete_volta", 0) or 0),
                                                      key=f"fb_{tk}_{g['id']}")
                        frete_obs = st.text_input("Sem frete? Explique (retirada pessoal, cliente pagou...)",
                                                  value=g.get("frete_obs", ""), key=f"fo_{tk}_{g['id']}")
                        salvar = st.form_submit_button("💾 Salvar atualização", type="primary")
                    if salvar:
                        problemas = []
                        # a partir de Aguardando peça já houve trabalho na máquina:
                        # não se salva sem contar O QUE FOI FEITO
                        if novo_status in ("Aguardando peça", "Confirmado — aguardando R$ frete",
                                           "Concluída") and not diag_obs.strip():
                            problemas.append("O QUE FOI FEITO (diagnóstico/serviço)")
                        if novo_status in ("Confirmado — aguardando R$ frete", "Concluída"):
                            if not causa or not resultado:
                                problemas.append("CAUSA e RESULTADO")
                            if not empresa_nf or not nf_entrada.strip():
                                problemas.append("EMPRESA e Nº da NF de entrada (o produto já passou por aqui)")
                        if novo_status == "Concluída":
                            if (frete_vinda <= 0 or frete_volta <= 0) and not frete_obs.strip():
                                problemas.append("os FRETES de vinda e volta (se não houve frete, "
                                                 "explique no campo 'Sem frete?')")
                        if problemas:
                            st.error("⚠️ Para esse status, informe: " + " e ".join(problemas) + ". Nada foi salvo.")
                        else:
                            upd = {"status": novo_status, "diagnostico_causa": causa or "",
                                   "diagnostico_obs": diag_obs.strip(), "pecas": pecas_novas,
                                   "resultado": resultado or "", "nf_saida": nf_saida.strip(),
                                   "empresa_nf": empresa_nf or "", "nf_entrada": nf_entrada.strip(),
                                   "rastreio_saida": rastreio_saida.strip(),
                                   "prioridade": prioridade_b,
                                   "data_chegada": data_chegada.strftime("%Y-%m-%d") if data_chegada else "",
                                   "data_envio": data_envio.strftime("%Y-%m-%d") if data_envio else "",
                                   "frete_vinda": float(frete_vinda), "frete_volta": float(frete_volta),
                                   "frete_obs": frete_obs.strip()}
                            upd["custo_total"] = _garantia_custo_total({**g, **upd}, custo_map)
                            # trabalho termina no Confirmado (a Concluída pode vir semanas depois, só pelo frete)
                            if novo_status in ("Confirmado — aguardando R$ frete", "Concluída") \
                                    and not g.get("concluido_em"):
                                upd["concluido_em"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                            _acao = f"Status → {novo_status}"
                            if _fechada:
                                _acao = f"CORREÇÃO PÓS-FECHAMENTO (era {g.get('status')}): {_acao}"
                            if update_garantia(g["id"], upd, _acao):
                                st.success(f"✅ {g['id']} atualizada (custo do caso: "
                                           f"{fmt_brl_full(upd['custo_total'])}).")
                                st.rerun()
                            else:
                                st.error("Não consegui salvar no GitHub agora. Tente de novo em instantes.")
                    if g.get("historico"):
                        st.caption(" ➤ " + " | ".join(f"{h['em']} {h['por']}: {h['acao']}"
                                                      for h in g["historico"][-4:]))
                    if st.session_state.get("role") == "admin":
                        # exclusão REAL, só admin (nem o master tem): limpar simulações.
                        # Caso real → status Cancelada, nunca excluir.
                        cdel1, cdel2 = st.columns([4, 1])
                        _confdel = cdel1.checkbox(
                            f"🗑️ Confirmo EXCLUIR {g['id']} DEFINITIVAMENTE — sem rastro "
                            "(só p/ simulação/teste; caso real usa 'Cancelada')",
                            key=f"delck_{tk}_{g['id']}")
                        if cdel2.button("Excluir", key=f"delbt_{tk}_{g['id']}",
                                        disabled=not _confdel):
                            if delete_garantia(g["id"]):
                                st.session_state["gar_flash_del"] = \
                                    f"🗑️ {g['id']} excluída definitivamente."
                                st.rerun()
                            else:
                                st.error("Não consegui excluir no GitHub agora. Tente de novo.")

        # Cancelada = exclusão lógica: só master/admin/diretor enxergam.
        # Para Marcos/Pedro não há sub-aba Canceladas e 'Todas' as omite.
        _ve_canceladas = can_edit_garantia_fechada() or has_full_data_access()
        _sub_defs = [("⚡ Ativas", STATUS_ATIVOS, "atv"),
                     ("📬 Aguard. chegada", ["Aguardando chegada"], "ab"),
                     ("🔧 Em bancada", ["Em bancada"], "bc"),
                     ("📦 Aguard. peça", ["Aguardando peça"], "pc"),
                     ("🚚 Aguard. R$ frete", ["Confirmado — aguardando R$ frete"], "fr"),
                     ("✅ Concluídas", ["Concluída"], "co")]
        if _ve_canceladas:
            _sub_defs += [("🚫 Canceladas", ["Cancelada"], "cn"),
                          ("📚 Todas", STATUS_GARANTIA, "td")]
        else:
            _sub_defs += [("📚 Todas", [s for s in STATUS_GARANTIA if s != "Cancelada"], "td")]
        _listas = {tk: _filtra(sts) for _, sts, tk in _sub_defs}
        _rotmap = {tk: f"{rot} ({len(_listas[tk])})" for rot, _, tk in _sub_defs}
        _opts = [tk for _, _, tk in _sub_defs]
        # seletor com key em vez de st.tabs: a CONTAGEM no rótulo muda a cada
        # save, e st.tabs com rótulo dinâmico voltaria pra 1ª aba — widget com
        # estado preserva a fila selecionada (e renderiza SÓ ela: página leve)
        if hasattr(st, "segmented_control"):
            tk_sel = st.segmented_control("Fila", _opts, format_func=lambda t: _rotmap[t],
                                          default="atv", key="gar_subtab",
                                          label_visibility="collapsed")
        else:
            tk_sel = st.radio("Fila", _opts, format_func=lambda t: _rotmap[t],
                              horizontal=True, key="gar_subtab",
                              label_visibility="collapsed")
        tk_sel = tk_sel if tk_sel in _listas else "atv"  # desmarcado/estado velho → padrão
        _render_fila(_listas[tk_sel], tk_sel)

    # ---------------- PAINEL ----------------
    with tab_painel:
        if not garantias:
            st.info("Nenhuma garantia registrada ainda. Os indicadores nascem conforme o time registra.")
        else:
            atng = [g for g in garantias if g.get("status") in STATUS_ATIVOS]
            # "concluída operacionalmente" = serviço feito e enviado (mesmo aguardando R$ do frete)
            concl = [g for g in garantias
                     if g.get("status") in ("Confirmado — aguardando R$ frete", "Concluída")]
            for g in garantias:
                if not g.get("custo_total"):
                    g["custo_total"] = _garantia_custo_total(g, custo_map)
            mes_atual = datetime.now().strftime("%Y-%m")
            custo_mes = sum(g["custo_total"] for g in concl if str(g.get("concluido_em", "")).startswith(mes_atual))

            # Mesmas datas/validações dos cards. Não mistura bases diferentes
            # na média; prioriza chegada→envio, depois usa a base disponível.
            tempos = [_garantia_tempo_info(g) for g in concl]
            bases_tempo = {
                "chegada_envio": "⏱️ Tempo na empresa (chegada→envio)",
                "chegada_confirmacao": "⏱️ Tempo médio (chegada→confirmação)",
                "abertura_envio": "⏱️ Tempo médio (abertura→envio)",
                "abertura_confirmacao": "⏱️ Tempo médio (abertura→confirmação)",
            }
            base_tempo = next((base for base in bases_tempo
                               if any(t["base"] == base and t["dias"] is not None for t in tempos)), None)
            dias_validos = [t["dias"] for t in tempos
                           if t["base"] == base_tempo and t["dias"] is not None]
            _t_label = bases_tempo.get(base_tempo, "⏱️ Tempo médio (datas válidas)")
            _t_valor = f"{sum(dias_validos)/len(dias_validos):.0f} dias" if dias_validos else "—"
            k1, k2, k3, k4 = st.columns(4)
            k1.metric("🔴 Em aberto", len(atng))
            k2.metric(_t_label, _t_valor)
            k3.metric("💸 Custo no mês", fmt_brl(custo_mes))
            k4.metric("📋 Total histórico", len(garantias))
            st.caption(f"Prazo calculado com {len(dias_validos)} de {len(concl)} casos encerrados "
                       "operacionalmente. A média usa apenas casos com a mesma base de datas; "
                       "datas ausentes, inconsistentes ou futuras não entram.")
            st.divider()
            dfg = pd.DataFrame(garantias)
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**Defeitos mais comuns (Pareto)**")
                st.dataframe(dfg["defeito"].value_counts().reset_index().rename(
                    columns={"defeito": "Defeito", "count": "Casos"}),
                    use_container_width=True, hide_index=True)
            with col2:
                st.markdown("**Causa raiz (diagnóstico da bancada)**")
                causas = dfg[dfg["diagnostico_causa"] != ""]["diagnostico_causa"].value_counts()
                if len(causas):
                    st.dataframe(causas.reset_index().rename(
                        columns={"diagnostico_causa": "Causa", "count": "Casos"}),
                        use_container_width=True, hide_index=True)
                else:
                    st.caption("Ainda sem diagnósticos concluídos.")
            st.markdown("**Casos registrados × unidades vendidas** (todos os canais)")
            periodo_vendas = _garantia_periodo_vendas(meta)
            if not periodo_vendas:
                st.info("Relação indisponível: a referência de vendas não informa um período "
                        "e uma data de publicação válidos.")
            else:
                inicio_ref, fim_ref = periodo_vendas
                ultimo_instante = fim_ref - timedelta(minutes=1)
                st.caption(f"Vendas: {meta['periodo']} · Referência publicada em {meta['gerado_em']}. "
                           f"Casos registrados de {inicio_ref:%d/%m/%Y} até "
                           f"{ultimo_instante:%d/%m/%Y %H:%M}, sem canceladas.")
                st.caption("O mês final pode ser parcial. A publicação não comprova a cobertura "
                           "integral das vendas ou dos registros de assistência. Esta relação "
                           "não representa a taxa real de defeito dos produtos vendidos.")
                casos_ref = _garantias_no_periodo_vendas(garantias, periodo_vendas)
                if not casos_ref:
                    st.info("Nenhum caso não cancelado com data de registro válida nessa referência.")
                else:
                    # O SKU é a identidade: nomes antigos/variantes não podem
                    # dividir o numerador e repetir o mesmo denominador de vendas.
                    nomes_sku = {}
                    for _, p in products_df.iterrows():
                        if pd.notna(p.get("code")) and pd.notna(p.get("name")):
                            sku, nome = str(p["code"]).strip(), str(p["name"]).strip()
                            if sku and nome:
                                nomes_sku.setdefault(sku, nome)
                    for g in casos_ref:
                        sku = str(g.get("produto_sku") or "").strip()
                        nome = str(g.get("produto_nome") or "").strip()
                        if sku and nome:
                            nomes_sku.setdefault(sku, nome)
                    df_ref = pd.DataFrame(casos_ref)
                    df_ref["produto_sku"] = df_ref["produto_sku"].fillna("").astype(str).str.strip()
                    agg = df_ref.groupby("produto_sku").agg(
                        Casos=("id", "count"), Custo=("custo_total", "sum")).reset_index()
                    agg.insert(1, "produto_nome", agg["produto_sku"].map(nomes_sku).fillna("Sem nome informado"))
                    agg["Unidades vendidas"] = agg["produto_sku"].map(
                        lambda s: vendas_map.get(str(s).strip()))
                    agg["Casos / unidades"] = agg.apply(
                        lambda r: _garantia_relacao_vendas(r["Casos"], r["Unidades vendidas"]), axis=1)
                    agg = agg.sort_values("Casos", ascending=False).head(20)
                    disp = agg.rename(columns={"produto_sku": "SKU", "produto_nome": "Produto"})
                    disp["Casos / unidades"] = disp["Casos / unidades"].apply(
                        lambda v: f"{v:.1%}" if pd.notna(v) else "—")
                    show_money_table(disp, ["Custo"], use_container_width=True, hide_index=True,
                                     height=min(420, 35 * len(disp) + 38))
                    st.caption("— indica ausência de uma quantidade vendida válida para calcular a relação.")
            if len(concl):
                st.markdown("**Peças e serviços mais usados** (planejar reposição / carga da bancada)")
                cons = defaultdict(lambda: {"qtd": 0, "custo": 0.0})
                for g in garantias:
                    for p in g.get("pecas", []):
                        c = cons[f"{p.get('sku','')} — {p.get('nome','')[:40]}"]
                        c["qtd"] += p.get("qtd", 1)
                        c["custo"] += (p.get("qtd", 1) or 1) * (p.get("custo", 0) or 0)
                if cons:
                    dfp = pd.DataFrame([{"Peça": k, "Qtd usada": v["qtd"], "Custo": round(v["custo"], 2)}
                                        for k, v in cons.items()]).sort_values("Qtd usada", ascending=False)
                    show_money_table(dfp, ["Custo"], use_container_width=True, hide_index=True,
                                     height=min(350, 35 * len(dfp) + 38))
            flat = []
            for g in garantias:
                flat.append({"ID": g["id"], "Status": g.get("status"),
                             "Prioridade": g.get("prioridade", "Normal"),
                             "Canal": _rotulo_outro(g.get("canal"), g.get("canal_outro")),
                             "Cliente": g.get("cliente"), "SKU": g.get("produto_sku"),
                             "Produto": g.get("produto_nome"),
                             "Empresa NF": g.get("empresa_nf", ""),
                             "Data compra": g.get("data_compra", ""),
                             "Cliente final": g.get("cliente_final", ""),
                             "NF cliente final": g.get("cliente_final_nf", ""),
                             "Chave NF cliente final": g.get("cliente_final_nf_chave", ""),
                             "NF entrada": g.get("nf_entrada"), "NF saída": g.get("nf_saida"),
                             "Rastreio entrada": g.get("rastreio_entrada", ""),
                             "Rastreio saída": g.get("rastreio_saida", ""),
                             "Defeito": _rotulo_outro(g.get("defeito"), g.get("defeito_outro")),
                             "Relato": g.get("defeito_obs"),
                             "Causa": g.get("diagnostico_causa"), "Serviço": g.get("diagnostico_obs"),
                             "Data chegada": g.get("data_chegada", ""), "Data envio": g.get("data_envio", ""),
                             "Peças": "; ".join(f"{p.get('qtd',1)}x {p.get('nome','')}" for p in g.get("pecas", [])),
                             "Frete vinda": g.get("frete_vinda", 0), "Frete volta": g.get("frete_volta", 0),
                             "Sem frete (justif.)": g.get("frete_obs", ""),
                             "Custo total": g.get("custo_total", 0), "Resultado": g.get("resultado"),
                             "Entrada": g.get("criado_em"), "Concluída": g.get("concluido_em", ""),
                             "Registrado por": g.get("criado_por")})
            _csv_download(pd.DataFrame(flat), "⬇️ Baixar garantias deste perfil (Excel/CSV)",
                          "garantias.csv", "dl_gar")

# ============================================================
# PAGE: PAINEL DO GESTOR (tela inicial do admin/diretor)
# ============================================================
def _mv_num(v, dflt=0.0):
    """float à prova de json corrompido (blindagem obrigatória da Fase 2).
    Rejeita NaN/Infinity — json.load os aceita e st.progress/format estouram."""
    try:
        v = float(v)
        return v if math.isfinite(v) else dflt
    except (TypeError, ValueError):
        return dflt


def _mv_int(v, dflt=0):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return dflt


def _mv_brl(v):
    """fmt_brl seguro p/ st.markdown/st.caption: dois 'R$' na mesma linha
    viram fórmula LaTeX no markdown do Streamlit — escapar o cifrão resolve."""
    return fmt_brl(v).replace("$", "\\$")


def _mes_vivo_bloco_vendedor(v, dia, dias_mes, mes_ant_nome):
    """Bloco de um vendedor na página Mês ao Vivo: progresso vs meta + ritmo."""
    nome = str(v.get("nome", "?")).replace(" Propetz Distribuição", "") \
                                  .replace(" La Maison Propetz", "")
    receita = _mv_num(v.get("receita"))
    meta = _mv_num(v.get("meta"))
    proj = receita / max(dia, 1) * dias_mes
    if meta > 0:
        pct = receita / meta
        st.markdown(f"**{nome}** — {_mv_brl(receita)} de {_mv_brl(meta)} "
                    f"(**{pct * 100:.0f}%** da meta) · projeção {_mv_brl(proj)} "
                    f"({proj / meta * 100:.0f}% da meta)")
        st.progress(max(0.0, min(pct, 1.0)))
        desb = _mv_num(v.get("meta_desbloqueio"))
        extra = []
        if desb > 0:
            extra.append("🔓 comissão desbloqueada" if receita >= desb else
                         f"desbloqueio da comissão: {_mv_brl(desb)} "
                         f"(faltam {_mv_brl(desb - receita)})")
    else:
        st.markdown(f"**{nome}** — {_mv_brl(receita)} *(sem meta cadastrada)*")
        extra = []
    extra.append(f"{_mv_int(v.get('notas'))} nota(s) · "
                 f"{_mv_int(v.get('clientes'))} cliente(s)")
    _dev_v = _mv_num(v.get("devolucoes"))
    if _dev_v > 0:
        extra.append(f"↩️ devoluções abatidas: {_mv_brl(_dev_v)}")
    ant_ate = _mv_num(v.get("anterior_ate_dia"))
    if ant_ate > 0 and mes_ant_nome:
        extra.append(f"vs mesmo dia de {mes_ant_nome}: "
                     f"{(receita / ant_ate - 1) * 100:+.0f}%")
    st.caption(" · ".join(extra))


def _mes_vivo_tabela_clientes(top):
    dt = pd.DataFrame([{
        "Cliente": str(c.get("nome", "")),
        "Vendedor": str(c.get("vendedor", "")).split(" ")[0].title(),
        "UF": str(c.get("uf", "")),
        "Notas": _mv_int(c.get("notas")),
        "Receita no Mês": _mv_num(c.get("receita")),
    } for c in top])
    show_money_table(dt, ["Receita no Mês"], use_container_width=True,
                     hide_index=True, height=min(400, 35 * len(dt) + 38))


def page_mes_vivo():
    st.header("🔴 Mês ao Vivo")
    mv = load_silver_mes_vivo()
    # BLINDAGEM (padrão Fase 2): json ausente/inesperado NUNCA derruba o app
    if not (isinstance(mv, dict) and isinstance(mv.get("total"), dict)
            and mv.get("mes_nome")):
        st.info("📡 Os dados ao vivo ainda não foram publicados pela rotina do "
                "banco. Eles chegam de hora em hora — volte em instantes.")
        return
    total = mv["total"]
    ant = mv.get("anterior") if isinstance(mv.get("anterior"), dict) else {}
    vendedores = [v for v in (mv.get("vendedores") or []) if isinstance(v, dict)]
    por_dia = [x for x in (mv.get("por_dia") or [])
               if isinstance(x, dict) and x.get("d")]
    top = [c for c in (mv.get("top_clientes") or []) if isinstance(c, dict)]
    dia = max(_mv_int(mv.get("dia"), 1), 1)
    dias_mes = max(_mv_int(mv.get("dias_no_mes"), 30), 1)
    mes_curto = str(mv.get("mes_nome", "")).split("/")[0]
    st.caption(f"**{esc(str(mv.get('mes_nome', '')))}**, dia {dia} de {dias_mes} — "
               f"direto do banco (carga de hora em hora; após o dia 25, a cada "
               f"30 min). Última publicação: {esc(str(mv.get('gerado_em', '?')))}. "
               "A nota não carrega hora — o dia cresce a cada carga. "
               "**Critério: receita LÍQUIDA** — vendas menos devoluções do mês, "
               "sem IPI e sem frete (regra 06/08).")

    # FRESCOR (auditoria 06/08): rotina parada não pode passar por "ao vivo" —
    # mês virou sem publicação nova, ou última publicação velha ⇒ aviso claro
    if str(mv.get("mes", "")) != datetime.now().strftime("%Y-%m"):
        st.warning(f"⚠️ Estes números são de **{esc(str(mv.get('mes_nome', '?')))}** — "
                   "a rotina do PC ainda não publicou o mês corrente. "
                   "Não use como 'ao vivo'.")
    else:
        try:
            _idade_h = (datetime.now() - datetime.strptime(
                str(mv.get("gerado_em", "")), "%Y-%m-%d %H:%M")).total_seconds() / 3600
            if _idade_h > 26:
                st.warning(f"⚠️ Última publicação há **{_idade_h / 24:.0f} dia(s)** — "
                           "a rotina do PC não está rodando; os números abaixo "
                           "estão defasados.")
        except ValueError:
            pass

    # ---- vendedor: SÓ os números dele. Gate por PAPEL (auditoria 06/08:
    # vendor_filter vazio não pode liberar a visão completa/comissões) ----
    if not has_full_data_access():
        vf = st.session_state.get("vendor_filter")
        if not vf:
            st.info("Seu usuário está sem carteira configurada — peça ao admin "
                    "para definir o vendedor no seu cadastro.")
            return
        # match pelo 1º nome (mesmo critério do resto da integração): não
        # quebra se o rótulo da planilha de metas for renomeado
        _vf1 = str(vf).split(" ")[0].lower()
        meus = [v for v in vendedores
                if str(v.get("nome", "")).split(" ")[0].lower() == _vf1]
        if not meus:
            st.info(f"Ainda não há vendas suas registradas em {mes_curto}.")
            return
        _mes_vivo_bloco_vendedor(meus[0], dia, dias_mes, ant.get("mes_nome"))
        meus_cli = [c for c in top
                    if str(c.get("vendedor", "")).split(" ")[0].lower() == _vf1]
        if meus_cli:
            st.subheader(f"Seus clientes em {mes_curto}")
            _mes_vivo_tabela_clientes(meus_cli)
        return

    # ---- visão completa (admin/diretor) ----
    receita = _mv_num(total.get("receita"))
    ant_ate = _mv_num(ant.get("receita_ate_dia"))
    ant_total = _mv_num(ant.get("receita_total"))
    proj = receita / max(dia, 1) * dias_mes
    k1, k2, k3, k4 = st.columns(4)
    k1.metric(f"Receita de {mes_curto}", fmt_brl(receita),
              (f"{(receita / ant_ate - 1) * 100:+.0f}% vs mesmo dia de "
               f"{ant.get('mes_nome')}" if ant_ate > 0 else None))
    k2.metric("Projeção do mês", fmt_brl(proj),
              (f"{(proj / ant_total - 1) * 100:+.0f}% vs {ant.get('mes_nome')} "
               f"fechado" if ant_total > 0 else None),
              help="Projeção linear: receita até agora ÷ dias corridos × dias do mês.")
    k3.metric("Notas no mês", f"{_mv_int(total.get('notas'))}")
    k4.metric("Clientes atendidos", f"{_mv_int(total.get('clientes'))}")
    _dev_mes = _mv_num(total.get("devolucoes"))
    if _dev_mes > 0:
        st.caption(f"↩️ Devoluções já abatidas do mês: {_mv_brl(_dev_mes)}")

    st.divider()
    st.subheader("🎯 Meta por vendedor")
    st.caption("Metas oficiais do plano 2026 (fonte: dash da TV, projeto Demanda).")
    for v in vendedores:
        _mes_vivo_bloco_vendedor(v, dia, dias_mes, ant.get("mes_nome"))

    if por_dia:
        st.divider()
        st.subheader("📈 Receita por dia")
        dias_x = [str(x["d"])[-2:] for x in por_dia]
        vals = [_mv_num(x.get("r")) for x in por_dia]
        acum, s = [], 0.0
        for r in vals:
            s += r
            acum.append(round(s, 2))
        fig = go.Figure()
        fig.add_bar(x=dias_x, y=vals, name="No dia", marker_color="#3b82f6")
        fig.add_scatter(x=dias_x, y=acum, name="Acumulado", mode="lines+markers",
                        line=dict(color="#FF6B35", width=3))
        fig.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10),
                          legend=dict(orientation="h", y=1.1),
                          xaxis_title=f"Dia de {mes_curto}")
        st.plotly_chart(fig, use_container_width=True)

    if top:
        st.divider()
        st.subheader(f"🏆 Top 30 clientes de {mes_curto}")
        if len(top) > 30:
            st.caption(f"Os 30 maiores de {len(top)} clientes atendidos no mês.")
        _mes_vivo_tabela_clientes(top[:30])


def _commercial_active_mask(df, inactive_ids):
    """Carteira ativa: status Ativo na planilha e sem inativação no app."""
    inactive_ids = {str(cid).strip() for cid in inactive_ids}
    return df['status'].eq('Ativo') & ~df['id'].astype(str).str.strip().isin(inactive_ids)


def _commercial_reactivation_candidates(df, inactive_ids):
    """O app só desfaz sua própria inativação; status da planilha exige correção na fonte."""
    inactive_ids = {str(cid).strip() for cid in inactive_ids}
    return df[df['status'].eq('Ativo')
              & df['id'].astype(str).str.strip().isin(inactive_ids)].copy()


def _commercial_period_recurrence(monthly, selected_indices):
    """Conta compras e meses dentro do mesmo período selecionado."""
    indices = {i for i in selected_indices if 0 <= i < len(monthly)}
    return sum(1 for i in indices if monthly[i] > 0), len(indices)


def _commercial_estimate_caption():
    st.caption("Estimativa anual = média dos meses com compra nos últimos 12 meses × 12; "
               "sem compras nesses 12 meses, usa a média histórica dos meses com compra. "
               "Pressupõe compra mensal; pode superar o faturamento realizado no ano.")


def page_manager(df, months, df_sku, products_df):
    st.header("🎛️ Painel do Gestor")
    st.caption(f"Acompanhamento objetivo do canal Distribuição — dados até **{months[-1]}**.")

    _n_pend = len(pending_inactivation_requests())
    if _n_pend > 0:
        st.warning(f"📋 **{_n_pend} solicitação(ões) de inativação** aguardando sua aprovação — "
                   f"veja a seção *Solicitações de Inativação* abaixo.")

    n = len(months)
    monthly_tot = [df['monthly'].apply(lambda m: m[i] if i < len(m) else 0).sum() for i in range(n)]

    # ---- LINHA 1: O MÊS E O ANO, SEM RODEIO ----
    last = monthly_tot[-1]
    prev = monthly_tot[-2] if n >= 2 else 0
    yoy = monthly_tot[n - 13] if n >= 13 else 0
    last3 = sum(monthly_tot[-3:])
    prev3 = sum(monthly_tot[-6:-3]) if n >= 6 else 0

    cur_year = _year_of_label(months[-1])
    prev_year = str(int(cur_year) - 1) if cur_year.isdigit() else ''
    ytd_idx = [i for i in range(n) if _year_of_label(months[i]) == cur_year]
    py_idx = [i for i in range(n) if _year_of_label(months[i]) == prev_year][:len(ytd_idx)]
    ytd = sum(monthly_tot[i] for i in ytd_idx)
    pytd = sum(monthly_tot[i] for i in py_idx)

    def _pct(cur, base):
        return f"{(cur - base) / base * 100:+.1f}%" if base > 0 else None

    k1, k2, k3, k4 = st.columns(4)
    k1.metric(f"Receita {months[-1]}", fmt_brl(last), _pct(last, prev),
              help="Variação vs mês anterior")
    k2.metric("vs Mesmo Mês Ano Passado", _pct(last, yoy) or "—", f"{fmt_brl(yoy)} em {months[n-13]}" if n >= 13 else "")
    k3.metric("Últimos 3 Meses", fmt_brl(last3), _pct(last3, prev3),
              help="Variação vs trimestre móvel anterior")
    k4.metric(f"Acumulado {cur_year}", fmt_brl(ytd), _pct(ytd, pytd),
              help=f"vs mesmo período de {prev_year}")

    # Concentração de receita (últimos 12 meses)
    df_conc = df.copy()
    df_conc['r12'] = df_conc['monthly'].apply(lambda m: sum(m[-12:]))
    r12_total = df_conc['r12'].sum()
    top5 = df_conc.nlargest(5, 'r12')['r12'].sum()
    if r12_total > 0:
        st.caption(f"⚖️ Concentração: os 5 maiores clientes respondem por **{top5/r12_total*100:.0f}%** "
                   f"da receita dos últimos 12 meses.")

    st.divider()

    # ---- LINHA 2: DESEMPENHO POR VENDEDOR ----
    st.subheader("👥 Desempenho por Vendedor")
    st.caption(f"Mês de referência: {months[-1]}. Tendência = mês atual vs média dos 3 meses anteriores. "
               "Carteira = clientes ATIVOS (inativados ficam fora da conta).")
    # carteira ativa: mesmo recorte do quadro de recuperações logo abaixo —
    # sem isso, cliente inativado seguia inflando o nº de clientes do vendedor
    _df_cart = df[_commercial_active_mask(df, load_inactive_clients())].copy()
    rows = []
    for v, g in _df_cart.groupby('vendor'):
        if not v:
            continue
        rev_m = g['monthly'].apply(lambda m: m[-1] if len(m) >= 1 else 0).sum()
        base3 = g['monthly'].apply(lambda m: sum(m[-4:-1])).sum() / 3 if n >= 4 else 0
        buyers = len(g[g['monthly'].apply(lambda m: m[-1] > 0 if len(m) >= 1 else False)])
        risk_rs = g[g['risk'].isin(['Recuperação', 'Atenção'])]['monthly'].apply(annual_value_estimate).sum()
        rows.append({
            'Vendedor': str(v).replace(' Propetz Distribuição', '').replace(' La Maison Propetz', ''),
            'Receita no Mês': round(rev_m, 2),
            'Média 3m Anteriores': round(base3, 2),
            'Tendência': f"{(rev_m / base3 - 1) * 100:+.0f}%" if base3 > 0 else '—',
            'Compraram no Mês': f"{buyers}/{len(g)}",
            'Cobertura': f"{buyers / len(g) * 100:.0f}%" if len(g) > 0 else '—',
            'R$ em Risco (ano)': round(risk_rs, 2),
        })
    if rows:
        vend_df = pd.DataFrame(rows).sort_values('Receita no Mês', ascending=False)
        show_money_table(vend_df, ['Receita no Mês', 'Média 3m Anteriores', 'R$ em Risco (ano)'],
                         use_container_width=True, hide_index=True,
                         height=min(350, 35 * len(vend_df) + 38))
    _commercial_estimate_caption()

    st.divider()

    # ---- LINHA 3: ONDE AGIR AGORA ----
    st.subheader("🚨 Maiores Recuperações em Jogo")
    st.caption("Top 10 clientes ativos esfriando, da base inteira, por receita anual em jogo — cobre isso nas reuniões com o time.")
    risky = _df_cart[_df_cart['risk'].isin(['Recuperação', 'Atenção'])].copy()
    risky['_cid'] = risky['id'].astype(str).str.strip()
    if len(risky) > 0:
        _busca_m = st.text_input("🔍 Buscar cliente (nome, UF, vendedor ou código)", key="mgr_risky_search",
                                 placeholder="Digite parte do nome para achar qualquer cliente em risco...")
        _pend_ids_m = {r['client_id'] for r in pending_inactivation_requests()}
        risky['valor'] = risky['monthly'].apply(annual_value_estimate)
        risky['Prioridade'] = risky['risk'].map({'Recuperação': '🔴 Urgente', 'Atenção': '🟡 Atenção'})
        risky.loc[risky['_cid'].isin(_pend_ids_m), 'Prioridade'] = \
            risky.loc[risky['_cid'].isin(_pend_ids_m), 'Prioridade'] + ' ⏳'
        risky = risky.sort_values('valor', ascending=False)
        # Sem busca: top 10. Com busca: TODOS os que casam (da base inteira em risco).
        if _busca_m and str(_busca_m).strip():
            risky = _filter_clients_by_term(risky, _busca_m).reset_index(drop=True)
            _nota_m = f"{len(risky)} resultado(s) para “{_busca_m}”."
        else:
            risky = risky.head(10).reset_index(drop=True)
            _nota_m = "Top 10 por valor — use a busca acima para achar qualquer outro cliente em risco."
        if len(risky) == 0:
            st.info("Nenhum cliente encontrado com esse termo.")
        else:
            disp_r = risky[['Prioridade', 'name', 'vendor', 'state', 'last_purchase', 'months_since', 'valor']].copy()
            disp_r['vendor'] = disp_r['vendor'].str.replace(' Propetz Distribuição', '').str.replace(' La Maison Propetz', '')
            disp_r.columns = ['Prioridade', 'Cliente', 'Vendedor', 'UF', 'Última Compra', 'Meses', 'R$ em Jogo (ano)']
            ev_r = show_money_table(disp_r, ['R$ em Jogo (ano)'], use_container_width=True, hide_index=True,
                                    height=min(400, 35 * len(disp_r) + 38),
                                    on_select="rerun", selection_mode="multi-row", key="mgr_risky_sel")
            _act = "Marque a caixinha na linha para inativar" if can_approve_inactivations() \
                else "Cliente fechou? Marque a caixinha na linha para sugerir a inativação"
            st.caption(f"☑️ {_act}. ⏳ = já há uma solicitação pendente.")
            try:
                _rows = list(ev_r.selection.rows) if ev_r and ev_r.selection and ev_r.selection.rows else []
            except Exception:
                _rows = []
            # Protege contra seleção "presa" de um rerun anterior apontando linhas que já não existem
            _rows = [i for i in _rows if 0 <= i < len(risky)]
            if _rows:
                _clientes = [{'cid': rw['_cid'], 'name': rw['name'], 'vendor': rw['vendor']}
                             for _, rw in risky.iloc[_rows].iterrows()]
                _inativacao_form(_clientes, "mgr_risky")
            st.caption(_nota_m)
    else:
        st.success("Nenhum cliente ativo em risco no momento.")

    st.divider()

    # ---- LINHA 4: SOLICITAÇÕES DE INATIVAÇÃO ----
    # Admin aprova/rejeita; diretora vê a fila como informativa (ela sugere, não decide).
    _can_approve = can_approve_inactivations()
    st.subheader("📋 Solicitações de Inativação")
    reqs = load_inactive_requests()
    pend = [r for r in reqs if r.get('status') == 'pendente']
    if not pend:
        st.caption("Nenhuma solicitação pendente.")
    elif not _can_approve:
        st.caption("Solicitações aguardando aprovação do administrador. Você pode sugerir novas "
                   "inativações nas tabelas abaixo e nas páginas de Ações e Churn.")
        for r in pend:
            _vend_short = r.get('vendor', '').replace(' Propetz Distribuição', '').replace(' La Maison Propetz', '')
            _mot = _fmt_motivo(r)
            st.markdown(f"⏳ **{r['client_name']}** ({_vend_short}) — solicitado por "
                        f"*{r.get('requested_by_name', '?')}* em {r.get('date', '')}" +
                        (f"  \n&nbsp;&nbsp;&nbsp;💬 {_mot}" if _mot else ""))
    else:
        st.caption("Aprovando, o cliente sai das listas de ação e churn de todo o time. "
                   "Rejeitando, ele continua como está.")
        for idx, r in enumerate(pend):
            c1, c2, c3 = st.columns([6, 1.2, 1.2])
            _vend_short = r.get('vendor', '').replace(' Propetz Distribuição', '').replace(' La Maison Propetz', '')
            _mot = _fmt_motivo(r)
            c1.markdown(f"**{r['client_name']}** ({_vend_short}) — solicitado por "
                        f"*{r.get('requested_by_name', '?')}* em {r.get('date', '')}" +
                        (f"  \n💬 **Motivo:** {_mot}" if _mot else "  \n💬 *sem motivo informado*"))
            if c2.button("✅ Aprovar", key=f"apr_{r['client_id']}_{idx}"):
                if decide_inactivation_request(r['client_id'], True, st.session_state.get('user_name', '')):
                    st.rerun()
                else:
                    st.error("Não consegui salvar no GitHub agora. Tente de novo em instantes.")
            if c3.button("❌ Rejeitar", key=f"rej_{r['client_id']}_{idx}"):
                if decide_inactivation_request(r['client_id'], False, st.session_state.get('user_name', '')):
                    st.rerun()
                else:
                    st.error("Não consegui salvar no GitHub agora. Tente de novo em instantes.")

    # ---- Histórico / banco de motivos das inativações ----
    if reqs:
        with st.expander(f"📚 Histórico de inativações e motivos ({len(reqs)} registros)"):
            hist = pd.DataFrame([{
                'Cliente': r.get('client_name', ''),
                'Vendedor': r.get('vendor', '').replace(' Propetz Distribuição', '').replace(' La Maison Propetz', ''),
                'Motivo': r.get('motivo', '') or '(não informado)',
                'Observação': r.get('observacao', ''),
                'Status': r.get('status', ''),
                'Solicitado por': r.get('requested_by_name', ''),
                'Data solicitação': r.get('date', ''),
                'Decidido por': r.get('decidido_por', ''),
                'Data decisão': r.get('decidido_em', ''),
            } for r in reversed(reqs)])
            st.dataframe(hist, use_container_width=True, hide_index=True,
                         height=min(420, 35 * len(hist) + 38))
            # Resumo: quantas inativações APROVADAS por motivo
            aprov = hist[hist['Status'] == 'aprovado']
            if len(aprov) > 0:
                st.markdown("**Resumo dos motivos (inativações já aprovadas):**")
                resumo = aprov['Motivo'].value_counts().reset_index()
                resumo.columns = ['Motivo', 'Clientes inativados']
                st.dataframe(resumo, use_container_width=True, hide_index=True)
            if _can_approve:
                _csv_download(hist, "⬇️ Baixar histórico de motivos (Excel/CSV)",
                              "historico_inativacoes.csv", "dl_hist_inativ")

    # Reativação de clientes inativados (somente admin)
    inact_ids = load_inactive_clients()
    _df_inact = df[df['id'].astype(str).str.strip().isin(inact_ids)]
    with st.expander(f"♻️ Clientes inativados no app ({len(_df_inact)})" + (" — reativar" if _can_approve else "")):
        if len(_df_inact) == 0:
            st.caption("Nenhum cliente inativado.")
        else:
            st.caption("Só retorna à carteira quem está Ativo na planilha. Outros status "
                       "precisam ser corrigidos na fonte antes da reativação pelo administrador.")
            st.dataframe(_df_inact[['id', 'name', 'vendor', 'state', 'status']].rename(
                columns={'id': 'Código', 'name': 'Cliente', 'vendor': 'Vendedor',
                         'state': 'UF', 'status': 'Status na Planilha'}),
                use_container_width=True, hide_index=True)
            if not _can_approve:
                st.caption("Reativação é feita pelo administrador.")
            else:
                _react = _commercial_reactivation_candidates(_df_inact, inact_ids).sort_values('name')
                _react_labels = {str(r['id']).strip(): f"{r['name']} · {str(r['id']).strip()}"
                                 for _, r in _react.iterrows()}
                if _react_labels:
                    sel_react = st.multiselect("Selecione para reativar:", list(_react_labels),
                                               format_func=lambda cid: _react_labels.get(cid, cid), key="mgr_react")
                    _ids = [cid for cid in sel_react if cid in _react_labels]
                    if _ids and st.button(f"♻️ Reativar {len(_ids)} cliente(s)", key="btn_mgr_react", type="primary"):
                        if reactivate_clients(_ids):
                            st.rerun()
                        else:
                            st.error("Não consegui salvar no GitHub agora. Tente de novo em instantes.")

    st.divider()

    # ---- LINHA 5: O TIME ESTÁ USANDO O BI? ----
    st.subheader("📡 Uso do BI pelo Time (últimos 14 dias)")
    access_log = _load_access_log()
    df_log = pd.DataFrame(access_log) if access_log else pd.DataFrame()
    if df_log.empty or 'action' not in df_log.columns or 'date' not in df_log.columns:
        st.info("Sem registros de acesso ainda. (Para o log sobreviver a reinícios do servidor, configure o GITHUB_TOKEN — ver COMO-USAR.md.)")
    else:
        cutoff = (datetime.now() - timedelta(days=14)).strftime('%Y-%m-%d')
        recent = df_log[(df_log['action'] == 'login') & (df_log['date'] >= cutoff)]
        users_all = load_users().get('users', {})
        urows = []
        for uname, info in users_all.items():
            if info.get('role') == 'admin':
                continue
            u_logins = recent[recent['user'] == uname]
            last_seen = df_log[df_log['user'] == uname]['date'].max() if len(df_log[df_log['user'] == uname]) > 0 else None
            urows.append({
                'Usuário': info.get('name', uname),
                'Logins (14d)': len(u_logins),
                'Último Acesso': last_seen or 'Nunca',
                'Situação': '✅ Ativo' if len(u_logins) >= 3 else ('🟡 Pouco uso' if len(u_logins) >= 1 else '🔴 Sem uso'),
            })
        if urows:
            st.dataframe(pd.DataFrame(urows).sort_values('Logins (14d)', ascending=False),
                         use_container_width=True, hide_index=True)
            n_inactive = sum(1 for r in urows if r['Logins (14d)'] == 0)
            if n_inactive > 0:
                st.warning(f"{n_inactive} usuário(s) sem nenhum login nos últimos 14 dias — ferramenta não vira resultado se o time não usa.")

# ============================================================
# PAGE: MINHAS AÇÕES (tela inicial — lista priorizada de trabalho)
# ============================================================
def _csv_download(df_export, label, filename, key):
    """Botão de download CSV no padrão Excel brasileiro (; e vírgula decimal)."""
    csv_bytes = df_export.to_csv(index=False, sep=';', decimal=',').encode('utf-8-sig')
    st.download_button(label, csv_bytes, file_name=filename, mime="text/csv", key=key)

def _filter_clients_by_term(df, term, name_col='name', vendor_col='vendor', state_col='state', id_col='id'):
    """Filtra um DataFrame de clientes por nome, UF, vendedor ou código.
    Termo vazio devolve o df inteiro. Busca sem acento/maiúscula."""
    if not term or not str(term).strip():
        return df
    s = str(term).strip().lower()
    mask = None
    for col in (name_col, vendor_col, state_col, id_col):
        if col and col in df.columns:
            m = df[col].astype(str).str.lower().str.contains(s, na=False, regex=False)
            mask = m if mask is None else (mask | m)
    return df[mask] if mask is not None else df

def _inativacao_form(clientes, key_prefix):
    """Formulário com MOTIVO + observação POR cliente. `clientes` é uma lista de
    dicts {cid, name, vendor}. Admin inativa direto (registro 'aprovado' com motivo);
    vendedor/diretora cria solicitação 'pendente' com motivo. st.rerun ao concluir.
    Tudo vira histórico em inactive_requests.json (banco de motivos)."""
    if not clientes:
        return
    pode = can_approve_inactivations()
    verbo = "Inativar agora" if pode else "Enviar solicitação"
    st.caption("✍️ Informe o **motivo** de cada cliente — vira histórico para o gestor analisar.")
    with st.form(f"{key_prefix}_motform"):
        entradas = []
        for c in clientes:
            st.markdown(f"**{c['name']}**")
            col1, col2 = st.columns([1, 1.3])
            # index=None: sem motivo pré-selecionado — força escolha consciente
            # (senão o banco enche de "Fechou" defaultado sem ninguém ter escolhido)
            mot = col1.selectbox("Motivo", MOTIVOS_INATIVACAO, index=None,
                                 placeholder="Selecione o motivo...",
                                 key=f"{key_prefix}_mot_{c['cid']}", label_visibility="collapsed")
            obs = col2.text_input("Observação", key=f"{key_prefix}_obs_{c['cid']}",
                                  placeholder="Detalhe (opcional)", label_visibility="collapsed")
            entradas.append((c, mot, obs))
        enviar = st.form_submit_button(f"✅ {verbo} ({len(clientes)})", type="primary")
    if enviar:
        _faltando = [c['name'] for c, mot, obs in entradas if not mot]
        if _faltando:
            st.error("⚠️ Escolha o motivo de: " + ", ".join(_faltando[:5]) +
                     ("…" if len(_faltando) > 5 else "") + ". Nenhuma inativação foi registrada.")
            return
        n = 0
        for c, mot, obs in entradas:
            if add_inactivation_request(c['cid'], c['name'], c['vendor'],
                                        motivo=mot, observacao=obs, direct_approve=pode):
                n += 1
        falhas = len(entradas) - n
        if falhas:
            # NÃO dá rerun: deixa o aviso visível e o formulário aberto para tentar de novo
            if n:
                st.success(f"{n} registrado(s).")
            st.warning(f"⚠️ {falhas} não confirmado(s): pode ser solicitação já pendente, ou falha de "
                       "conexão com o GitHub. Se foi conexão, **clique de novo** — o sistema completa o "
                       "que faltou, sem duplicar.")
        else:
            if pode:
                st.success(f"{n} cliente(s) inativado(s) com motivo registrado.")
            else:
                st.success(f"{n} solicitação(ões) enviada(s) ao administrador, com motivo.")
            st.rerun()

def page_actions(df, df_sku, products_df, df_client_products, months):
    st.header("✅ Minhas Ações")
    st.caption(f"Sua lista de trabalho priorizada — dados até {months[-1]}. "
               "Comece pelo topo: são os contatos com mais receita em jogo.")

    work = df.copy()

    # Admin/diretor escolhem a carteira; vendedor já chega filtrado
    if has_full_data_access():
        vendors = ["Todas"] + sorted(work['vendor'].unique().tolist())
        sel_v = st.selectbox("Carteira", vendors, key="act_vendor")
        if sel_v != "Todas":
            work = work[work['vendor'] == sel_v]

    # Mesmo recorte de carteira ativa usado no Gestor e no Churn.
    inactive_ids = load_inactive_clients()
    work['_cid'] = work['id'].astype(str).str.strip()
    work = work[_commercial_active_mask(work, inactive_ids)].copy()
    st.caption("Carteira ativa = status Ativo na planilha e sem inativação no app.")

    if len(work) == 0:
        st.info("Nenhum cliente ativo na carteira selecionada.")
        return

    work['valor_anual'] = work['monthly'].apply(annual_value_estimate)
    work['vendor_short'] = work['vendor'].str.replace(' Propetz Distribuição', '').str.replace(' La Maison Propetz', '')

    # ---- Produtos favoritos por cliente (o que ele sempre comprou) ----
    fav = {}
    if len(df_sku) > 0:
        g = df_sku.groupby(['cod_cliente', 'produto'])['quantidade'].sum().reset_index()
        g['cod_cliente'] = g['cod_cliente'].astype(str).str.strip()
        for cid, grp in g.groupby('cod_cliente'):
            fav[cid] = ', '.join(grp.nlargest(3, 'quantidade')['produto'].tolist())

    # ---- 1. CONTATOS PRIORITÁRIOS (clientes esfriando, por receita em jogo) ----
    calls = work[work['risk'].isin(['Recuperação', 'Atenção'])].copy()
    calls = calls.sort_values('valor_anual', ascending=False)

    # ---- 2. OFERTAS PRONTAS (produtos Curva A que o cliente ainda não compra) ----
    offers = []
    if len(df_client_products) > 0 and len(products_df) > 0:
        carteira_ids = set(work['_cid'])
        cp = df_client_products.copy()
        cp['client_id'] = cp['client_id'].astype(str).str.strip()
        cp = cp[cp['client_id'].isin(carteira_ids)]
        n_cart = max(cp['client_id'].nunique(), 1)
        pen = cp.groupby('product_code')['client_id'].nunique() / n_cart * 100
        bought_by = cp.groupby('client_id')['product_code'].apply(set).to_dict()
        prod_a = products_df[products_df['abc'] == 'A']
        _preco = _preco_medio_map(products_df)
        _typ, _nbuy, _ = _sku_stats(df_sku)

        # Foca nos 30 clientes ativos mais valiosos da carteira
        targets = work.sort_values('valor_anual', ascending=False).head(30)
        for _, cl in targets.iterrows():
            owned = bought_by.get(cl['_cid'], set())
            if not owned:
                continue  # sem histórico de produto para cruzar
            for _, pr in prod_a.iterrows():
                code = pr['code']
                if code in owned:
                    continue
                p = float(pen.get(code, 0))
                if p >= 30:
                    pot = round((_typ.get(code) or 0) * (_preco.get(code) or 0), 2)
                    offers.append({
                        'Cliente': cl['name'],
                        'Vendedor': cl['vendor_short'],
                        'Produto Sugerido': pr['name'],
                        'Categoria': pr['category'],
                        '% da Carteira que Compra': round(p),
                        'R$ Potencial/Mês': pot,
                        '_valor_cliente': cl['valor_anual'],
                        '_score': p * max(cl['valor_anual'], 1),
                    })
    offers_df = pd.DataFrame(offers)
    if len(offers_df) > 0:
        offers_df = offers_df.sort_values(['R$ Potencial/Mês', '_score'], ascending=False)

    # ---- KPIs ----
    n_urgent = len(calls[calls['risk'] == 'Recuperação'])
    n_warn = len(calls[calls['risk'] == 'Atenção'])
    at_stake = calls['valor_anual'].sum()
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("🔴 Contatos Urgentes", f"{n_urgent}", "6+ meses sem comprar")
    k2.metric("🟡 Contatos de Atenção", f"{n_warn}", "3-5 meses sem comprar")
    k3.metric("💰 Receita em Jogo", fmt_brl(at_stake), "estimativa anual")
    k4.metric("🎯 Ofertas Identificadas", f"{len(offers_df)}", "produtos Curva A")
    _commercial_estimate_caption()

    st.divider()

    # ---- SEÇÃO 1: LIGAR / VISITAR ----
    st.subheader("📞 Quem contatar primeiro")
    if len(calls) == 0:
        st.success("Nenhum cliente esfriando na carteira. Foque nas ofertas abaixo!")
    else:
        _busca = st.text_input("🔍 Buscar cliente na lista (nome, UF, vendedor ou código)", key="calls_search",
                               placeholder="Digite parte do nome para encontrar qualquer cliente da lista...")
        _calls_f = _filter_clients_by_term(calls, _busca)
        # Sem busca: mostra os 20 maiores. Com busca: mostra TODOS os que casam.
        if _busca and str(_busca).strip():
            top_calls = _calls_f.copy()
            _nota = f"{len(top_calls)} resultado(s) para “{_busca}”."
        else:
            top_calls = calls.head(20).copy()
            _nota = (f"Mostrando os 20 maiores de {len(calls)} clientes esfriando — "
                     "use a busca acima para achar qualquer outro." if len(calls) > 20 else "")
        if len(top_calls) == 0:
            st.info("Nenhum cliente encontrado com esse termo.")
        else:
            _pend_ids = {r['client_id'] for r in pending_inactivation_requests()}
            top_calls['Prioridade'] = top_calls['risk'].map({'Recuperação': '🔴 Urgente', 'Atenção': '🟡 Atenção'})
            # ⏳ = já existe solicitação de inativação aguardando aprovação
            top_calls.loc[top_calls['_cid'].isin(_pend_ids), 'Prioridade'] = \
                top_calls.loc[top_calls['_cid'].isin(_pend_ids), 'Prioridade'] + ' ⏳'
            top_calls['Sempre Comprou'] = top_calls['_cid'].map(fav).fillna('—')
            top_calls = top_calls.reset_index(drop=True)
            disp_calls = top_calls[['Prioridade', 'name', 'vendor_short', 'state', 'last_purchase',
                                    'months_since', 'valor_anual', 'Sempre Comprou']].copy()
            disp_calls.columns = ['Prioridade', 'Cliente', 'Vendedor', 'UF', 'Última Compra',
                                  'Meses sem Comprar', 'Valor Anual Est.', 'Sempre Comprou']
            ev = show_money_table(disp_calls, ['Valor Anual Est.'], use_container_width=True, hide_index=True,
                                  height=min(500, 35 * len(disp_calls) + 38),
                                  on_select="rerun", selection_mode="multi-row", key="calls_sel")
            st.caption("☑️ Cliente fechou ou não existe mais? Marque a caixinha à esquerda da linha "
                       "e use o botão que aparece abaixo. ⏳ = solicitação já enviada.")
            try:
                _sel_rows = list(ev.selection.rows) if ev and ev.selection and ev.selection.rows else []
            except Exception:
                _sel_rows = []
            # Protege contra seleção "presa" de um rerun anterior apontando linhas que já não existem
            _sel_rows = [i for i in _sel_rows if 0 <= i < len(top_calls)]
            if _sel_rows:
                _clientes = [{'cid': rw['_cid'], 'name': rw['name'], 'vendor': rw['vendor']}
                             for _, rw in top_calls.iloc[_sel_rows].iterrows()]
                _inativacao_form(_clientes, "act_calls")
            if _nota:
                st.caption(_nota)

        export_calls = calls[['name', 'vendor_short', 'state', 'risk', 'last_purchase',
                              'months_since', 'valor_anual']].copy()
        export_calls.columns = ['Cliente', 'Vendedor', 'UF', 'Risco', 'Última Compra',
                                'Meses sem Comprar', 'Valor Anual Estimado']
        export_calls['Sempre Comprou'] = calls['_cid'].map(fav).fillna('')
        _csv_download(export_calls, "⬇️ Baixar lista de contatos (Excel/CSV)",
                      "contatos_prioritarios.csv", "dl_calls")

    # ---- SOLICITAR INATIVAÇÃO (vendedor/diretora pedem, admin aprova) ----
    if not can_approve_inactivations():
        with st.expander("🚫 Solicitar inativação de cliente FORA da lista acima"):
            st.caption("Para os clientes da tabela, use a caixinha de seleção na própria linha. "
                       "A solicitação vai para aprovação do administrador — até lá, o cliente continua nas listas.")
            _pend = pending_inactivation_requests()
            _pend_ids = {r['client_id'] for r in _pend}
            _opts = work[~work['_cid'].isin(_pend_ids)].sort_values('name')['name'].tolist()
            sel_inact = st.multiselect("Clientes para inativar:", _opts, key="req_inact")
            if sel_inact:
                _clientes = []
                for nm in sel_inact:
                    row = work[work['name'] == nm]
                    if len(row) > 0:
                        _clientes.append({'cid': row.iloc[0]['_cid'], 'name': nm, 'vendor': row.iloc[0]['vendor']})
                _inativacao_form(_clientes, "act_fora")
            _mine = [r for r in _pend if r.get('requested_by') == st.session_state.get('username')]
            if _mine:
                st.info("⏳ Aguardando aprovação: " + ", ".join(r['client_name'] for r in _mine))

    st.divider()

    # ---- SEÇÃO 2: OFERTAS PRONTAS ----
    st.subheader("🎯 Ofertas prontas para os seus melhores clientes")
    st.caption("Produtos Curva A (os que mais faturam no canal Distribuição) que pelo menos 30% da carteira compra, "
               "mas estes clientes ainda não — argumento de venda pronto.")
    if len(offers_df) == 0:
        st.info("Nenhuma oferta identificada — clientes principais já compram os produtos Curva A relevantes, "
                "ou não há dados de produto por cliente na planilha.")
    else:
        disp_offers = offers_df[['Cliente', 'Vendedor', 'Produto Sugerido', 'Categoria',
                                 '% da Carteira que Compra', 'R$ Potencial/Mês']].head(40)
        show_money_table(disp_offers, ['R$ Potencial/Mês'], use_container_width=True, hide_index=True,
                         height=min(500, 35 * len(disp_offers) + 38))
        _csv_download(offers_df.drop(columns=['_valor_cliente', '_score']),
                      "⬇️ Baixar lista de ofertas (Excel/CSV)",
                      "ofertas_mix.csv", "dl_offers")

# ============================================================
# PAGE: AGENDA COMERCIAL
# ============================================================
AGENDA_FILE = os.path.join(os.path.dirname(__file__), 'agenda_comercial.json')


def _agenda_now():
    return datetime.now(ZoneInfo('America/Sao_Paulo'))


def _agenda_local_state():
    if not os.path.exists(AGENDA_FILE):
        return {'schema_version': 1, 'clientes': {}}
    with open(AGENDA_FILE, encoding='utf-8') as file:
        state = json.load(file)
    agenda.validate_state(state)
    return state


def _agenda_remote_state(token):
    """GET estrito: erro remoto nunca vira agenda vazia nem fallback local."""
    url = f'{_GH_API}/repos/{_GH_REPO}/contents/agenda_comercial.json'
    headers = _gh_headers(token)
    try:
        response = requests.get(url, params={'ref': _GH_STATE_BRANCH}, headers=headers, timeout=20)
        if response.status_code == 404:
            # GitHub tambem usa 404 para repo sem acesso. Somente um branch
            # acessivel confirma que este arquivo realmente ainda nao existe.
            branch = requests.get(f'{_GH_API}/repos/{_GH_REPO}/git/ref/heads/{_GH_STATE_BRANCH}',
                                  headers=headers, timeout=20)
            if branch.status_code != 200:
                raise ValueError('Não foi possível confirmar o acesso à agenda no servidor.')
            state, sha = {'schema_version': 1, 'clientes': {}}, None
        elif response.status_code == 200:
            metadata = response.json()
            if not isinstance(metadata, dict) or not isinstance(metadata.get('sha'), str) or not metadata['sha']:
                raise ValueError('A referência do histórico comercial está inválida no servidor.')
            sha = metadata['sha']
            if metadata.get('encoding') == 'none':
                raw_headers = {**headers, 'Accept': 'application/vnd.github.raw+json'}
                raw = requests.get(url, params={'ref': _GH_STATE_BRANCH}, headers=raw_headers, timeout=60)
                if raw.status_code != 200:
                    raise ValueError('Não foi possível ler o histórico comercial completo.')
                content = raw.content
            elif metadata.get('encoding') == 'base64' and isinstance(metadata.get('content'), str):
                content = base64.b64decode(metadata['content'].replace('\n', '').replace('\r', ''), validate=True)
            else:
                raise ValueError('O formato do histórico comercial está inválido no servidor.')
            state = json.loads(content.decode('utf-8'))
            agenda.validate_state(state)
        else:
            raise ValueError('A agenda no servidor está indisponível. Tente novamente; nenhum registro foi substituído.')
    except requests.RequestException:
        raise ValueError('Não foi possível conectar à agenda no servidor. Tente novamente.') from None
    except (UnicodeError, KeyError, TypeError, ValueError) as error:
        raise ValueError('Não foi possível validar a agenda no servidor. O histórico foi preservado.') from error
    local = _agenda_local_state()
    remote_events = {event['id']: (cid, event) for cid, record in state['clientes'].items()
                     for event in record['historico']}
    unpublished = any(remote_events.get(event['id']) != (cid, event)
                      for cid, record in local['clientes'].items() for event in record['historico'])
    if unpublished:
        raise ValueError('Há histórico comercial local que ainda não está no servidor. '
                         'Solicite a conciliação antes de registrar novos contatos; nada foi substituído.')
    return state, sha


def _agenda_write_local(state):
    """Troca atomica; erros chegam ao chamador, sem falso sucesso no modo local."""
    temp_path = AGENDA_FILE + '.' + uuid.uuid4().hex + '.tmp'
    try:
        with open(temp_path, 'w', encoding='utf-8') as file:
            json.dump(state, file, ensure_ascii=False, separators=(',', ':'))
            file.flush()
            os.fsync(file.fileno())
        os.replace(temp_path, AGENDA_FILE)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


def load_agenda():
    token = _gh_token()
    return _agenda_remote_state(token)[0] if token else _agenda_local_state()


def _agenda_authorized_clients():
    if not st.session_state.get('authenticated') or _session_expired():
        raise ValueError('Entre novamente para registrar o contato.')
    error = _refresh_session_access()
    if error:
        raise ValueError(error)
    if st.session_state.get('role') not in ('admin', 'diretor', 'vendedor'):
        raise ValueError('Seu perfil não possui acesso à agenda comercial.')
    source = load_data()[0]
    if source is None:
        raise ValueError('A base de clientes está indisponível.')
    scoped = _clients_for_access(source, st.session_state.get('role'), st.session_state.get('vendor_filter'))
    return scoped[_commercial_active_mask(scoped, load_inactive_clients())].copy()


def save_agenda_contact(client_id, *, expected_version, event_id, channel, outcome,
                        note, next_action, return_date, closed):
    """Revalida carteira na gravação e nunca anuncia sucesso sem persistir."""
    allowed = _agenda_authorized_clients()
    cid = str(client_id).strip()
    if cid not in set(allowed['id'].astype(str).str.strip()):
        raise ValueError('Esse cliente não está na sua carteira ativa. Atualize a página.')
    now = _agenda_now()
    def apply(state):
        # Revalida tambem em cada retry, apos a leitura remota potencialmente lenta.
        current = _agenda_authorized_clients()
        if cid not in set(current['id'].astype(str).str.strip()):
            raise ValueError('Esse cliente não está mais na sua carteira ativa. Atualize a página.')
        return agenda.register_contact(state, client_id=cid, actor=st.session_state['username'], channel=channel,
            outcome=outcome, note=note, next_action=next_action, return_date=return_date,
            closed=closed, expected_version=expected_version, event_id=event_id, now=now)

    token = _gh_token()
    with _GH_WRITE_LOCK:
        if token:
            confirmed = None
            for _ in range(5):
                state, sha = _agenda_remote_state(token)
                new_state = apply(state)
                if new_state == state:
                    confirmed = state  # mesmo UUID ja persistido: nao duplica o evento.
                    break
                body = json.dumps(new_state, ensure_ascii=False, separators=(',', ':')).encode('utf-8')
                done, status = _gh_put_file_status('agenda_comercial.json', body,
                    'Estado: agenda_comercial.json', _GH_STATE_BRANCH, sha, token)
                if done:
                    confirmed = new_state
                    break
                if status not in (409, 422):
                    # O servidor pode ter salvo antes de a conexao cair. Uma nova
                    # leitura confirma o mesmo evento sem criar outro UUID.
                    try:
                        check, _ = _agenda_remote_state(token)
                        if apply(check) == check:
                            confirmed = check
                    except (ValueError, OSError):
                        pass
                    break
            if confirmed is None:
                raise ValueError('Não foi possível confirmar o salvamento. Seu preenchimento foi mantido; tente novamente.')
            try:
                _agenda_write_local(confirmed)
            except OSError:
                pass  # remoto ja confirmou: falha no cache local nao invalida o contato.
            _STATE_RAW_CACHE.pop('agenda_comercial.json', None)
            return 'Contato salvo com confirmação no servidor.'
        _agenda_write_local(apply(_agenda_local_state()))
        _STATE_RAW_CACHE.pop('agenda_comercial.json', None)
    return 'Contato salvo apenas neste servidor. Sem persistência remota, pode ser perdido no reinício do Cloud.'


def page_agenda(df, months):
    if (not st.session_state.get('authenticated') or _session_expired()
            or st.session_state.get('role') not in ('admin', 'diretor', 'vendedor')):
        st.error('A agenda está disponível para o time comercial.')
        return
    # Defesa também na entrada da página; não depende apenas da navegação.
    try:
        scoped = _clients_for_access(df, st.session_state.get('role'), st.session_state.get('vendor_filter'))
        state = load_agenda()
    except ValueError as error:
        st.error(str(error))
        return
    except OSError:
        st.error('Não foi possível ler o histórico local da agenda. Contate o administrador.')
        return
    today = _agenda_now().date()
    ui.page_hero('COMERCIAL / HOJE', 'Hoje, sua próxima ação.',
                 'Sua carteira, com uma próxima ação para cada oportunidade.',
                 f'{today:%d/%m/%Y} · Base mensal até {months[-1] if months else "—"}')
    if '_agenda_notice' in st.session_state:
        st.success(st.session_state.pop('_agenda_notice'))
    if not _gh_token():
        st.warning('Modo local: os registros ficam apenas neste servidor e podem ser perdidos no reinício do Cloud.')

    work = scoped[_commercial_active_mask(scoped, load_inactive_clients())].copy()
    if has_full_data_access():
        options = ['Todas as carteiras'] + _vendor_options(work)
        vendor = st.selectbox('Acompanhar carteira', options, key='agenda_vendor')
        if vendor != options[0]:
            work = work[work['vendor'].astype(str).str.strip() == vendor].copy()
    if work.empty:
        st.info('Nenhum cliente ativo nesta carteira. Confira o cadastro e os filtros.')
        return
    work['id'] = work['id'].astype(str).str.strip()
    work['valor_anual'] = work['monthly'].apply(annual_value_estimate)
    items = agenda.build_agenda(work.to_dict('records'), state, today)
    counts = {category: sum(item['category'] == category for item in items)
              for category in ('Atrasados', 'Hoje', 'Recuperação', 'Atenção', 'Programados')}
    ui.stats_grid(list(zip(
            ['Retornos atrasados', 'Combinados para hoje', 'Contatos em recuperação', 'Retornos programados'],
            [counts['Atrasados'], counts['Hoje'], counts['Recuperação'], counts['Programados']],
            ['Retome estes compromissos', 'Sua agenda do dia', 'Sem retorno já agendado', 'Próximos dias'],
            ['red', 'teal', 'amber', 'neutral'])))
    st.caption('Sugestões usam a régua atual de risco e a carteira ativa. Retornos combinados têm preferência; '
               'encerrar um acompanhamento não inativa o cliente.')

    left, right = st.columns([1.15, 1], gap='large')
    with left:
        st.subheader('Onde agir agora')
        queue_filter = st.radio('Mostrar', ['Prioridades', 'Retornos', 'Programados'], horizontal=True,
                                key='agenda_queue', label_visibility='collapsed')
        search = st.text_input('Buscar na agenda', placeholder='Nome, código ou vendedor', key='agenda_search')
        category_map = {'Prioridades': {'Atrasados', 'Hoje', 'Recuperação', 'Atenção'},
                        'Retornos': {'Atrasados', 'Hoje'}, 'Programados': {'Programados'}}
        visible = [item for item in items if item['category'] in category_map[queue_filter]]
        if search.strip():
            term = search.strip().casefold()
            visible = [item for item in visible if term in ' '.join(str(item.get(k, '')) for k in ('name', 'cid', 'vendor')).casefold()]
        if not visible:
            st.info('Nenhum contato neste filtro. Você pode escolher qualquer cliente da carteira ao lado.')
        else:
            pages = max(1, math.ceil(len(visible) / 5))
            page_number = st.selectbox('Página da agenda', list(range(1, pages + 1)),
                                       format_func=lambda n: f'{n} de {pages}', key='agenda_page') if pages > 1 else 1
            st.caption(f'{len(visible)} contato(s) · até 5 por página')
            for item in visible[(page_number - 1) * 5:page_number * 5]:
                with st.container(border=True):
                    due = f" · {date.fromisoformat(item['due_date']):%d/%m}" if item.get('due_date') else ''
                    st.markdown(f"<div class='agenda-kicker'>{esc(item['category'] + due)}</div>"
                                f"<div class='agenda-client'>{esc(item['name'])}</div>", unsafe_allow_html=True)
                    st.caption(f"{item['cid']} · {item['vendor']}")
                    st.write(item['reason'])
                    st.markdown(f"<div class='agenda-next'>{esc(item['suggested_action'])}</div>", unsafe_allow_html=True)
                    if st.button('Abrir cliente →', key=f"agenda_open_{item['cid']}", use_container_width=True):
                        st.session_state['agenda_client'] = item['cid']

    with right:
        st.subheader('Conversa e próximo passo')
        clients = work.drop_duplicates('id').set_index('id')
        ids = sorted(clients.index, key=lambda cid: (str(clients.loc[cid, 'name']), cid))
        current_client = st.session_state.get('agenda_client')
        # Streamlit 1.41 refaz o widget ao mudar opções: reafirma o ID ainda válido.
        st.session_state['agenda_client'] = current_client if current_client in ids else (items[0]['cid'] if items else ids[0])
        cid = st.selectbox('Cliente da carteira', ids,
                          format_func=lambda value: f"{clients.loc[value, 'name']} · {value}", key='agenda_client')
        client = clients.loc[cid]
        record = state['clientes'].get(cid, {})
        version = record.get('version', 0)
        prefix = f'agenda_contact_{cid}'
        version_key = prefix + '_version'
        event_key = prefix + '_event'
        st.session_state.setdefault(version_key, version)
        st.session_state.setdefault(event_key, str(uuid.uuid4()))
        with st.container(border=True):
            st.markdown(f"<div class='agenda-client'>{esc(client['name'])}</div>", unsafe_allow_html=True)
            st.caption(f"Código {cid} · {client.get('state', '—')} · {client['vendor']}")
            a, b = st.columns(2)
            a.metric('Situação comercial', str(client['risk']))
            b.metric('Realizado · últimos 12 meses', fmt_brl(sum(client['monthly'][-12:])))
            st.caption(f"Última compra: {client.get('last_purchase', '—')}. "
                       'O realizado acima usa a base mensal; a recência pode incluir a atualização diária.')
            if record.get('encerrado'):
                st.info('Acompanhamento encerrado. Registre um novo contato com retorno para retomá-lo.')
            elif record.get('retorno_em'):
                st.info(f"Próximo retorno: {date.fromisoformat(record['retorno_em']):%d/%m/%Y} · {record['proxima_acao']}")

        saved_event = any(event['id'] == st.session_state[event_key] for event in record.get('historico', []))
        conflict = st.session_state[version_key] != version and not saved_event
        if saved_event:
            st.success('O contato anterior já está confirmado no histórico. Ele não será registrado novamente.')
            if st.button('Iniciar outro registro', key=prefix + '_restart'):
                for key in list(st.session_state):
                    if key == prefix or str(key).startswith(prefix + '_'):
                        del st.session_state[key]
                st.rerun()
        if conflict:
            st.warning('Há um contato mais recente para este cliente. Revise o histórico antes de salvar.')
            if st.button('Revisei: usar histórico atualizado', key=prefix + '_refresh'):
                st.session_state[version_key] = version
                st.session_state[event_key] = str(uuid.uuid4())
                st.rerun()
        with st.form(prefix, clear_on_submit=False):
            st.markdown('**Registrar contato**')
            channel_col, result_col = st.columns(2)
            channel = channel_col.selectbox('Canal do contato', agenda.CHANNELS, key=prefix + '_channel')
            outcome = result_col.selectbox('Resultado', agenda.OUTCOMES, index=None,
                                           placeholder='Selecione o resultado', key=prefix + '_outcome')
            note = st.text_area('Resumo da conversa', max_chars=2000, height=95,
                                placeholder='O que foi conversado, interesse e objeções.', key=prefix + '_note')
            next_action = st.text_input('Próxima ação', max_chars=300,
                                        placeholder='Ex.: retornar sobre a proposta de lâminas', key=prefix + '_action')
            return_date = st.date_input('Próximo retorno', value=today + timedelta(days=1),
                                        min_value=today, format='DD/MM/YYYY', key=prefix + '_date')
            closed = st.checkbox('Encerrar este acompanhamento', key=prefix + '_closed',
                                  help='Retira este cliente da agenda automática até um novo registro com retorno. Não altera seu cadastro.')
            st.caption('Para manter o acompanhamento, informe a próxima ação e a data. '
                       'Ao encerrar, esses dois campos são desconsiderados. Pedido informado é um registro manual.')
            if st.form_submit_button('Salvar contato', type='primary', use_container_width=True, disabled=conflict or saved_event):
                try:
                    message = save_agenda_contact(cid, expected_version=st.session_state[version_key],
                        event_id=st.session_state[event_key], channel=channel, outcome=outcome, note=note,
                        next_action=next_action, return_date=return_date, closed=closed)
                except (ValueError, OSError) as error:
                    st.error(str(error))
                else:
                    for key in list(st.session_state):
                        if key == prefix or str(key).startswith(prefix + '_'):
                            del st.session_state[key]
                    st.session_state['_agenda_notice'] = message
                    st.rerun()
        with st.expander(f"Histórico de contatos · {len(record.get('historico', []))}", expanded=True):
            history = record.get('historico', [])
            if not history:
                st.caption('O primeiro contato registrado aparecerá aqui.')
            for event in reversed(history[-10:]):
                when = datetime.fromisoformat(event['em'])
                st.markdown(f"**{when:%d/%m/%Y %H:%M} · {event['resultado']}**")
                st.caption(f"{event['user']} · {event['canal']}")
                if event.get('observacao'):
                    st.text(event['observacao'])
                if event.get('retorno_em'):
                    st.caption(f"Retorno em {date.fromisoformat(event['retorno_em']):%d/%m/%Y}: {event['proxima_acao']}")
                elif event.get('encerrado'):
                    st.caption('Acompanhamento encerrado neste registro.')
                st.divider()
            if len(history) > 10:
                st.caption('Exibindo os 10 contatos mais recentes. O histórico completo está na exportação abaixo.')
            if history:
                export = pd.DataFrame(history).drop(columns=['id'], errors='ignore')
                # Observações livres devem abrir como texto no Excel, nunca fórmulas.
                export = export.map(lambda value: "'" + value if isinstance(value, str)
                    and value.lstrip().startswith(('=', '+', '-', '@')) else value)
                _csv_download(export, 'Baixar histórico deste cliente', f'contatos_{cid}.csv', prefix + '_export')

# ============================================================
# PAGE: VISÃO GERAL
# ============================================================
def page_overview(df, months, year_ranges, sel_indices, sel_indices_sorted, sel_months):
    st.header("📊 Visão Geral")

    n_sel_months = len(sel_months)

    # --- FILTERS (Vendedor, Estado, Status) ---
    fc1, fc2, fc3 = st.columns(3)
    vendors = ["Todos"] + sorted(df['vendor'].unique().tolist())
    states = ["Todos"] + sorted(df['state'].unique().tolist())
    statuses = ["Todos"] + sorted(df['status'].unique().tolist())

    with fc1:
        sel_vendor = st.selectbox("Vendedor", vendors, key="ov_vendor")
    with fc2:
        sel_state = st.selectbox("Estado", states, key="ov_state")
    with fc3:
        sel_status = st.selectbox("Status", statuses, key="ov_status")

    # Apply client filters
    filtered = df.copy()
    if sel_vendor != "Todos":
        filtered = filtered[filtered['vendor'] == sel_vendor]
    if sel_state != "Todos":
        filtered = filtered[filtered['state'] == sel_state]
    if sel_status != "Todos":
        filtered = filtered[filtered['status'] == sel_status]

    # Helper: sum monthly values in selected period (uses sel_indices set)
    def period_sum(m):
        return sum(m[i] for i in sel_indices_sorted if i < len(m))

    # Helper: sum monthly values for a specific set of indices
    def range_sum_set(m, idx_set):
        return sum(m[i] for i in idx_set if i < len(m))

    # --- KPIs ---
    period_rev = filtered['monthly'].apply(period_sum).sum()
    # status EFETIVO: inativado pelo app conta como Inativo, mesmo que a planilha
    # ainda diga Ativo (a planilha envelhece entre uploads; o app é a verdade)
    _inat_app = filtered['id'].astype(str).str.strip().isin(load_inactive_clients())
    n_active = len(filtered[(filtered['status'] == 'Ativo') & ~_inat_app])
    n_inactive = len(filtered[(filtered['status'] == 'Inativo') | _inat_app])
    n_risk = len(filtered[filtered['risk'].isin(['Recuperação', 'Atenção']) & ~_inat_app])

    # Buyers in period = clients with any revenue > 0 in selected months
    period_buyers = len(filtered[filtered['monthly'].apply(
        lambda m: any(m[i] > 0 for i in sel_indices_sorted if i < len(m))
    )])
    avg_ticket = period_rev / period_buyers / max(n_sel_months, 1) if period_buyers > 0 else 0

    # Compare with same-length previous period (shift selected indices back by n_sel_months)
    prev_indices = set(max(0, i - n_sel_months) for i in sel_indices_sorted)
    prev_indices = prev_indices - sel_indices  # remove overlap
    prev_rev = filtered['monthly'].apply(lambda m: range_sum_set(m, prev_indices)).sum()
    period_growth = ((period_rev - prev_rev) / prev_rev * 100) if prev_rev > 0 else 0

    # Retention: of clients who bought in prev period, how many bought in current period
    if len(prev_indices) > 0:
        bought_prev = set(filtered[filtered['monthly'].apply(
            lambda m: any(m[i] > 0 for i in prev_indices if i < len(m))
        )].index)
        bought_curr = set(filtered[filtered['monthly'].apply(
            lambda m: any(m[i] > 0 for i in sel_indices_sorted if i < len(m))
        )].index)
        retained = len(bought_prev & bought_curr)
        retention_rate = (retained / len(bought_prev) * 100) if len(bought_prev) > 0 else 0
    else:
        retention_rate = 0
        retained = 0
        bought_prev = set()
        bought_curr = set()

    k1, k2, k3, k4, k5 = st.columns(5)
    period_label = f"{sel_months[0]} - {sel_months[-1]}" if len(sel_months) > 1 else sel_months[0] if sel_months else ""
    k1.metric("Receita do Período", fmt_brl(period_rev), period_label)
    k2.metric("vs Período Anterior", f"{'↑' if period_growth >= 0 else '↓'} {abs(period_growth):.1f}%",
              f"{fmt_brl(prev_rev)} anterior")
    k3.metric("Retenção", f"{retention_rate:.0f}%",
              f"{retained} de {len(bought_prev)} retidos")
    k4.metric("Base Ativa", f"{n_active}", f"Inativos: {n_inactive} | Risco: {n_risk}")
    k5.metric("Ticket Médio/Mês", fmt_brl(avg_ticket), f"{period_buyers} comprando no período")

    st.divider()

    # Insights aparecem AQUI no topo; o conteúdo é calculado ao longo da
    # página e renderizado neste container no final.
    st.subheader("🧠 Insights Automáticos")
    insights_box = st.container()

    st.divider()

    # --- REVENUE TREND (clickable bar chart with CTRL+click) ---
    monthly_totals = []
    for i in range(len(months)):
        total = filtered['monthly'].apply(lambda m: m[i] if i < len(m) else 0).sum()
        monthly_totals.append(total)

    # Bars: selected period = bold color, others = subtle
    bar_colors = ['#3b82f6' if i in sel_indices else 'rgba(59,130,246,0.2)' for i in range(len(months))]

    fig_rev = go.Figure()
    fig_rev.add_trace(go.Bar(
        x=months, y=monthly_totals, name='Receita',
        marker_color=bar_colors,
        hovertemplate='%{x}: R$ %{y:,.0f}<extra></extra>',
        selectedpoints=None
    ))
    fig_rev.update_layout(
        title="Receita Mensal (R$) — clique para filtrar | CTRL+clique para adicionar | arraste para selecionar",
        height=400,
        template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
        yaxis=dict(gridcolor='#e2e8f0'),
        xaxis=dict(gridcolor='#e2e8f0'),
        showlegend=False,
        clickmode='event+select',
        dragmode='select',
        selectdirection='h'
    )

    # Use on_select for box/lasso selection (drag)
    event = st.plotly_chart(fig_rev, use_container_width=True, on_select="rerun",
                            selection_mode=["points", "box", "lasso"], key="rev_chart_sel")

    # Process chart selection (click or drag)
    if event and event.selection and event.selection.points:
        clicked_labels = []
        for pt in event.selection.points:
            x_val = pt.get("x", None) if isinstance(pt, dict) else getattr(pt, 'x', None)
            if x_val and x_val in months:
                clicked_labels.append(x_val)
        # Only update if selection changed
        current_chart_sel = set(st.session_state.get("chart_sel_months", []))
        new_sel = set(clicked_labels)
        if new_sel and new_sel != current_chart_sel:
            st.session_state["chart_sel_months"] = list(new_sel)
            st.rerun()

    # --- Interactive month grid below chart ---
    chart_sel_current = st.session_state.get("chart_sel_months", [])

    with st.expander("Seleção individual de meses (clique para selecionar/remover)", expanded=False):
        # Group months by year for compact display
        _year_groups = {}
        for i, lbl in enumerate(months):
            parts = lbl.replace('-', '/').split('/')
            yr = parts[-1].strip() if len(parts) >= 2 else '??'
            _year_groups.setdefault(yr, []).append(lbl)

        for yr, yr_months in _year_groups.items():
            yr_label = f"20{yr}" if len(yr) == 2 else yr
            n_m = len(yr_months)
            cols = st.columns([0.6] + [1] * n_m + [1] * max(0, 12 - n_m))
            cols[0].markdown(f"**{yr_label}**")
            for j, m_lbl in enumerate(yr_months):
                is_selected = m_lbl in chart_sel_current or (not chart_sel_current and months.index(m_lbl) in sel_indices)
                btn_label = m_lbl.split('/')[0] if '/' in m_lbl else m_lbl
                with cols[j + 1]:
                    if is_selected:
                        if st.button(f"**{btn_label}**", key=f"mbtn_{m_lbl}", use_container_width=True, type="primary"):
                            new_sel = [x for x in chart_sel_current if x != m_lbl]
                            if new_sel:
                                st.session_state["chart_sel_months"] = new_sel
                            else:
                                if "chart_sel_months" in st.session_state:
                                    del st.session_state["chart_sel_months"]
                            st.rerun()
                    else:
                        if st.button(btn_label, key=f"mbtn_{m_lbl}", use_container_width=True):
                            if chart_sel_current:
                                st.session_state["chart_sel_months"] = chart_sel_current + [m_lbl]
                            else:
                                st.session_state["chart_sel_months"] = [m_lbl]
                            st.rerun()

    st.divider()

    # --- RECEITA NOVA vs RECORRENTE ---
    col_nr, col_ret = st.columns(2)

    with col_nr:
        # For each month in the period, classify revenue as new/reactivated vs recurring
        new_rev_list = []
        rec_rev_list = []
        for i in sel_indices_sorted:
            new_r = 0
            rec_r = 0
            for _, row in filtered.iterrows():
                m = row['monthly']
                if i >= len(m) or m[i] <= 0:
                    continue
                # Check if client bought in any of the previous 3 months
                had_recent = any(m[j] > 0 for j in range(max(0, i - 3), i))
                if had_recent:
                    rec_r += m[i]
                else:
                    new_r += m[i]
            new_rev_list.append(new_r)
            rec_rev_list.append(rec_r)

        fig_nr = go.Figure()
        fig_nr.add_trace(go.Bar(x=sel_months, y=rec_rev_list, name='Recorrente', marker_color='#3b82f6'))
        fig_nr.add_trace(go.Bar(x=sel_months, y=new_rev_list, name='Nova/Reativação', marker_color='#22c55e'))
        fig_nr.update_layout(
            title="Receita Recorrente vs Nova", barmode='stack', height=380,
            template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
            yaxis=dict(gridcolor='#e2e8f0', title=''),
            xaxis=dict(gridcolor='#e2e8f0'),
            legend=dict(orientation='h', y=1.12)
        )
        fig_nr.update_traces(hovertemplate='%{x}: R$ %{y:,.0f}<extra></extra>')
        st.plotly_chart(fig_nr, use_container_width=True)

        total_rec = sum(rec_rev_list)
        total_new = sum(new_rev_list)
        pct_rec = total_rec / (total_rec + total_new) * 100 if (total_rec + total_new) > 0 else 0
        st.caption(f"Recorrente: {pct_rec:.0f}% | Nova/Reativação: {100 - pct_rec:.0f}%")

    with col_ret:
        # Retention curve: for each month, % of prev month buyers who bought again
        ret_months = []
        ret_rates = []
        for i in sel_indices_sorted:
            if i < 1:
                continue
            prev_buyers = set()
            curr_buyers = set()
            for idx, row in filtered.iterrows():
                m = row['monthly']
                if i - 1 < len(m) and m[i - 1] > 0:
                    prev_buyers.add(idx)
                if i < len(m) and m[i] > 0:
                    curr_buyers.add(idx)
            if len(prev_buyers) > 0:
                ret_months.append(months[i])
                ret_rates.append(len(prev_buyers & curr_buyers) / len(prev_buyers) * 100)

        if ret_rates:
            fig_ret = go.Figure()
            fig_ret.add_trace(go.Scatter(
                x=ret_months, y=ret_rates, mode='lines+markers',
                line=dict(color='#22c55e', width=2), marker=dict(size=5),
                hovertemplate='%{x}: %{y:.1f}%<extra></extra>'
            ))
            avg_ret = sum(ret_rates) / len(ret_rates)
            fig_ret.add_hline(y=avg_ret, line_dash="dash", line_color="rgba(234,179,8,0.5)",
                             annotation_text=f"Média: {avg_ret:.0f}%", annotation_position="top right")
            fig_ret.update_layout(
                title="Retenção Mensal (%)", height=380,
                template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                yaxis=dict(gridcolor='#e2e8f0', range=[0, 100], title=''),
                xaxis=dict(gridcolor='#e2e8f0'),
                showlegend=False
            )
            st.plotly_chart(fig_ret, use_container_width=True)
            st.caption(f"Média de retenção no período: {avg_ret:.0f}%")

    st.divider()

    # --- COBERTURA POR VENDEDOR + STATUS ---
    col_left, col_right = st.columns(2)

    with col_left:
        # Vendor coverage: % of their clients that bought in the period
        vendor_cov = filtered.groupby('vendor').apply(
            lambda g: pd.Series({
                'total_clients': len(g),
                'active_buyers': len(g[g['monthly'].apply(
                    lambda m: any(m[i] > 0 for i in sel_indices_sorted if i < len(m))
                )]),
                'revenue': g['monthly'].apply(period_sum).sum()
            })
        ).reset_index()
        vendor_cov = vendor_cov[vendor_cov['total_clients'] > 0].copy()
        vendor_cov['coverage'] = (vendor_cov['active_buyers'] / vendor_cov['total_clients'] * 100).round(1)
        vendor_cov['vendor_short'] = vendor_cov['vendor'].str.replace(' Propetz Distribuição', '').str.replace(' La Maison Propetz', '')
        vendor_cov = vendor_cov.sort_values('revenue', ascending=False)

        fig_cov = go.Figure()
        fig_cov.add_trace(go.Bar(
            x=vendor_cov['vendor_short'], y=vendor_cov['coverage'],
            marker_color=['#22c55e' if c >= 50 else '#eab308' if c >= 30 else '#ef4444' for c in vendor_cov['coverage']],
            text=[f"{c:.0f}%" for c in vendor_cov['coverage']], textposition='outside',
            hovertemplate='%{x}<br>Cobertura: %{y:.1f}%<br>Comprando: %{customdata[0]} de %{customdata[1]}<extra></extra>',
            customdata=list(zip(vendor_cov['active_buyers'].astype(int), vendor_cov['total_clients'].astype(int)))
        ))
        fig_cov.update_layout(
            title="Cobertura por Vendedor (% clientes comprando)", height=380,
            template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
            yaxis=dict(gridcolor='#e2e8f0', range=[0, 100], title=''),
            xaxis=dict(title=''), showlegend=False
        )
        st.plotly_chart(fig_cov, use_container_width=True)

    with col_right:
        status_counts = filtered['status'].value_counts()
        colors = {'Ativo': '#22c55e', 'Inativo': '#ef4444', 'Inadimplente': '#eab308', 'Permuta': '#f97316'}
        fig_st = px.pie(values=status_counts.values, names=status_counts.index,
                        title="Clientes por Status", color=status_counts.index,
                        color_discrete_map=colors)
        fig_st.update_layout(template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff', height=380)
        st.plotly_chart(fig_st, use_container_width=True)

    st.divider()

    # --- TOP CLIENTES + RECEITA POR VENDEDOR ---
    col_l2, col_r2 = st.columns(2)

    with col_l2:
        top10 = filtered.copy()
        top10['period_total'] = top10['monthly'].apply(period_sum)
        top10 = top10.nlargest(10, 'period_total')
        fig_top = px.bar(top10, y='name', x='period_total', orientation='h',
                         title="Top 10 Clientes (Receita no Período)", color_discrete_sequence=['#3b82f6'])
        fig_top.update_layout(template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                              yaxis=dict(autorange='reversed', title=''),
                              xaxis=dict(title='Receita', gridcolor='#e2e8f0'),
                              showlegend=False, height=400)
        fig_top.update_traces(hovertemplate='%{y}: R$ %{x:,.0f}<extra></extra>')
        st.plotly_chart(fig_top, use_container_width=True)

    with col_r2:
        vendor_data = filtered.groupby('vendor').apply(
            lambda g: pd.Series({
                'total': g['monthly'].apply(period_sum).sum(),
                'clients': len(g),
                'active': len(g[g['status'] == 'Ativo'])
            })
        ).reset_index()
        vendor_data = vendor_data[vendor_data['total'] > 0].sort_values('total', ascending=False)
        vendor_data['vendor_short'] = vendor_data['vendor'].str.replace(' Propetz Distribuição', '').str.replace(' La Maison Propetz', '')

        fig_vend = px.bar(vendor_data, x='vendor_short', y='total',
                          title="Receita por Vendedor (Período)", color_discrete_sequence=['#2563eb'])
        fig_vend.update_layout(template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                               xaxis=dict(title=''), yaxis=dict(title='Receita', gridcolor='#e2e8f0'),
                               showlegend=False, height=400)
        fig_vend.update_traces(hovertemplate='%{x}: R$ %{y:,.0f}<extra></extra>')
        st.plotly_chart(fig_vend, use_container_width=True)

    st.divider()

    # --- CLIENTES INATIVOS COM POTENCIAL ---
    st.subheader("💰 Clientes Inativos com Maior Potencial")
    inactive_pot = filtered.copy()
    inactive_pot['total_hist'] = inactive_pot['monthly'].apply(sum)
    inactive_pot['period_rev'] = inactive_pot['monthly'].apply(period_sum)
    # Clients with zero revenue in last 6 months but significant historical revenue
    last_6_start = max(0, len(months) - 6)
    inactive_pot['last6'] = inactive_pot['monthly'].apply(lambda m: sum(m[last_6_start:]) if len(m) > last_6_start else 0)
    dormant = inactive_pot[(inactive_pot['last6'] == 0) & (inactive_pot['total_hist'] > 0)].copy()
    dormant = dormant.nlargest(10, 'total_hist')

    if len(dormant) > 0:
        dormant_display = dormant[['name', 'vendor', 'state', 'total_hist']].copy()
        dormant_display.columns = ['Cliente', 'Vendedor', 'UF', 'Receita Histórica']
        dormant_display['Vendedor'] = dormant_display['Vendedor'].str.replace(' Propetz Distribuição', '').str.replace(' La Maison Propetz', '')
        dormant_display = dormant_display.reset_index(drop=True)
        dormant_display.index = dormant_display.index + 1
        show_money_table(dormant_display, ['Receita Histórica'], use_container_width=True, height=min(400, 35 * len(dormant_display) + 38))
        total_dormant_rev = dormant['total_hist'].sum()
        st.caption(f"Estes {len(dormant)} clientes já geraram {fmt_brl_full(total_dormant_rev)} em receita total. Uma ligação pode reativá-los.")
    else:
        st.success("Todos os clientes com histórico relevante estão comprando!")

    st.divider()

    # --- INSIGHTS AUTOMÁTICOS (renderizados no insights_box, no topo) ---
    insights = []

    # Insight: Churn risk
    at_risk = filtered[filtered['risk'] == 'Recuperação']
    if len(at_risk) > 0:
        lost = at_risk['monthly'].apply(annual_value_estimate).sum()
        insights.append(insight_html('danger', 'RISCO DE CHURN',
            f"{len(at_risk)} clientes há 6+ meses sem comprar. Perda anual estimada: {fmt_brl_full(lost)}.",
            "Priorize contato imediato com os maiores tickets."))

    # Insight: Period trend
    monthly_sel = [monthly_totals[i] for i in sel_indices_sorted if i < len(monthly_totals)]
    if len(monthly_sel) >= 6:
        half = len(monthly_sel) // 2
        first_half = sum(monthly_sel[:half])
        second_half = sum(monthly_sel[half:])
        if first_half > 0:
            trend = (second_half - first_half) / first_half
            if trend > 0.05:
                insights.append(insight_html('success', 'TENDÊNCIA POSITIVA',
                    f"A 2ª metade do período selecionado cresceu {trend*100:.1f}% vs a 1ª metade.",
                    "Manter estratégia atual e identificar o que está funcionando."))
            elif trend < -0.05:
                insights.append(insight_html('warning', 'ALERTA DE QUEDA',
                    f"A 2ª metade do período selecionado caiu {abs(trend)*100:.1f}% vs a 1ª metade.",
                    "Investigar: clientes perdidos? Sazonalidade? Problema de produto?"))

    # Insight: Concentration
    sorted_by_rev = filtered.copy()
    sorted_by_rev['p_total'] = sorted_by_rev['monthly'].apply(period_sum)
    sorted_by_rev = sorted_by_rev.sort_values('p_total', ascending=False)
    if len(sorted_by_rev) >= 5:
        top5_rev = sorted_by_rev.head(5)['p_total'].sum()
        all_rev = sorted_by_rev['p_total'].sum()
        conc = top5_rev / all_rev if all_rev > 0 else 0
        if conc > 0.3:
            insights.append(insight_html('warning', 'CONCENTRAÇÃO',
                f"Os 5 maiores clientes representam {conc*100:.1f}% da receita no período. Risco de dependência.",
                "Diversificar: focar em clientes médios com potencial de crescimento."))

    # Insight: Retention
    if retention_rate > 0 and retention_rate < 50:
        insights.append(insight_html('danger', 'RETENÇÃO BAIXA',
            f"Apenas {retention_rate:.0f}% dos clientes do período anterior voltaram a comprar.",
            "Ação urgente: entender por que clientes não estão recomprando."))
    elif retention_rate >= 70:
        insights.append(insight_html('success', 'RETENÇÃO SAUDÁVEL',
            f"{retention_rate:.0f}% dos clientes do período anterior voltaram a comprar.",
            "Base fiel. Foco em aumentar ticket médio dos recorrentes."))

    # Insight: Revenue mix
    if (total_rec + total_new) > 0 and total_new / (total_rec + total_new) > 0.4:
        insights.append(insight_html('success', 'CONQUISTA FORTE',
            f"{total_new/(total_rec+total_new)*100:.0f}% da receita vem de clientes novos ou reativados.",
            "Crescimento saudável. Garantir que esses novos clientes se tornem recorrentes."))

    # Insight: Vendor coverage
    if len(vendor_cov) > 0:
        low_cov = vendor_cov[vendor_cov['coverage'] < 30]
        if len(low_cov) > 0:
            names = ', '.join(low_cov['vendor_short'].tolist())
            insights.append(insight_html('warning', 'COBERTURA BAIXA',
                f"Vendedor(es) com menos de 30% dos clientes comprando: {names}.",
                "Revisar: carteira grande demais? Clientes desatualizados? Falta de follow-up?"))

    with insights_box:
        for ins in insights:
            st.markdown(ins, unsafe_allow_html=True)
        if not insights:
            st.info("Nenhum insight crítico para o período selecionado. Operação estável.")

# ============================================================
# PAGE: CLIENTES
# ============================================================
def page_clients(df, df_sku, months, year_ranges, sel_indices_sorted, sel_months):
    st.header("👤 Visão por Cliente")

    search = st.text_input("🔍 Buscar cliente (nome, estado, código ou vendedor)", key="client_search")

    filtered = df.copy()
    if search:
        s = search.lower()
        filtered = filtered[
            filtered['name'].str.lower().str.contains(s, na=False) |
            filtered['state'].str.lower().str.contains(s, na=False) |
            filtered['id'].str.contains(s, na=False) |
            filtered['vendor'].str.lower().str.contains(s, na=False)
        ]

    # Period-based revenue
    def _period_sum(m):
        return sum(m[i] for i in sel_indices_sorted if i < len(m))

    filtered = filtered.copy()
    filtered['total_rev'] = filtered['monthly'].apply(_period_sum)
    filtered = filtered.sort_values('total_rev', ascending=False)

    # Client selector
    client_names = filtered['name'].tolist()

    if not client_names:
        st.info("Nenhum cliente encontrado.")
        return

    # Show table
    period_label = f"{sel_months[0]} - {sel_months[-1]}" if len(sel_months) > 1 else (sel_months[0] if sel_months else "")
    display_df = filtered[['name','state','vendor','status','risk','total_rev','last_purchase']].head(50).copy()
    rev_col = f'Receita ({period_label})'
    display_df.columns = ['Cliente','UF','Vendedor','Status','Risco',rev_col,'Última Compra']
    display_df['Vendedor'] = display_df['Vendedor'].str.replace(' Propetz Distribuição','').str.replace(' La Maison Propetz','')

    show_money_table(display_df, [rev_col], use_container_width=True, height=300, hide_index=True)

    st.divider()

    # Client detail
    selected = st.selectbox("Selecione um cliente para ver detalhes:", client_names[:100], key="client_select")

    if selected:
        c = df[df['name'] == selected].iloc[0]
        client_id = str(c['id']).strip()
        monthly = c['monthly']
        period_total = _period_sum(monthly)
        total = sum(monthly)
        months_active, n_sel = _commercial_period_recurrence(monthly, sel_indices_sorted)
        avg_ticket = period_total / months_active if months_active > 0 else 0

        st.subheader(f"📋 {c['name']}")

        meta_cols = st.columns(6)
        meta_cols[0].markdown(f"**UF:** {c['state']}")
        meta_cols[1].markdown(f"**ID:** {c['id']}")
        meta_cols[2].markdown(f"**Vendedor:** {c['vendor'].replace(' Propetz Distribuição','').replace(' La Maison Propetz','')}")
        meta_cols[3].markdown(f"**Status:** {status_badge(c['status'])}", unsafe_allow_html=True)
        meta_cols[4].markdown(f"**Risco:** {risk_badge(c['risk'])}", unsafe_allow_html=True)
        if c['credit_limit'] > 0:
            meta_cols[5].markdown(f"**Limite:** {fmt_brl_full(c['credit_limit'])}")

        # KPIs
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Receita no Período", fmt_brl(period_total), f"Total histórico: {fmt_brl(total)}")
        k2.metric("Ticket Médio/Mês", fmt_brl(avg_ticket), f"{months_active} meses com compra")
        k3.metric("Última Compra", c['last_purchase'], f"{c['months_since']} meses atrás" if c['months_since'] < 999 else "Nunca")
        k4.metric("Frequência no Período", f"{months_active}/{n_sel}", f"{months_active/n_sel*100:.0f}% dos meses" if n_sel > 0 else "")

        # Trend chart
        ma3 = []
        for i in range(len(monthly)):
            if i < 2:
                ma3.append(monthly[i])
            else:
                ma3.append((monthly[i] + monthly[i-1] + monthly[i-2]) / 3)

        # Color bars: selected period = bold, others = subtle
        bar_colors = ['#3b82f6' if i in set(sel_indices_sorted) else 'rgba(59,130,246,0.2)' for i in range(len(monthly))]
        fig = go.Figure()
        fig.add_trace(go.Bar(x=months, y=monthly, name='Receita Mensal', marker_color=bar_colors))
        fig.add_trace(go.Scatter(x=months, y=ma3, name='Média Móvel (3m)', line=dict(color='#f97316', width=2), mode='lines'))
        fig.update_layout(title="Evolução de Vendas", height=400,
                         template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                         yaxis=dict(gridcolor='#e2e8f0'),
                         hovermode='x unified')
        st.plotly_chart(fig, use_container_width=True)

        # Yearly comparison
        col1, col2 = st.columns(2)
        yt = c['yearly_totals'] if isinstance(c['yearly_totals'], dict) else {}
        am = c['avg_month'] if isinstance(c['avg_month'], dict) else {}

        with col1:
            years = ['2021','2022','2023','2024','2025','2026']
            fig_yr = px.bar(x=years, y=[yt.get(y,0) for y in years],
                           title="Receita por Ano", color_discrete_sequence=['#3b82f6'])
            fig_yr.update_layout(template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                                showlegend=False, height=300, yaxis=dict(gridcolor='#e2e8f0'))
            st.plotly_chart(fig_yr, use_container_width=True)

        with col2:
            fig_am = px.bar(x=['2021','2022','2023','2024','2025'], y=[am.get(y,0) for y in ['2021','2022','2023','2024','2025']],
                           title="Ticket Médio por Ano", color_discrete_sequence=['#22c55e'])
            fig_am.update_layout(template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                                showlegend=False, height=300, yaxis=dict(gridcolor='#e2e8f0'))
            st.plotly_chart(fig_am, use_container_width=True)

        # Client insights
        st.subheader("🧠 Insights do Cliente")

        last6 = sum(monthly[-6:])
        prev6 = sum(monthly[-12:-6])
        if prev6 > 0:
            change = (last6 - prev6) / prev6
            if change < -0.3:
                st.markdown(insight_html('danger', 'QUEDA ACENTUADA',
                    f"Volume caiu {abs(change)*100:.0f}% nos últimos 6 meses ({fmt_brl(last6)}) vs anteriores ({fmt_brl(prev6)}).",
                    "Ação urgente: agendar visita ou ligação para entender a causa."), unsafe_allow_html=True)
            elif change < -0.1:
                st.markdown(insight_html('warning', 'TENDÊNCIA DE QUEDA',
                    f"Volume reduziu {abs(change)*100:.0f}% nos últimos 6 meses.",
                    "Monitorar e investigar causas possíveis."), unsafe_allow_html=True)
            elif change > 0.2:
                st.markdown(insight_html('success', 'CRESCIMENTO',
                    f"Volume cresceu {change*100:.0f}%! De {fmt_brl(prev6)} para {fmt_brl(last6)}.",
                    "Aproveitar momento para ampliar mix de produtos."), unsafe_allow_html=True)

        if n_sel > 0 and months_active / n_sel < 0.3 and total > 10000:
            st.markdown(insight_html('warning', 'BAIXA RECORRÊNCIA',
                f"Comprou em apenas {months_active} de {n_sel} meses do período selecionado "
                f"({months_active/n_sel*100:.0f}%), mas tem faturamento histórico relevante.",
                "Oportunidade: criar rotina de compras recorrentes."), unsafe_allow_html=True)

        if c['months_since'] >= 3 and c['months_since'] < 6 and c['status'] == 'Ativo':
            st.markdown(insight_html('warning', 'ATENÇÃO - INATIVIDADE',
                f"Cliente ativo sem compras há {c['months_since']} meses. Última: {c['last_purchase']}.",
                "Entrar em contato antes que vire churn."), unsafe_allow_html=True)
        elif c['months_since'] >= 6 and c['status'] == 'Ativo':
            st.markdown(insight_html('danger', 'URGENTE - RECUPERAÇÃO',
                f"Cliente ativo sem compras há {c['months_since']} meses! Risco iminente de perda.",
                "Ação imediata: contato direto + oferta especial."), unsafe_allow_html=True)

        avg_all = df[df['status']=='Ativo']['monthly'].apply(sum).mean() / len(months)
        if avg_ticket > avg_all * 2:
            st.markdown(insight_html('success', 'CLIENTE PREMIUM',
                f"Ticket médio de {fmt_brl(avg_ticket)} é {avg_ticket/avg_all:.1f}x acima da média ({fmt_brl(avg_all)}).",
                "Cliente estratégico: garantir atendimento diferenciado."), unsafe_allow_html=True)

        if c['status'] == 'Ativo' and c['months_since'] < 3 and (prev6 == 0 or (last6 - prev6) / prev6 >= -0.1):
            st.markdown(insight_html('success', 'CLIENTE SAUDÁVEL',
                f"Cliente ativo com compras recentes ({c['last_purchase']}). Manter relacionamento.",
                "Explorar oportunidades de mix de produtos."), unsafe_allow_html=True)

        # ============================================================
        # SEÇÃO: DETALHES DE PRODUTOS COM SKU E QUANTIDADE (from df_sku)
        # ============================================================
        st.subheader("📦 Produtos Comprados (Detalhes por SKU)")

        if len(df_sku) > 0:
            # Get products this client bought
            client_skus = df_sku[df_sku['cod_cliente'].astype(str).str.strip() == str(client_id).strip()].copy()

            if len(client_skus) > 0:
                # Aggregate by SKU and product
                sku_detail = client_skus.groupby(['sku', 'produto']).agg({
                    'quantidade': 'sum',
                    'mes': 'nunique'
                }).reset_index()
                sku_detail.columns = ['SKU', 'Produto', 'Qtd Total', 'Meses']

                # Calculate global mix % (from total all quantities across all clients/products)
                total_all_qty = df_sku['quantidade'].sum()
                sku_detail['% Mix Global'] = (sku_detail['Qtd Total'] / total_all_qty * 100).round(2) if total_all_qty > 0 else 0

                sku_detail = sku_detail.sort_values('Qtd Total', ascending=False)

                st.dataframe(sku_detail, use_container_width=True, hide_index=True, 
                           height=min(400, 35 * len(sku_detail) + 38))
            else:
                st.info("Nenhum dado de quantidade por SKU disponível para este cliente.")
        else:
            st.info("Dados de SKU não carregados.")

# ============================================================
# PAGE: MIX
# ============================================================
def page_mix(df, products_df, df_client_products, df_sku, months, sel_indices_sorted, sel_months):
    st.header("🧩 Oportunidades de Mix de Produtos")

    _abc_meta = load_abc_valor()
    if _abc_meta:
        st.caption(f"Curvas A/B/C e estimativas em R$ calculadas pelo **faturamento real** do canal "
                   f"Distribuição ({_abc_meta.get('periodo', 'últimos 12 meses')}, Base Mãe).")

    def _period_sum(m):
        return sum(m[i] for i in sel_indices_sorted if i < len(m))

    active_clients = df[_commercial_active_mask(df, load_inactive_clients())].copy()
    active_clients['total'] = active_clients['monthly'].apply(_period_sum)
    active_clients = active_clients.sort_values('total', ascending=False)

    period_label = f"{sel_months[0]} - {sel_months[-1]}" if len(sel_months) > 1 else (sel_months[0] if sel_months else "")

    selected = st.selectbox("Selecione um cliente:", active_clients['name'].tolist(), key="mix_client")
    if not selected:
        return

    c = df[df['name'] == selected].iloc[0]
    client_id = str(c['id']).strip()
    total = _period_sum(c['monthly'])
    months_active = sum(1 for i in sel_indices_sorted if i < len(c['monthly']) and c['monthly'][i] > 0)
    avg = total / months_active if months_active > 0 else 0
    is_admin = has_full_data_access()

    # Motor de valor: preço médio real por SKU + estatísticas dos compradores
    preco_map = _preco_medio_map(products_df)
    typical, buyers_n, per_buyer = _sku_stats(df_sku)
    client_pm = {}
    if len(per_buyer) > 0:
        _mine = per_buyer[per_buyer['cod_cliente'] == client_id]
        client_pm = dict(zip(_mine['sku'], _mine['pm']))

    # Produtos que o cliente compra (histórico completo)
    cp_client = pd.DataFrame()
    if len(df_client_products) > 0:
        cp_client = df_client_products[df_client_products['client_id'].astype(str).str.strip() == client_id].copy()
    bought_codes = set(cp_client['product_code']) if len(cp_client) > 0 else set(client_pm.keys())
    has_product_data = len(bought_codes) > 0

    # ---- OPORTUNIDADES: produtos A/B onde há dinheiro na mesa ----
    ops = []
    if has_product_data:
        catalog = products_df[products_df['abc'].isin(['A', 'B'])]
        for code, name, abc in zip(catalog['code'], catalog['name'], catalog['abc']):
            prc = preco_map.get(code)
            typ = typical.get(code)
            nb = buyers_n.get(code, 0)
            if not prc or not typ or typ <= 0 or nb < 5:
                continue  # sem base estatística suficiente
            if code not in bought_codes:
                pot = typ * prc
                tipo = '🆕 Não compra'
                why = f"{nb} clientes compram (típico {typ:.0f}/mês)"
            else:
                atual = client_pm.get(code)
                if atual is None or atual >= typ * 0.5:
                    continue  # já compra em nível razoável
                pot = (typ - atual) * prc
                tipo = '📉 Compra pouco'
                why = f"compra {atual:.1f}/mês vs típico {typ:.0f}/mês"
            if pot < 100:
                continue  # materialidade mínima: R$ 100/mês
            ops.append({
                'Tipo': tipo,
                'Código': code,
                'Produto': name,
                'Curva': abc,
                'Por quê': why,
                'R$ Potencial/Mês': round(pot, 2),
            })
    ops_df = pd.DataFrame(ops)
    if len(ops_df) > 0:
        ops_df = ops_df.sort_values('R$ Potencial/Mês', ascending=False)
    pot_total = ops_df['R$ Potencial/Mês'].sum() if len(ops_df) > 0 else 0

    # ---- KPIs ----
    k1, k2, k3, k4 = st.columns(4)
    k1.metric(f"Receita ({period_label})", fmt_brl(total))
    k2.metric("Ticket Médio/Mês", fmt_brl(avg), f"{months_active} meses com compra")
    k3.metric("Produtos Comprados", f"{len(bought_codes)}", f"de {len(products_df)} no catálogo")
    k4.metric("💰 Potencial de Mix", f"{fmt_brl(pot_total)}/mês", f"{len(ops_df)} oportunidades")

    st.divider()

    # ---- SEÇÃO 1: O QUE OFERECER (a receita vem primeiro) ----
    st.subheader("🎯 O que oferecer para este cliente")
    if not has_product_data:
        st.info("Sem histórico de produtos para este cliente — não é possível calcular oportunidades.")
    elif len(ops_df) == 0:
        st.success("Este cliente já compra os produtos Curva A/B em nível típico — sem gaps relevantes.")
    else:
        st.caption("Produtos **Curva A/B** com 5+ compradores, priorizados pelo R$ estimado "
                   "(quantidade típica dos compradores × preço médio real). Entram na lista apenas "
                   "oportunidades acima de R$ 100/mês.")
        show_money_table(ops_df.head(15), ['R$ Potencial/Mês'], use_container_width=True, hide_index=True,
                         height=min(560, 35 * min(len(ops_df), 15) + 38))
        if len(ops_df) > 15:
            st.caption(f"Mostrando as 15 maiores de {len(ops_df)} oportunidades — baixe a lista completa.")
        _csv_download(ops_df, "⬇️ Baixar oportunidades (Excel/CSV)", "oportunidades_mix.csv", "dl_mix_ops")

    st.divider()

    # ---- SEÇÃO 2: RAIO-X — o que o cliente compra hoje (em valor) ----
    st.subheader("📊 O que o cliente compra hoje")
    if len(cp_client) == 0:
        st.info("Sem dados de produtos comprados para este cliente.")
        return

    cp = cp_client.merge(
        products_df[['code', 'abc']].rename(columns={'code': 'product_code'}),
        on='product_code', how='left'
    )
    cp['abc'] = cp['abc'].fillna('C')
    cp['preco'] = cp['product_code'].map(preco_map)
    cp['valor_est'] = (cp['total_qty'] * cp['preco']).fillna(0)
    _tot_val = cp['valor_est'].sum()
    cp['mix_pct'] = (cp['valor_est'] / _tot_val * 100).round(1) if _tot_val > 0 else 0.0
    cp = cp.sort_values('valor_est', ascending=False)

    top12 = cp.head(12)
    if is_admin:
        fig_top = px.bar(top12, y='product_name', x='valor_est', orientation='h', color='abc',
                         title="Top 12 do Cliente (R$ estimado, histórico)",
                         color_discrete_map={'A': '#22c55e', 'B': '#eab308', 'C': '#ef4444'})
        fig_top.update_traces(hovertemplate='%{y}: R$ %{x:,.0f}<extra></extra>')
        fig_top.update_layout(xaxis=dict(title='R$ estimado'))
    else:
        fig_top = px.bar(top12, y='product_name', x='mix_pct', orientation='h', color='abc',
                         title="Top 12 do Cliente (% do mix em valor)",
                         color_discrete_map={'A': '#22c55e', 'B': '#eab308', 'C': '#ef4444'})
        fig_top.update_traces(hovertemplate='%{y}: %{x:.1f}%<extra></extra>')
        fig_top.update_layout(xaxis=dict(title='% do mix'))
    fig_top.update_layout(template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                          yaxis=dict(autorange='reversed', title=''), height=420, showlegend=True)
    st.plotly_chart(fig_top, use_container_width=True)

    if is_admin:
        disp_cp = cp[['product_code', 'product_name', 'abc', 'total_qty', 'valor_est', 'mix_pct']].copy()
        disp_cp.columns = ['Código', 'Produto', 'Curva', 'Qtd Total', 'R$ Estimado', '% Mix (valor)']
        show_money_table(disp_cp, ['R$ Estimado'], use_container_width=True, hide_index=True,
                         height=min(420, 35 * len(disp_cp) + 38))
    else:
        disp_cp = cp[['product_code', 'product_name', 'abc', 'mix_pct']].copy()
        disp_cp.columns = ['Código', 'Produto', 'Curva', '% Mix (valor)']
        st.dataframe(disp_cp, use_container_width=True, hide_index=True,
                     height=min(420, 35 * len(disp_cp) + 38))

# ============================================================
# PAGE: CHURN
# ============================================================
def page_churn(df, months, sel_indices_sorted, sel_months):
    st.header("⚠️ Gestão de Churn")

    _sv = load_silver_distribuicao()
    if isinstance(_sv, dict) and _sv.get("clientes"):
        st.caption("🔄 Fonte DarosCorp conectada — última publicação: "
                   f"{_sv.get('gerado_em', '?')} (a classificação reflete em até 1h). "
                   "Clientes sem nota em 2026 seguem pela planilha (histórico).")
        _novos = _sv.get("novos_sem_cadastro")
        _novos = [n for n in _novos if isinstance(n, dict)] if isinstance(_novos, list) else []
        if _novos and has_full_data_access():
            with st.expander(f"🆕 Comprando na fonte SEM cadastro na planilha ({len(_novos)}) — "
                             "cadastrar no fechamento do mês"):
                for n in _novos:
                    st.markdown(f"- **{n.get('nome', '?')}** — {fmt_brl(n.get('receita_2026', 0))} "
                                f"em 2026, última NF {n.get('ultima', '?')} "
                                f"({n.get('cidade', '?')}/{n.get('uf', '?')})")

    period_label = f"{sel_months[0]} - {sel_months[-1]}" if len(sel_months) > 1 else (sel_months[0] if sel_months else "")

    def _period_sum(m):
        return sum(m[i] for i in sel_indices_sorted if i < len(m))

    # Load inactive clients
    inactive_ids = load_inactive_clients()

    # Mesmo recorte de carteira ativa das páginas Ações e Gestor.
    active_mask = _commercial_active_mask(df, inactive_ids)
    df_active = df[active_mask].copy()
    df_inactive = df[~active_mask].copy()
    st.caption("Indicadores e ranking consideram apenas clientes Ativos na planilha "
               "e sem inativação no app. Os demais aparecem em Inativos / Outros.")

    recup = df_active[df_active['risk'] == 'Recuperação'].copy()
    atencao = df_active[df_active['risk'] == 'Atenção'].copy()
    saudavel = df_active[df_active['risk'] == 'Saudável']

    recup_impact = recup['monthly'].apply(annual_value_estimate).sum()
    atencao_impact = atencao['monthly'].apply(annual_value_estimate).sum()

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("🔴 Recuperação (6+ meses)", f"{len(recup)}", f"Impacto: {fmt_brl(recup_impact)}/ano")
    k2.metric("🟡 Atenção (3-5 meses)", f"{len(atencao)}", f"Impacto: {fmt_brl(atencao_impact)}/ano")
    k3.metric("🟢 Saudáveis", f"{len(saudavel)}")
    k4.metric("💰 Receita Total em Risco", fmt_brl(recup_impact + atencao_impact), "Estimativa anual")
    _commercial_estimate_caption()

    st.divider()

    tab1, tab2, tab3, tab4 = st.tabs(["🔴 Recuperação", "🟡 Atenção", "📊 Ranking Vendedores", f"🚫 Inativos / Outros ({len(df_inactive)})"])

    # --- Helper to render a churn table with inactivate buttons ---
    def _render_churn_table(data, tab_key):
        if len(data) == 0:
            st.success(f"Nenhum cliente nesta categoria!")
            return

        data['total_rev'] = data['monthly'].apply(_period_sum)
        data['impact'] = data['monthly'].apply(annual_value_estimate)
        data['vendor_short'] = data['vendor'].str.replace(' Propetz Distribuição','').str.replace(' La Maison Propetz','')
        data = data.sort_values('total_rev', ascending=False)

        # Busca: filtra a tabela E o seletor de inativação ao mesmo tempo
        _busca_c = st.text_input("🔍 Buscar cliente (nome, UF, vendedor ou código)", key=f"churn_search_{tab_key}",
                                 placeholder="Digite parte do nome para filtrar a lista...")
        data = _filter_clients_by_term(data, _busca_c)
        if len(data) == 0:
            st.info("Nenhum cliente encontrado com esse termo.")
            return

        # Multiselect to choose clients to inactivate
        _is_full = can_approve_inactivations()
        client_names = data['name'].tolist()
        selected_to_inactivate = st.multiselect(
            "Selecione clientes para INATIVAR:" if _is_full else "Selecione clientes para SOLICITAR inativação:",
            options=client_names,
            key=f"inactivate_{tab_key}",
            help="Clientes inativados saem das listas de churn e ações" if _is_full
                 else "A solicitação vai para aprovação do administrador"
        )

        if selected_to_inactivate:
            _clientes = []
            for name in selected_to_inactivate:
                match = data[data['name'] == name]
                if len(match) > 0:
                    _clientes.append({'cid': str(match.iloc[0]['id']).strip(), 'name': name,
                                      'vendor': match.iloc[0].get('vendor', '')})
            _inativacao_form(_clientes, f"churn_{tab_key}")

        display = data[['name','state','vendor_short','last_purchase','months_since','impact','total_rev']].copy()
        display.columns = ['Cliente','UF','Vendedor','Última Compra','Meses Inativo','Impacto Anual Est.',f'Receita ({period_label})']
        show_money_table(display, ['Impacto Anual Est.', f'Receita ({period_label})'],
                         use_container_width=True, hide_index=True, height=500)

    with tab1:
        _render_churn_table(recup, "recup")

    with tab2:
        _render_churn_table(atencao, "atencao")

    with tab3:
        if df_active.empty:
            st.info("Nenhum cliente ativo para compor o ranking de vendedores.")
        else:
            vendor_risk = df_active.groupby('vendor')[['risk', 'monthly']].apply(lambda g: pd.Series({
                'total': len(g),
                'recuperacao': len(g[g['risk']=='Recuperação']),
                'atencao': len(g[g['risk']=='Atenção']),
                'impact': g[g['risk'].isin(['Recuperação','Atenção'])]['monthly'].apply(annual_value_estimate).sum()
            })).reset_index()
            vendor_risk['vendor_short'] = vendor_risk['vendor'].str.replace(' Propetz Distribuição','').str.replace(' La Maison Propetz','')
            vendor_risk['pct_risco'] = ((vendor_risk['recuperacao'] + vendor_risk['atencao']) / vendor_risk['total'] * 100).round(1)
            vendor_risk = vendor_risk[vendor_risk['total'] > 0].sort_values('pct_risco', ascending=False)

            display = vendor_risk[['vendor_short','total','recuperacao','atencao','pct_risco']].copy()
            display.columns = ['Vendedor','Total Clientes','Recuperação','Atenção','% em Risco']
            st.dataframe(display, use_container_width=True, hide_index=True)

            fig = px.bar(vendor_risk, x='vendor_short', y=['recuperacao','atencao'],
                        title="Clientes em Risco por Vendedor",
                        color_discrete_map={'recuperacao':'#ef4444','atencao':'#eab308'},
                        labels={'value':'Clientes','vendor_short':'Vendedor'},
                        barmode='stack')
            fig.update_layout(template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff', height=350)
            st.plotly_chart(fig, use_container_width=True)

    with tab4:
        st.subheader("🚫 Clientes fora da carteira ativa")
        st.caption("A tabela distingue o status da planilha da inativação feita no app. "
                   "O administrador pode desfazer a inativação no app de quem já está Ativo na planilha. "
                   "Outros status precisam ser corrigidos na fonte antes de retornar à carteira.")

        if len(df_inactive) == 0:
            st.info("Nenhum cliente fora da carteira ativa.")
        else:
            df_inactive['total_rev'] = df_inactive['monthly'].apply(_period_sum)
            df_inactive['vendor_short'] = df_inactive['vendor'].str.replace(' Propetz Distribuição','').str.replace(' La Maison Propetz','')
            df_inactive['inactive_app'] = df_inactive['id'].astype(str).str.strip().isin(
                {str(cid).strip() for cid in inactive_ids}).map({True: 'Sim', False: 'Não'})
            df_inactive_sorted = df_inactive.sort_values('name')

            _busca_i = st.text_input("🔍 Buscar cliente fora da carteira ativa (nome, UF, vendedor ou código)", key="inativ_search",
                                     placeholder="Digite parte do nome para filtrar...")
            df_inactive_sorted = _filter_clients_by_term(df_inactive_sorted, _busca_i)
            if len(df_inactive_sorted) == 0:
                st.info("Nenhum cliente encontrado com esse termo.")
                return

            # Só admin; nunca tratar status da planilha como uma flag removível no app.
            if can_approve_inactivations():
                _react = _commercial_reactivation_candidates(df_inactive_sorted, inactive_ids)
                _react_labels = {str(r['id']).strip(): f"{r['name']} · {str(r['id']).strip()}"
                                 for _, r in _react.iterrows()}
                if _react_labels:
                    selected_to_reactivate = st.multiselect(
                        "Selecione clientes para REATIVAR:",
                        options=list(_react_labels),
                        format_func=lambda cid: _react_labels.get(cid, cid),
                        key="reactivate_clients",
                        help="Somente clientes Ativos na planilha e inativados no app podem retornar aqui"
                    )
                    _ids = [cid for cid in selected_to_reactivate if cid in _react_labels]
                    if _ids and st.button(f"✅ Reativar {len(_ids)} cliente(s)", key="btn_reactivate", type="primary"):
                        if reactivate_clients(_ids):
                            st.rerun()
                        else:
                            st.error("Não consegui salvar no GitHub agora. Tente de novo em instantes.")
            else:
                st.caption("Reativação de clientes é feita pelo administrador.")

            display_inact = df_inactive_sorted[['id','name','state','vendor_short','status','inactive_app',
                                                'risk','last_purchase','months_since','total_rev']].copy()
            display_inact.columns = ['Código','Cliente','UF','Vendedor','Status na Planilha','Inativação no App',
                                     'Risco Original','Última Compra','Meses sem Comprar',f'Receita ({period_label})']
            show_money_table(display_inact, [f'Receita ({period_label})'],
                             use_container_width=True, hide_index=True, height=500)

# ============================================================
# PAGE: PRODUTOS
# ============================================================
def page_products(products_df, df_sku):
    st.header("📦 Análise de Produtos")

    is_admin = has_full_data_access()

    # Critério da curva: faturamento (abc_valor.json) ou quantidade (fallback)
    abc_meta = load_abc_valor()
    has_val = abc_meta is not None and 'valor_12m' in products_df.columns and products_df['valor_12m'].sum() > 0
    if has_val:
        st.caption(f"Curva ABC calculada por **faturamento** do canal Distribuição — "
                   f"{abc_meta.get('periodo', 'últimos 12 meses')} (Base Mãe, atualizado em "
                   f"{abc_meta.get('gerado_em', '—')}). A = 80% do faturamento acumulado, B = 15%, C = 5%. "
                   f"Produtos sem venda no período ficam na curva C.")
        abc_base = products_df['valor_12m']
        base_label = "do faturamento"
    else:
        st.caption("Curva ABC por quantidade (planilha). Gere o abc_valor.json para usar faturamento.")
        abc_base = products_df['total_qty']
        base_label = "do volume"

    count_a = len(products_df[products_df['abc']=='A'])
    count_b = len(products_df[products_df['abc']=='B'])
    count_c = len(products_df[products_df['abc']=='C'])
    total_qty = products_df['total_qty'].sum()
    total_base = abc_base.sum()
    pct_a = abc_base[products_df['abc']=='A'].sum() / total_base * 100 if total_base > 0 else 0
    pct_b = abc_base[products_df['abc']=='B'].sum() / total_base * 100 if total_base > 0 else 0
    pct_c = abc_base[products_df['abc']=='C'].sum() / total_base * 100 if total_base > 0 else 0

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Total Produtos", f"{len(products_df)}")
    k2.metric("Curva A", f"{count_a}", f"{pct_a:.1f}% {base_label}")
    k3.metric("Curva B", f"{count_b}", f"{pct_b:.1f}% {base_label}")
    k4.metric("Curva C", f"{count_c}", f"{pct_c:.1f}% {base_label}")

    col1, col2 = st.columns(2)

    with col1:
        _df_abc = products_df.copy()
        _df_abc['_base'] = abc_base
        abc_data = _df_abc.groupby('abc')['_base'].sum().reset_index().rename(columns={'_base': 'total_qty'})
        abc_data['pct'] = (abc_data['total_qty'] / abc_data['total_qty'].sum() * 100).round(1)
        fig_abc = px.pie(abc_data, values='total_qty', names='abc',
                        title=f"Curva ABC - % {base_label.replace('do ', '')}",
                        color='abc', color_discrete_map={'A':'#22c55e','B':'#eab308','C':'#ef4444'})
        fig_abc.update_traces(textinfo='label+percent', hovertemplate='Curva %{label}: %{percent}<extra></extra>')
        fig_abc.update_layout(template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff', height=400)
        st.plotly_chart(fig_abc, use_container_width=True)

    with col2:
        top20 = products_df.head(20).copy()
        if is_admin:
            fig_top = px.bar(top20, y='name', x='total_qty', orientation='h', color='abc',
                            title="Top 20 Produtos (Volume)",
                            color_discrete_map={'A':'#22c55e','B':'#eab308','C':'#ef4444'})
            fig_top.update_layout(template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                                 yaxis=dict(autorange='reversed', title=''), height=500, showlegend=True)
        else:
            # Vendor view: show % of total instead of quantities
            top20['pct'] = (top20['total_qty'] / total_qty * 100).round(2)
            fig_top = px.bar(top20, y='name', x='pct', orientation='h', color='abc',
                            title="Top 20 Produtos (% do Volume Total)",
                            color_discrete_map={'A':'#22c55e','B':'#eab308','C':'#ef4444'})
            fig_top.update_traces(hovertemplate='%{y}: %{x:.2f}%<extra></extra>')
            fig_top.update_layout(template='plotly_white', paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                                 yaxis=dict(autorange='reversed', title=''),
                                 xaxis=dict(title='% do Volume'), height=500, showlegend=True)
        st.plotly_chart(fig_top, use_container_width=True)

    st.subheader("Catálogo Completo")
    search = st.text_input("🔍 Buscar produto", key="prod_search")

    if is_admin:
        cols = ['code','name','category','abc','total_qty'] + (['valor_12m'] if has_val else [])
        display = products_df[cols].copy()
        if has_val:
            display = display.sort_values('valor_12m', ascending=False)
    else:
        # Vendor view: show % instead of raw quantities
        display = products_df[['code','name','category','abc','total_qty']].copy()
        display['pct_volume'] = (display['total_qty'] / total_qty * 100).round(2)
        display = display.drop(columns=['total_qty'])

    if search:
        s = search.lower()
        display = display[
            display['name'].str.lower().str.contains(s, na=False) |
            display['code'].str.lower().str.contains(s, na=False) |
            display['category'].str.lower().str.contains(s, na=False)
        ]

    if is_admin:
        display.columns = ['Código','Produto','Categoria','Curva','Volume Total'] + (['Faturamento 12m'] if has_val else [])
        show_money_table(display, ['Faturamento 12m'], use_container_width=True, hide_index=True, height=500)
    else:
        display.columns = ['Código','Produto','Categoria','Curva','% do Volume']
        st.dataframe(display, use_container_width=True, hide_index=True, height=500)

    # ---- ANÁLISE DE GAP GLOBAL (visão de portfólio do gestor) ----
    if is_admin and len(df_sku) > 0:
        st.divider()
        st.subheader("📉 Análise de Gap — Potencial da Base Inteira")
        st.caption("Produtos com 5+ compradores: se os demais compradores chegassem ao nível do top 25%, "
                   "quanto a mais venderíamos por mês. Use para escolher campanhas de produto.")
        sku_cov = max(df_sku['mes'].nunique(), 1)
        preco_map = _preco_medio_map(products_df)
        nome_map = df_sku.groupby('sku')['produto'].first().to_dict()
        per = df_sku.groupby(['sku', 'cod_cliente'])['quantidade'].sum().reset_index()
        gap_rows = []
        for sku_code, grp in per.groupby('sku'):
            if grp['cod_cliente'].nunique() < 5:
                continue
            qs = grp['quantidade'].sort_values(ascending=False)
            n_top = max(1, int(len(qs) * 0.25))
            top25 = qs.head(n_top).mean()
            avg_all = qs.mean()
            extra_q = (top25 - avg_all) * len(qs) / sku_cov
            prc = preco_map.get(str(sku_code), 0)
            gap_rows.append({
                'Produto': nome_map.get(sku_code, ''),
                'SKU': sku_code,
                'Clientes': grp['cod_cliente'].nunique(),
                'Média/Cliente': round(avg_all, 1),
                'Média Top 25%': round(top25, 1),
                'Qtd Extra/Mês': round(extra_q, 1),
                'R$ Potencial/Mês': round(extra_q * prc, 2),
            })
        if gap_rows:
            gap_df = pd.DataFrame(gap_rows).sort_values('R$ Potencial/Mês', ascending=False)
            show_money_table(gap_df.head(30), ['R$ Potencial/Mês'], use_container_width=True,
                             hide_index=True, height=500)
            _csv_download(gap_df, "⬇️ Baixar análise completa (Excel/CSV)", "gap_portfolio.csv", "dl_gap")

# ============================================================
# PAGE: ADMIN
# ============================================================
def page_admin(vendor_options=None):
    if not st.session_state.get('authenticated') or st.session_state.get('role') != 'admin':
        st.error("A administração está disponível somente para administradores.")
        return
    vendor_options = vendor_options or []
    st.header("⚙️ Administração")

    st.subheader("Gerenciar Usuários")
    users = load_users()

    # Show current users
    user_data = []
    for username, info in users["users"].items():
        user_data.append({
            "Usuário": username,
            "Nome": info["name"],
            "Papel": info["role"],
            "Filtro Vendedor": (info.get("vendor_filter") or "⚠️ Sem carteira") if info.get("role") == "vendedor" else "Não se aplica"
        })
    st.dataframe(pd.DataFrame(user_data), use_container_width=True, hide_index=True)

    st.divider()

    # Add new user
    st.subheader("Adicionar Novo Usuário")
    with st.form("add_user"):
        col1, col2 = st.columns(2)
        with col1:
            new_username = st.text_input("Usuário (login)")
            new_name = st.text_input("Nome completo")
        with col2:
            new_password = st.text_input("Senha", type="password")
            new_role = st.selectbox("Papel", ["vendedor", "admin"])

        new_vendor = st.selectbox("Carteira do vendedor", vendor_options, index=None,
                                   placeholder="Selecione a carteira",
                                   help="Obrigatória para vendedor. Administradores têm acesso completo.")

        if st.form_submit_button("Adicionar Usuário", type="primary"):
            error = _new_user_error(users, new_username, new_name, new_password,
                                    new_role, new_vendor, vendor_options)
            if error:
                st.error(error)
            else:
                new_username = new_username.strip().lower()
                users["users"][new_username] = {
                    "name": new_name.strip(),
                    "password": hash_password(new_password),
                    "role": new_role,
                    "vendor_filter": new_vendor if new_role == 'vendedor' else None
                }
                save_users(users)
                st.success(f"Usuário '{new_username}' criado com sucesso!")
                st.rerun()

    vendor_users = [username for username, info in users['users'].items() if info.get('role') == 'vendedor']
    if vendor_users:
        st.divider()
        st.subheader("Ajustar Carteira de Vendedor")
        with st.form("change_vendor"):
            wallet_user = st.selectbox("Vendedor (login)", vendor_users)
            wallet_vendor = st.selectbox("Nova carteira", vendor_options, index=None,
                                         placeholder="Selecione a carteira")
            st.caption("O vendedor verá a carteira atualizada na próxima interação com o app neste servidor.")
            if st.form_submit_button("Salvar Carteira"):
                error = _access_configuration_error('vendedor', wallet_vendor, vendor_options)
                if error:
                    st.error(error)
                else:
                    users['users'][wallet_user]['vendor_filter'] = wallet_vendor
                    save_users(users)
                    st.success(f"Carteira de '{wallet_user}' ajustada.")
                    st.rerun()

    st.divider()

    # Change password
    st.subheader("Alterar Senha")
    with st.form("change_pwd"):
        pwd_user = st.selectbox("Usuário", list(users["users"].keys()))
        new_pwd = st.text_input("Nova senha", type="password", key="new_pwd")
        if st.form_submit_button("Alterar Senha"):
            if len(new_pwd) < 12:
                st.error("Use uma senha com pelo menos 12 caracteres.")
            else:
                users["users"][pwd_user]["password"] = hash_password(new_pwd)
                save_users(users)
                st.success(f"Senha de '{pwd_user}' alterada!")

    st.divider()

    # ============================================================
    # DIAGNÓSTICO DE PERSISTÊNCIA (GitHub)
    # ============================================================
    st.subheader("🔌 Persistência (salvamento permanente)")
    if _gh_token():
        st.caption("Há um token configurado. Clique para testar se ele realmente salva no GitHub — "
                   "o teste tenta gravar um arquivo de verdade no branch 'state'.")
    else:
        st.warning("Nenhum token detectado nos secrets agora. Inativações, usuários e logs "
                   "serão perdidos no próximo reinício. Configure o GITHUB_TOKEN (ver COMO-USAR.md) "
                   "e use o botão abaixo para validar.")
    if st.button("🔍 Testar conexão com o GitHub", key="btn_gh_diag"):
        with st.spinner("Testando..."):
            resultados = _gh_diagnose()
        for ok, titulo, detalhe in resultados:
            if ok:
                st.success(f"**{titulo}** — {detalhe}")
            else:
                st.error(f"**{titulo}** — {detalhe}")
        if resultados and resultados[-1][0] and 'Escrita' in resultados[-1][1]:
            st.info("Tudo certo! Agora refaça as inativações uma vez — daqui em diante elas ficam permanentes.")

    st.divider()

    # Upload new data
    st.subheader("Atualizar Base de Dados")
    if _gh_token():
        st.caption("✅ Persistência ativada: a planilha enviada é salva no GitHub e sobrevive a reinícios do servidor.")
    else:
        st.warning("⚠️ GITHUB_TOKEN não configurado nos secrets do Streamlit Cloud. "
                   "O upload vale só até o próximo reinício do servidor — para tornar permanente, "
                   "salve a planilha na pasta do projeto e rode o deploy.bat, ou configure o token (ver COMO-USAR.md).")
    uploaded = st.file_uploader("Envie a planilha atualizada (.xlsx)", type=['xlsx'])
    if uploaded:
        pushed = _handle_planilha_upload(uploaded)
        if pushed:
            st.success("Planilha atualizada e salva no GitHub! O app pode reiniciar em ~1 min para aplicar — os dados ficam permanentes.")
        else:
            st.success("Planilha atualizada nesta sessão! (Sem token GitHub: será perdida no próximo reinício.)")
        st.rerun()

    st.divider()

    # ============================================================
    # ACCESS MONITORING PANEL
    # ============================================================
    st.subheader("📡 Monitoramento de Acessos")
    st.caption("Acompanhe quando e como o time está utilizando o sistema")

    access_log = _load_access_log()
    df_log = pd.DataFrame(access_log) if access_log else pd.DataFrame()

    if df_log.empty or not {'user', 'date', 'action'}.issubset(df_log.columns):
        st.info("Nenhum acesso registrado ainda. Os logs começam a ser gerados a partir do próximo login.")
    else:
        # --- Filters ---
        log_col1, log_col2 = st.columns(2)
        with log_col1:
            _all_users_log = sorted(df_log['user'].unique().tolist())
            _filter_users = st.multiselect("Filtrar por usuário", options=_all_users_log, default=_all_users_log, key="log_filter_users")
        with log_col2:
            _all_dates = sorted(df_log['date'].unique().tolist(), reverse=True)
            _default_dates = _all_dates[:7] if len(_all_dates) > 7 else _all_dates
            _filter_dates = st.multiselect("Filtrar por data", options=_all_dates, default=_default_dates, key="log_filter_dates")

        df_filtered = df_log[df_log['user'].isin(_filter_users) & df_log['date'].isin(_filter_dates)].copy()

        # --- KPIs ---
        _logins = df_filtered[df_filtered['action'] == 'login']
        _page_views = df_filtered[df_filtered['action'] == 'page_view']
        _unique_users = _logins['user'].nunique()
        _total_logins = len(_logins)
        _total_views = len(_page_views)

        mk1, mk2, mk3 = st.columns(3)
        mk1.metric("👥 Usuários Ativos", f"{_unique_users}")
        mk2.metric("🔑 Total de Logins", f"{_total_logins}")
        mk3.metric("📄 Páginas Visitadas", f"{_total_views}")

        # --- Logins per user per day ---
        if len(_logins) > 0:
            st.markdown("**📊 Logins por Usuário por Dia**")
            login_pivot = _logins.groupby(['date', 'name']).size().reset_index(name='logins')
            fig_logins = px.bar(login_pivot, x='date', y='logins', color='name',
                               title="Histórico de Logins",
                               labels={'date': 'Data', 'logins': 'Logins', 'name': 'Usuário'},
                               barmode='group')
            fig_logins.update_layout(template='plotly_white', paper_bgcolor='#ffffff',
                                    plot_bgcolor='#ffffff', height=350)
            st.plotly_chart(fig_logins, use_container_width=True)

        # --- Pages most visited ---
        if len(_page_views) > 0:
            st.markdown("**📄 Páginas Mais Acessadas por Usuário**")
            page_usage = _page_views.groupby(['name', 'page']).size().reset_index(name='visitas')
            page_usage = page_usage.sort_values('visitas', ascending=False)
            st.dataframe(page_usage.rename(columns={'name': 'Usuário', 'page': 'Página', 'visitas': 'Visitas'}),
                        use_container_width=True, hide_index=True, height=300)

        # --- Full log table ---
        st.markdown("**📋 Log Completo de Acessos**")
        display_log = df_filtered.sort_values('timestamp', ascending=False)
        display_cols = ['timestamp', 'name', 'action']
        if 'page' in display_log.columns:
            display_cols.append('page')
        display_log_show = display_log[display_cols].copy()
        col_names = {'timestamp': 'Data/Hora', 'name': 'Usuário', 'action': 'Ação', 'page': 'Página'}
        display_log_show = display_log_show.rename(columns=col_names)
        display_log_show['Ação'] = display_log_show['Ação'].replace({'login': '🔑 Login', 'page_view': '📄 Página'})
        st.dataframe(display_log_show, use_container_width=True, hide_index=True, height=400)

# ============================================================
# MAIN APP
# ============================================================
def main():
    # Zera o memo de leitura de estado a cada rerun (cada run relê o GitHub 1x por arquivo)
    _STATE_RAW_CACHE.clear()
    # Restaura estado persistido no GitHub (1x por boot do container) — fallback local
    _sync_state_from_github()

    # Autenticação: SÓ sessão do servidor. Links antigos com ?u=&t= não logam
    # ninguém — os params são apenas apagados da URL.
    _strip_stale_auth_params()
    if not st.session_state.get("authenticated"):
        login_page()
        return
    # Sessão expira por inatividade (3h) ou tempo máximo (12h) → novo login
    if _session_expired():
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.warning("🔒 Sua sessão expirou por segurança. Faça login novamente.")
        login_page()
        return
    _touch_session()

    access_error = _refresh_session_access()
    if access_error:
        _show_access_denied(access_error)
        return

    # Load data
    result = load_data()
    if result[0] is None or len(result) < 6:
        st.warning("Não foi possível carregar os dados. Verifique se o arquivo Excel está na pasta do app.")
        if st.session_state.get("role") == "admin":
            st.subheader("Upload da planilha")
            uploaded = st.file_uploader("Envie a planilha (.xlsx)", type=['xlsx'])
            if uploaded:
                pushed = _handle_planilha_upload(uploaded)
                if pushed:
                    st.success("Planilha salva no GitHub! Recarregando...")
                else:
                    st.success("Planilha salva nesta sessão! Recarregando...")
                st.rerun()
        return

    df_clients, df_products, df_client_products, months, year_ranges, df_sku = result

    vendor_options = _vendor_options(df_clients)
    try:
        df_clients = _clients_for_access(df_clients, st.session_state.get('role'),
                                         st.session_state.get('vendor_filter'))
    except ValueError as error:
        _show_access_denied(str(error))
        return

    # --- Parse unique years and month names from labels (e.g. "jan/21") ---
    all_years_ordered = []
    _seen_years = set()
    for lbl in months:
        parts = lbl.replace('-', '/').split('/')
        if len(parts) >= 2:
            y_raw = parts[-1].strip()
            y_full = f"20{y_raw}" if len(y_raw) == 2 else y_raw
        else:
            y_full = ""
        if y_full and y_full not in _seen_years:
            all_years_ordered.append(y_full)
            _seen_years.add(y_full)

    # ========== COMPACT SIDEBAR ==========

    # --- Map month names to numbers (jan=1, fev=2, ...) ---
    MONTH_NAME_TO_NUM = {'jan':'1','fev':'2','mar':'3','abr':'4','mai':'5','jun':'6',
                         'jul':'7','ago':'8','set':'9','out':'10','nov':'11','dez':'12'}
    # Build unique month numbers that exist in data
    _existing_month_nums = []
    _seen_mnums = set()
    for lbl in months:
        parts = lbl.replace('-', '/').split('/')
        if len(parts) >= 2:
            m_name = parts[0].strip().lower()
            m_num = MONTH_NAME_TO_NUM.get(m_name, m_name)
            if m_num not in _seen_mnums:
                _existing_month_nums.append(m_num)
                _seen_mnums.add(m_num)
    # Sort numerically
    _existing_month_nums.sort(key=lambda x: int(x) if x.isdigit() else 0)

    # Default year: most recent with 6+ months of data
    _best_default_year = all_years_ordered[-1] if all_years_ordered else ""
    for yr in reversed(all_years_ordered):
        yr_months = [lbl for lbl in months if lbl.split('/')[-1].strip() in [yr[-2:], yr]]
        if len(yr_months) >= 6:
            _best_default_year = yr
            break

    with st.sidebar:
        # --- User greeting (compact) ---
        _role = st.session_state['role']
        ui.identity(st.session_state['user_name'], _role)

        # --- Navigation ---
        if st.session_state["role"] in ("garantia", "garantia_master"):
            pages = {"🔧 Garantias": "garantia"}
        elif has_full_data_access():
            pages = {
                "◉ Hoje · Agenda": "agenda",
                "🎛️ Painel do Gestor": "manager",
                "🔴 Mês ao Vivo": "mesvivo",
                "✅ Ações do Time": "actions",
                "📊 Visão Geral": "overview",
                "👤 Clientes": "clients",
                "🧩 Mix de Produtos": "mix",
                "⚠️ Churn": "churn",
                "📦 Produtos": "products",
                "🔧 Garantias": "garantia",
            }
            if st.session_state["role"] == "admin":
                pages["⚙️ Admin"] = "admin"
        else:
            pages = {
                "◉ Hoje · Minha Agenda": "agenda",
                "✅ Minhas Ações": "actions",
                "🔴 Meu Mês ao Vivo": "mesvivo",
                "📊 Minha Visão Geral": "overview",
                "👤 Meus Clientes": "clients",
                "🧩 Mix de Produtos": "mix",
                "⚠️ Churn": "churn",
                "📦 Produtos": "products",
            }

        selected_page = st.radio("Navegação", list(pages.keys()), label_visibility="collapsed")

        st.markdown("---")

        with st.expander('Período das análises', expanded=pages[selected_page] not in ('agenda', 'garantia')):
            st.caption('Aplica-se às análises mensais. A agenda usa as datas dos retornos.')
            # --- Period Filter: compact multiselects ---
            st.markdown("**📅 Período**")

            # Chart click override indicator
            chart_override_active = "chart_sel_months" in st.session_state and st.session_state["chart_sel_months"]
            if chart_override_active:
                n_chart = len(st.session_state["chart_sel_months"])
                st.info(f"📊 Seleção via gráfico ({n_chart} {'mês' if n_chart == 1 else 'meses'})")
                if st.button("✕ Limpar seleção", use_container_width=True, key="clear_chart"):
                    del st.session_state["chart_sel_months"]
                    st.rerun()

            # Seletor único com presets (substitui os multiselects de ano + mês)
            _cur_year = all_years_ordered[-1] if all_years_ordered else ""
            _prev_year = all_years_ordered[-2] if len(all_years_ordered) >= 2 else ""
            preset_options = ["Últimos 12 meses"]
            if _cur_year:
                preset_options.append(f"Este ano ({_cur_year})")
            preset_options += ["Últimos 6 meses", "Últimos 3 meses", "Este mês"]
            if _prev_year:
                preset_options.append(f"Ano passado ({_prev_year})")
            preset_options += ["Tudo (desde o início)", "Personalizado…"]

            sel_preset = st.selectbox("Período", preset_options, key="global_preset",
                                      label_visibility="collapsed")

            # Modo avançado: escolher anos e meses manualmente
            selected_years, selected_month_nums = [], []
            if sel_preset == "Personalizado…":
                selected_years = st.multiselect(
                    "Ano",
                    options=all_years_ordered,
                    default=[_best_default_year] if _best_default_year else [],
                    key="global_years"
                )
                selected_month_nums = st.multiselect(
                    "Mês (1-12)",
                    options=_existing_month_nums,
                    default=_existing_month_nums,
                    key="global_months"
                )

            st.divider()


        with st.expander('Resumo da carteira'):
            _card_active = df_clients[_commercial_active_mask(df_clients, load_inactive_clients())]
            st.write(f"{len(_card_active)} clientes ativos · {len(df_clients) - len(_card_active)} inativos / outros")
            _risk_counts = _card_active['risk'].value_counts() if 'risk' in _card_active.columns else {}
            st.caption(f"Saudáveis: {_risk_counts.get('Saudável', 0)} · Atenção: {_risk_counts.get('Atenção', 0)} · Recuperação: {_risk_counts.get('Recuperação', 0)}")
        st.divider()


        if st.button("🚪 Sair", use_container_width=True):
            _strip_stale_auth_params()
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

    # --- Build selected indices ---
    # Chart click override takes priority over sidebar filters
    chart_override_active = "chart_sel_months" in st.session_state and st.session_state["chart_sel_months"]

    if chart_override_active:
        sel_indices = set()
        for lbl in st.session_state["chart_sel_months"]:
            if lbl in months:
                sel_indices.add(months.index(lbl))
    elif sel_preset != "Personalizado…":
        sel_indices = compute_preset_indices(sel_preset, months)
    else:
        # Personalizado: match months by year + month number
        sel_indices = set()
        for i, lbl in enumerate(months):
            parts = lbl.replace('-', '/').split('/')
            if len(parts) >= 2:
                m_name = parts[0].strip().lower()
                y_raw = parts[-1].strip()
                y_full = f"20{y_raw}" if len(y_raw) == 2 else y_raw
                m_num = MONTH_NAME_TO_NUM.get(m_name, "")
            else:
                y_full = ""
                m_num = ""
            if y_full in selected_years and m_num in selected_month_nums:
                sel_indices.add(i)

    # Fallback: if nothing selected, select all
    if not sel_indices:
        sel_indices = set(range(len(months)))

    sel_indices_sorted = sorted(sel_indices)
    sel_months = [months[i] for i in sel_indices_sorted]

    if pages[selected_page] != 'agenda':
        ui.page_hero('PROPETZ BI', 'Clareza para decidir.', 'Distribuição e assistência em um só lugar.',
                     f'Dados: {months[0]} a {months[-1]} · {len(df_clients)} clientes')

    # Alerta de persistência: sem GITHUB_TOKEN, inativações/usuários/logs são
    # perdidos quando o servidor reinicia. Visível só para admin/diretor.
    if has_full_data_access() and not _gh_token():
        st.error(
            "⚠️ **Persistência DESATIVADA.** As inativações de clientes, novos usuários e o "
            "log de acessos ficam só na memória temporária do servidor e **são perdidos quando "
            "ele reinicia**. Para tornar permanente, configure o `GITHUB_TOKEN` nos secrets do "
            "Streamlit Cloud (passo a passo no COMO-USAR.md). Enquanto isso não for feito, evite "
            "depender das inativações."
        )
    elif has_full_data_access():
        # Aviso ANTES de vencer: o token tem validade e, se expirar sem troca,
        # o app para de salvar. A data vem de graça do header do GitHub.
        _dias = _dias_p_expirar_token()
        if _dias is not None and _dias <= 30:
            _txt = (f"expira em **{_dias} dia(s)** ({_TOKEN_EXPIRA[0]})" if _dias > 0
                    else f"**EXPIROU** em {_TOKEN_EXPIRA[0]}")
            st.warning(
                f"🔑 **O token do GitHub {_txt}.** Quando vencer, o app PARA de salvar "
                "inativações, garantias e log. Gere um novo token fine-grained e troque nos "
                "Secrets do Streamlit (passo a passo no COMO-USAR.md)."
            )

    # Route to page
    page = pages[selected_page]

    # Log page view (only once per page per session to avoid spam)
    _page_log_key = f"_logged_page_{page}"
    if _page_log_key not in st.session_state:
        st.session_state[_page_log_key] = True
        log_page_view(st.session_state.get("username", ""), selected_page)

    if page == "agenda":
        page_agenda(df_clients, months)
    elif page == "garantia":
        page_garantias(df_products, df_clients)
    elif page == "mesvivo":
        page_mes_vivo()
    elif page == "manager":
        page_manager(df_clients, months, df_sku, df_products)
    elif page == "actions":
        page_actions(df_clients, df_sku, df_products, df_client_products, months)
    elif page == "overview":
        page_overview(df_clients, months, year_ranges, sel_indices, sel_indices_sorted, sel_months)
    elif page == "clients":
        page_clients(df_clients, df_sku, months, year_ranges, sel_indices_sorted, sel_months)
    elif page == "mix":
        page_mix(df_clients, df_products, df_client_products, df_sku, months, sel_indices_sorted, sel_months)
    elif page == "churn":
        page_churn(df_clients, months, sel_indices_sorted, sel_months)
    elif page == "products":
        page_products(df_products, df_sku)
    elif page == "admin":
        page_admin(vendor_options)

if __name__ == "__main__":
    main()
