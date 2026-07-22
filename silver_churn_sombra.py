# -*- coding: utf-8 -*-
"""SOMBRA do churn — banco silver vs planilha do BI (NADA muda no app). v2.

v2 pós-auditoria adversarial de 21/07/2026 (13 achados confirmados):
  - norm() agressiva: remove sufixos de cadastro do Tiny ("| SN", "- RN", "|"),
    TODA pontuação, tokens só-dígitos (CPF/CNPJ embutido) e stopwords societárias
    (LTDA/ME/EIRELI/SA...) — antes, metade da receita ficava sem código.
  - Casamento em 4 camadas (exato → contido → conjunto de tokens → cobertura por
    prefixo de tokens com âncora no 1º token), sempre exigindo candidato ÚNICO;
    ambiguidade agora é LISTADA (antes era descartada em silêncio — caso J.W).
  - Régua do APP: meses ancorados no ÚLTIMO MÊS CARREGADO da planilha (o app
    conta por índice de coluna, não por "hoje") e mês-calendário nos 2 lados
    (dias/30,44 criava divergência de borda fantasma).
  - Inativados (origin/state:inactive_clients.json) e Status marcados;
    código -1 (placeholder) e 1603 (UF inválida, app descarta) tratados.
Saídas locais (fora do repo): depara_clientes_silver.json,
Relatorio_Sombra_Churn.md, divergencias_churn.csv.
"""
import sys, os, json, csv, re, shutil, subprocess, tempfile, unicodedata
from datetime import date
sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
from ponte_db_silver import consultar
import openpyxl
from datetime import datetime

HOJE = date.today()  # rotina diária: sempre a data corrente
PLANILHA = os.path.join(BASE, "Relatorio Distribuidores Mensal.xlsx")
MESES_PT = {"jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
            "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12}
STOP = {"LTDA", "ME", "MEI", "EIRELI", "EPP", "CIA", "LT", "SA",
        # conectores: sem identidade, e "E" com a regra de inicial viraria curinga
        "E", "DE", "DA", "DO", "DAS", "DOS", "PARA"}


def norm(nome):
    """Normalização agressiva p/ casar variantes de cadastro do MESMO cliente."""
    s = str(nome or "").strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"[\|\-/]+\s*(SN|RN)\s*$", " ", s)   # sufixo de regime tributário
    s = re.sub(r"[\|]+\s*$", " ", s)                  # "|" sobrando no fim
    s = s.replace("S/A", " SA ").replace("S.A", " SA ")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)                 # toda pontuação vira espaço
    toks = [t for t in s.split()
            if not (t.isdigit() and len(t) >= 4) and t not in STOP]
    return " ".join(toks)


def toks_de(n):
    return n.split()


def tok_casa(a, b):
    """Token igual, um prefixo do outro (MAN~MANUT) ou inicial única (S~SILVA)."""
    if a == b:
        return True
    if len(a) >= 3 and len(b) >= 3:
        return a.startswith(b) or b.startswith(a)
    if len(a) == 1 and len(b) >= 3:
        return b.startswith(a)
    if len(b) == 1 and len(a) >= 3:
        return a.startswith(b)
    return False


def cobre(curto, longo):
    """Todo token do lado curto casa (prefixo) com algum token do lado longo."""
    usados = set()
    for t in curto:
        achou = None
        for i, u in enumerate(longo):
            if i not in usados and tok_casa(t, u):
                achou = i
                break
        if achou is None:
            return False
        usados.add(achou)
    return True


def classifica(meses):
    if meses is None:
        return "sem compra"
    if meses >= 6:
        return "Recuperação"
    if meses >= 3:
        return "Atenção"
    return "Saudável"


# ---------------- 1. PLANILHA (o que o app vê hoje) ----------------
tmp = os.path.join(tempfile.gettempdir(), "_sombra_rel.xlsx")
shutil.copy2(PLANILHA, tmp)
wb = openpyxl.load_workbook(tmp, data_only=True, read_only=True)
ws = wb["Propetz"]
linhas = list(ws.iter_rows(values_only=True))
cab = linhas[1]

meses_cols = []
for idx, v in enumerate(cab):
    ym = None
    if isinstance(v, datetime):
        ym = (v.year, v.month)
    elif isinstance(v, str) and "/" in v:
        p = v.strip().lower().split("/")
        if len(p) == 2 and p[0][:3] in MESES_PT:
            try:
                ym = (2000 + int(p[1]), MESES_PT[p[0][:3]])
            except ValueError:
                pass
    if ym:
        meses_cols.append((idx, ym))

planilha = {}
ancora = None  # último mês com DADO em qualquer cliente = âncora do app
for row in linhas[2:]:
    nome, cod = row[1], row[3]
    if not nome or cod is None:
        continue
    ultima, fat_total = None, 0.0
    for idx, ym in meses_cols:
        v = row[idx] if idx < len(row) else None
        if isinstance(v, (int, float)) and v > 0:
            fat_total += v
            if ultima is None or ym > ultima:
                ultima = ym
            if ancora is None or ym > ancora:
                ancora = ym
    planilha[str(cod).strip()] = {
        "nome": str(nome).strip(), "norm": norm(nome), "uf": str(row[2] or "").strip().upper(),
        "vendedor": str(row[5] or "").strip(), "status": str(row[6] or "").strip(),
        "ultima_ym": ultima, "fat_total": fat_total}

# STATUS que o app exibe vem da aba IA (contraprova: Propetz diverge em 105 códigos)
try:
    ws_ia = wb["IA"]
    ia = list(ws_ia.iter_rows(values_only=True))
    hdr_i = next(i for i, r in enumerate(ia[:6]) if r and "ID Cliente" in [str(v) for v in r])
    hdr = [str(v) for v in ia[hdr_i]]
    c_id, c_st = hdr.index("ID Cliente"), hdr.index("Status")
    for r in ia[hdr_i + 1:]:
        cid = r[c_id] if c_id < len(r) else None
        if cid is not None and str(cid).strip() in planilha:
            planilha[str(cid).strip()]["status"] = str(r[c_st] or "").strip()
except Exception as e:
    print(f"aviso: status da aba IA indisponível ({e}) — usando o da Propetz")

# universo do APP: a aba IA descarta UF inválida (ex. 1603 'TJ') — replicar
_UFS = {"AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS","MG","PA",
        "PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO","EX"}
fora_do_app = {c for c, p in planilha.items() if p["uf"] and p["uf"] not in _UFS}
if fora_do_app:
    print(f"fora do universo do app (UF inválida): {sorted(fora_do_app)}")
print(f"planilha: {len(planilha)} clientes | âncora do app (último mês carregado): "
      f"{ancora[1]:02d}/{ancora[0]}")

# classe segundo a RÉGUA DO APP (ancorada no último mês carregado)
for cod, p in planilha.items():
    if p["ultima_ym"]:
        p["meses_app"] = (ancora[0] * 12 + ancora[1]) - (p["ultima_ym"][0] * 12 + p["ultima_ym"][1])
    else:
        p["meses_app"] = None
    p["classe_app"] = classifica(p["meses_app"])

# inativados no app (fonte da verdade: branch state; fallback local)
inativos = set()
try:
    subprocess.run(["git", "-C", BASE, "fetch", "origin", "state"],
                   capture_output=True, timeout=60)
    out = subprocess.run(["git", "-C", BASE, "show", "origin/state:inactive_clients.json"],
                         capture_output=True, text=True, timeout=30)
    if out.returncode == 0:
        inativos = {str(i) for i in json.loads(out.stdout).get("inactive_ids", [])}
except Exception:
    pass
if not inativos:
    try:
        loc = json.load(open(os.path.join(BASE, "inactive_clients.json"), encoding="utf-8"))
        inativos = {str(i) for i in loc.get("inactive_ids", [])}
    except Exception:
        pass
print(f"inativados no app: {len(inativos)}")

# ---------------- 2. BANCO (a fonte, 2026+) ----------------
banco_rows = consultar("""
    SELECT cliente_nome,
           max(data_emissao)  AS ultima_compra,
           count(*)           AS notas,
           round(sum(valor_total)::numeric, 2) AS receita,
           max(cidade)        AS cidade,
           max(uf)            AS uf,
           max(vendedor_nome) AS vendedor
    FROM silver.faturamento
    WHERE ano >= 2026 AND tipo_faturamento = 'NF de Venda'
      AND modelo_negocio_descricao = 'Distribuição PROPETZ'
    GROUP BY cliente_nome
""")
print(f"banco: {len(banco_rows)} cadastros distintos (2026+)")

# ---------------- 3. DE-PARA nome(banco) -> código ----------------
_depara_arq = os.path.join(BASE, "depara_clientes_silver.json")
_manuais = {}
if os.path.exists(_depara_arq):
    try:
        _ant = json.load(open(_depara_arq, encoding="utf-8"))
        _manuais = {k: v for k, v in _ant.items() if v.get("como") == "manual"}
    except Exception:
        pass

por_norm = {}
for cod, p in planilha.items():
    por_norm.setdefault(p["norm"], []).append(cod)

# CAMADA 0 — TABELA BASE do Leonardo (21/07): a AUTORIDADE nome->código.
# "Mesmo que tenha 25 nomes diferentes, é aquele código que vale."
base_map = {}
_tb = os.path.join(BASE, "Tabela Base clientes distribuição.xlsx")
if os.path.exists(_tb):
    _tmpb = os.path.join(tempfile.gettempdir(), "_tabela_base_cli.xlsx")
    shutil.copy2(_tb, _tmpb)
    wsb = openpyxl.load_workbook(_tmpb, data_only=True, read_only=True)["Sheet1"]
    for row in wsb.iter_rows(min_row=6, values_only=True):
        if row[2] and row[4] is not None:
            base_map.setdefault(norm(row[2]), set()).add(str(row[4]).strip())
    print(f"tabela base: {sum(len(v) for v in base_map.values())} vínculos, "
          f"{len(base_map)} nomes normalizados")

for _nm, _v in _manuais.items():
    _nb = norm(_nm)
    if _nb in base_map and _v["codigo"] not in base_map[_nb]:
        print(f"AVISO: manual '{_nm[:40]}'->{_v['codigo']} conflita com a "
              f"tabela base {sorted(base_map[_nb])} — a base é a autoridade, conferir!")

depara, sem_match, ambiguos = dict(_manuais), [], []
for b in banco_rows:
    if b["cliente_nome"] in depara:
        continue
    nb = norm(b["cliente_nome"])
    tb = toks_de(nb)
    cands, como = [], None
    # camada 0: TABELA BASE (autoridade — vence qualquer heurística)
    if nb in base_map:
        cands, como = sorted(base_map[nb]), "tabela-base"
    # camada 1: exato normalizado
    if not cands and nb in por_norm:
        cands, como = por_norm[nb], "exato-normalizado"
    # camada 2: contido (um dentro do outro)
    if not cands and len(nb) >= 8:
        cs = {c for c, p in planilha.items() if nb in p["norm"] or p["norm"] in nb}
        if cs:
            cands, como = list(cs), "contido"
    # camada 3/4: tokens — âncora no 1º token + cobertura por prefixo.
    # EXIGE UF igual e vendedor compatível (contraprova 21/07: sem isso, a
    # regra de inicial única colou 6 clientes errados — JAIME→J.HENRIQUE etc.)
    b_uf = str(b["uf"] or "").strip().upper()
    b_vd = str(b.get("vendedor") or "").strip().split(" ")[0].lower()
    if not cands and len(tb) >= 2:
        cs = set()
        for c, p in planilha.items():
            if b_uf and p["uf"] and b_uf != p["uf"]:
                continue
            p_vd = p["vendedor"].split(" ")[0].lower()
            if b_vd and p_vd and b_vd != p_vd:
                continue
            tp = toks_de(p["norm"])
            if len(tp) >= 2 and tok_casa(tb[0], tp[0]):
                curto, longo = (tb, tp) if len(tb) <= len(tp) else (tp, tb)
                if cobre(curto, longo):
                    cs.add(c)
        if cs:
            cands, como = list(cs), "tokens+uf"
    cands = sorted(set(cands))
    # desempate de ambiguidade por UF (contraprova: Mega Pet tinha resposta óbvia)
    if len(cands) > 1 and b_uf:
        so_uf = [c for c in cands if planilha.get(c, {}).get("uf") == b_uf]
        if len(so_uf) == 1:
            cands, como = so_uf, (como or "") + "+desempate-uf"
    # REGRA LEONARDO (21/07): entre códigos duplicados, vale o código do ÚLTIMO
    # pedido (empate de mês: maior faturamento acumulado)
    if len(cands) > 1:
        def _recencia(c):
            p = planilha.get(c, {})
            return (p.get("ultima_ym") or (0, 0), p.get("fat_total", 0.0))
        melhor = max(cands, key=_recencia)
        if sum(1 for c in cands if _recencia(c) == _recencia(melhor)) == 1:
            cands, como = [melhor], (como or "") + "+recencia"
    if len(cands) == 1:
        depara[b["cliente_nome"]] = {"codigo": cands[0], "como": como,
                                     "nome_planilha": planilha[cands[0]]["nome"]}
    elif len(cands) > 1:
        ambiguos.append({"nome_banco": b["cliente_nome"], "codigos": cands,
                         "receita": float(b["receita"])})
        sem_match.append(b)
    else:
        sem_match.append(b)

json.dump(depara, open(_depara_arq, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
n_cod = len({d["codigo"] for d in depara.values()})
print(f"de-para: {len(depara)} nomes -> {n_cod} códigos | sem código: {len(sem_match)} "
      f"| ambíguos: {len(ambiguos)}")

# ---------------- 4. AGREGA por código e COMPARA (régua do app) ----------------
banco_por_cod = {}
for b in banco_rows:
    d = depara.get(b["cliente_nome"])
    if not d:
        continue
    cod = d["codigo"]
    atual = banco_por_cod.get(cod)
    if atual is None:
        banco_por_cod[cod] = dict(b)
    else:
        atual["notas"] += b["notas"]
        atual["receita"] += b["receita"]
        if b["ultima_compra"] > atual["ultima_compra"]:
            atual["ultima_compra"] = b["ultima_compra"]

diverg, frescor, defasagem, iguais = [], [], [], 0
for cod, b in banco_por_cod.items():
    p = planilha.get(cod)
    if p is None or cod in fora_do_app:  # o app nem exibe (UF inválida)
        continue
    u = b["ultima_compra"]
    # MESMA âncora do app (último mês carregado da planilha): compara igual
    # com igual; o mês a mais que o banco conhece vira FRESCOR, não divergência
    meses_banco = max(0, (ancora[0] * 12 + ancora[1]) - (u.year * 12 + u.month))
    cl_banco = classifica(meses_banco)
    # DEFASAGEM: com dado diário, a classe de HOJE já seria outra (o tempo
    # passou desde o último upload) — é o que a atualização diária muda na tela
    meses_hoje = (HOJE.year * 12 + HOJE.month) - (u.year * 12 + u.month)
    cl_hoje = classifica(meses_hoje)
    if cl_banco == p["classe_app"]:
        if cl_hoje != p["classe_app"]:
            defasagem.append({"codigo": cod, "cliente": p["nome"], "status": p["status"],
                              "vendedor": p["vendedor"], "classe_app": p["classe_app"],
                              "classe_hoje": cl_hoje, "ultima_banco": str(u),
                              "receita_2026": float(b["receita"])})
        iguais += 1
        continue
    item = {"codigo": cod, "cliente": p["nome"], "vendedor": p["vendedor"],
            "status": p["status"], "classe_app": p["classe_app"], "classe_banco": cl_banco,
            "ultima_planilha": f"{p['ultima_ym'][1]:02d}/{p['ultima_ym'][0]}" if p["ultima_ym"] else "—",
            "ultima_banco": str(u), "notas_2026": b["notas"],
            "receita_2026": float(b["receita"])}
    banco_alem_da_planilha = (u.year, u.month) > ancora
    (frescor if banco_alem_da_planilha and cl_banco == "Saudável" else diverg).append(item)

# cegos: códigos sem nenhuma nota 2026 casada no banco
cods_ambiguos = {c for a in ambiguos for c in a["codigos"]}
cegos, cegos_saudaveis = [], []
for cod, p in planilha.items():
    if cod in banco_por_cod or cod in fora_do_app:
        continue
    marca = []
    if cod in inativos:
        marca.append("INATIVADO no app")
    if cod == "-1":
        marca.append("código placeholder")
    if cod in cods_ambiguos:
        marca.append("banco TEM nota — retida na seção Ambíguos")
    if p["status"].lower() not in ("ativo",):
        marca.append(f"status: {p['status'] or '?'}")
    cegos.append((cod, p, marca))
    if p["classe_app"] == "Saudável" and cod not in inativos and cod != "-1":
        cegos_saudaveis.append((cod, p, marca))

# candidatos prováveis p/ os sem-código (sem exigir unicidade — só sugestão)
sugestoes = []
skus_casados = set(depara)
for b in sorted(sem_match, key=lambda x: -float(x["receita"])):
    nb = norm(b["cliente_nome"])
    tb = toks_de(nb)
    b_uf = str(b["uf"] or "").strip().upper()
    cands = set()
    for c, p in planilha.items():
        # mesmo filtro de UF do matcher: sem ele, a sugestão apontaria
        # exatamente os falsos positivos que a contraprova derrubou
        if b_uf and p["uf"] and b_uf != p["uf"]:
            continue
        tp = toks_de(p["norm"])
        if tb and tp and tok_casa(tb[0], tp[0]):
            curto, longo = (tb, tp) if len(tb) <= len(tp) else (tp, tb)
            if cobre(curto, longo):
                cands.add(c)
    sugestoes.append((b, sorted(cands)))

# ---------------- 5. RELATÓRIO ----------------
rel = os.path.join(BASE, "Relatorio_Sombra_Churn.md")
with open(rel, "w", encoding="utf-8") as f:
    f.write(f"# Sombra do churn v3 — banco silver vs planilha do BI ({HOJE})\n\n")
    f.write("> NADA foi alterado no app. Régua idêntica à do app (ancorada em "
            f"{ancora[1]:02d}/{ancora[0]}, o último mês carregado). Cliente = CÓDIGO.\n\n")
    f.write(f"- Planilha: **{len(planilha)}** clientes com código "
            f"({len(inativos)} inativados no app)\n")
    f.write(f"- Banco 2026+: **{len(banco_rows)}** cadastros → casados **{len(depara)}** "
            f"nomes = **{n_cod}** códigos | sem código **{len(sem_match)}** "
            f"| ambíguos **{len(ambiguos)}**\n")
    f.write(f"- Classificação IGUAL: **{iguais}** | Frescor (banco viu compra depois de "
            f"{ancora[1]:02d}/{ancora[0]}): **{len(frescor)}** | "
            f"**Divergência real: {len(diverg)}**\n")
    f.write(f"- Sem nota 2026 no banco: **{len(cegos)}** códigos (destes, "
            f"**{len(cegos_saudaveis)}** que o app mostra como Saudáveis — ver seção)\n\n")

    if diverg:
        f.write("## Divergências reais (o app mostra uma classe, a fonte mostra outra)\n\n")
        f.write("| Código | Cliente | Vendedor | Status | App | Banco | Últ. planilha | Últ. NF real |\n")
        f.write("|---|---|---|---|---|---|---|---|\n")
        for d in sorted(diverg, key=lambda x: -x["receita_2026"]):
            f.write(f"| {d['codigo']} | {d['cliente'][:32]} | {d['vendedor'][:18]} | "
                    f"{d['status'][:12]} | {d['classe_app']} | {d['classe_banco']} | "
                    f"{d['ultima_planilha']} | {d['ultima_banco']} |\n")
    if frescor:
        f.write(f"\n## Frescor ({len(frescor)}) — compras que a planilha ainda não recebeu "
                "(o ganho da atualização diária)\n\n")
        for d in sorted(frescor, key=lambda x: -x["receita_2026"]):
            f.write(f"- {d['codigo']} {d['cliente'][:38]}: app '{d['classe_app']}' → "
                    f"banco '{d['classe_banco']}' (última NF {d['ultima_banco']})\n")
    if defasagem:
        f.write(f"\n## Defasagem de calendário ({len(defasagem)}) — dado igual, mas com "
                "atualização diária a classe de HOJE já seria outra\n\n")
        f.write("> A planilha para no mês do último upload; o tempo continua passando. "
                "É isto que a atualização diária resolve sozinha.\n\n")
        for d in sorted(defasagem, key=lambda x: -x["receita_2026"]):
            st = f" [status: {d['status']}]" if d["status"].lower() != "ativo" else ""
            f.write(f"- {d['codigo']} {d['cliente'][:38]}: tela mostra '{d['classe_app']}', "
                    f"hoje já seria '{d['classe_hoje']}' (última NF {d['ultima_banco']}){st}\n")
    if cegos_saudaveis:
        f.write("\n## App diz 'Saudável', banco sem nota 2026 — conferir 1 a 1\n\n")
        for cod, p, marca in cegos_saudaveis:
            m = f" [{'; '.join(marca)}]" if marca else ""
            ult = f"{p['ultima_ym'][1]:02d}/{p['ultima_ym'][0]}" if p["ultima_ym"] else "?"
            f.write(f"- {cod} {p['nome'][:40]} (últ. planilha {ult}){m}\n")
    if ambiguos:
        f.write("\n## Ambíguos — ARBITRAGEM DO LEONARDO (1 nome do banco casa com 2+ códigos)\n\n")
        for a in ambiguos:
            f.write(f"- '{a['nome_banco'][:45]}' (R$ {a['receita']:,.2f}) ↔ códigos "
                    f"{', '.join(a['codigos'])}\n")
    # clientes NOVOS confirmados pelo Leonardo (21/07): cadastro no fechamento
    NOVOS_PENDENTES = ("SIGMAVET", "DESENCAIXE DIGITAL", "MAJOR PET")
    if sugestoes:
        f.write("\n## Sem código no de-para — novos e pendentes\n\n")
        for b, cands in sugestoes:
            nb2 = norm(b["cliente_nome"])
            if any(nb2.startswith(x) for x in NOVOS_PENDENTES):
                cand = " | ✅ CLIENTE NOVO — Leonardo cadastra no fechamento do mês"
            elif cands:
                cand = " | candidato: " + ", ".join(f"{c} {planilha[c]['nome'][:25]}" for c in cands[:2])
            else:
                cand = " | SEM candidato na planilha (conferir)"
            f.write(f"- {b['cliente_nome'][:48]} (R$ {float(b['receita']):,.2f}, "
                    f"últ. {b['ultima_compra']}, {b['cidade']}/{b['uf']}){cand}\n")
    f.write("\n## Observações de qualidade de cadastro (achadas pela auditoria)\n\n")
    f.write("- Códigos duplicados ARBITRADOS pela regra do Leonardo (código do último "
            "pedido vence): J.W→1257, FZ PET→1105, TRES PETS→1543, V.S.FONTES→1540, "
            "ARQUELINO→1230, CAMILLA→1608, UNIVERSO PET→1140. Os códigos perdedores "
            "seguem na planilha como histórico (aparecem 'sem nota 2026').\n")
    f.write("- UF errada na planilha: 1581 'RN'→BA (Itaberaba); 1603 'TJ'→RJ "
            "(São Gonçalo — o app DESCARTA este cliente por UF inválida).\n")
    f.write("- Filiais somadas na matriz (validar): 1240 André Shatoshi (filial PR), "
            "1442 Pup's Vet (filial TO).\n")
    f.write("- 'Life Distribuidora De Ferragens' tem código -1 (placeholder) e é canal "
            "NR — sempre ficará fora do filtro Distribuição PROPETZ.\n")

with open(os.path.join(BASE, "divergencias_churn.csv"), "w", encoding="utf-8-sig",
          newline="") as f:
    w = csv.DictWriter(f, fieldnames=["codigo", "cliente", "vendedor", "status",
                                      "classe_app", "classe_banco", "ultima_planilha",
                                      "ultima_banco", "notas_2026", "receita_2026"])
    w.writeheader()
    for d in diverg + frescor:
        w.writerow(d)

# ---------------- 6. SAÍDA P/ O APP (Fase 2: silver_distribuicao.json) ----------------
# O app AINDA NÃO lê este arquivo — ele é publicado no branch state pela rotina
# diária e ficará pronto p/ a Fase 2 (churn com data real de NF) após auditoria.
saida_app = {
    "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M"),
    "fonte": "silver.faturamento (NF de Venda, Distribuição PROPETZ, 2026+)",
    "ancora_planilha": f"{ancora[1]:02d}/{ancora[0]}",
    "resumo": {"iguais": iguais, "frescor": len(frescor), "divergencias": len(diverg),
               "defasagem": len(defasagem), "nomes_casados": len(depara),
               "sem_codigo": len(sem_match), "ambiguos": len(ambiguos)},
    "clientes": {cod: {
        "ultima_compra_real": str(b["ultima_compra"]),
        "notas_2026": int(b["notas"]),
        "receita_2026": round(float(b["receita"]), 2),
        "cidade": b.get("cidade") or "", "uf": b.get("uf") or ""}
        for cod, b in banco_por_cod.items()},
    "novos_sem_cadastro": [{"nome": b["cliente_nome"],
                            "receita_2026": round(float(b["receita"]), 2),
                            "ultima": str(b["ultima_compra"]),
                            "cidade": b.get("cidade") or "", "uf": b.get("uf") or ""}
                           for b in sem_match],
}
with open(os.path.join(BASE, "silver_distribuicao.json"), "w", encoding="utf-8") as f:
    json.dump(saida_app, f, ensure_ascii=False, separators=(",", ":"))

print(f"\nIGUAIS: {iguais} | FRESCOR: {len(frescor)} | DIVERGÊNCIA REAL: {len(diverg)} "
      f"| DEFASAGEM (diária mudaria hoje): {len(defasagem)}")
print(f"silver_distribuicao.json: {len(saida_app['clientes'])} clientes por código")
print(f"cegos: {len(cegos)} (saudáveis a conferir: {len(cegos_saudaveis)}) | "
      f"ambíguos: {len(ambiguos)}")
print(f"relatório: {rel}")
