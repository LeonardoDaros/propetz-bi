# -*- coding: utf-8 -*-
"""MÊS AO VIVO — coletor do banco silver p/ a página "Mês ao Vivo" do app.

Gera silver_mes_vivo.json com a receita do MÊS CORRENTE do canal Distribuição
(por dia, por vendedor, top clientes) + comparativo com o mês anterior no mesmo
dia + METAS por vendedor. Publicado no branch `state` pela silver_diaria.py
(mesma rotina horária do churn); o app online lê de lá.

REGRAS DO MANUAL (Manual_Integracao_Banco_Silver.md — não reinventar):
- receita = sum(valor_total), só tipo_faturamento='NF de Venda'
- canal = modelo_negocio_descricao='Distribuição PROPETZ'
- data_emissao é DATE sem hora (decisão 22/07: sem curva horária)
- METAS: NUNCA das tabelas meta_* do banco — fonte da verdade é o dash da TV
  (Metas_Vendedores.xlsx do projeto Demanda), bloco "100% da meta".
"""
import json
import os
import sys
from datetime import date, datetime

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
from ponte_db_silver import consultar          # noqa: E402
from util_comum import (abrir_ou_copiar, norm_cliente,      # noqa: E402
                        primeiro_nome_vendedor)             # reuso, regra nº 6

SAIDA = os.path.join(BASE, "silver_mes_vivo.json")
METAS_XLSX = os.path.join(os.path.dirname(BASE), "Demanda Curva abc",
                          "Demanda Curva Abc - Pet", "Metas_Vendedores.xlsx")
FILTRO = ("tipo_faturamento = 'NF de Venda' "
          "AND modelo_negocio_descricao = 'Distribuição PROPETZ'")
MESES_PT = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
            "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]


def _f(v):
    return float(v or 0)


def metas_do_mes(ano, mes):
    """Lê Metas_Vendedores.xlsx (aba METAS): segmento Propetz Distribuição,
    blocos lado a lado — esquerda=desbloqueia comissão, direita=100% da meta.
    Devolve {nome_vendedor: {"meta": 100%, "desbloqueio": 50%}}."""
    import openpyxl
    ws = openpyxl.load_workbook(abrir_ou_copiar(METAS_XLSX, "_metas_mv.xlsx"),
                                data_only=True)["METAS"]
    metas, col_esq, col_dir, dentro = {}, None, None, False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        b = row[1].value if len(row) > 1 else None
        b_txt = str(b or "").strip()
        eh_header = any(isinstance(c.value, datetime) for c in row[2:8])
        if eh_header:
            dentro = "distribui" in b_txt.lower()
            if dentro:
                col_esq = col_dir = None
                for c in row:
                    if isinstance(c.value, datetime) and c.value.year == ano \
                            and c.value.month == mes:
                        if col_esq is None:
                            col_esq = c.column          # 1º bloco: desbloqueio
                        else:
                            col_dir = c.column           # 2º bloco: 100% da meta
            continue
        if dentro and b_txt and col_esq:
            desb = _f(ws.cell(row[0].row, col_esq).value)
            meta = _f(ws.cell(row[0].row, col_dir).value) if col_dir else 0.0
            if meta > 0 or desb > 0:
                metas[b_txt] = {"meta": round(meta, 2),
                                "desbloqueio": round(desb, 2)}
    return metas


def coletar():
    hoje = date.today()
    ano, mes, dia = hoje.year, hoje.month, hoje.day
    ano_ant, mes_ant = (ano, mes - 1) if mes > 1 else (ano - 1, 12)
    if ano_ant < 2026:
        ano_ant = mes_ant = None    # banco cego pré-2026 (manual, regra 9)

    # -- mês corrente: por vendedor × dia --
    atual = consultar(f"""
        SELECT coalesce(vendedor_nome,'(sem vendedor)') AS v, data_emissao AS d,
               count(*) AS n, round(sum(valor_total)::numeric,2) AS r
        FROM silver.faturamento
        WHERE ano = %s AND mes = %s AND {FILTRO}
        GROUP BY 1, 2""", [ano, mes])
    # -- prova independente do total (guarda anti-besteira minha) --
    prova = consultar(f"""
        SELECT round(sum(valor_total)::numeric,2) AS r, count(*) AS n
        FROM silver.faturamento
        WHERE ano = %s AND mes = %s AND {FILTRO}""", [ano, mes])
    # TODOS os clientes do mês, por (cliente, vendedor) — sem LIMIT (auditoria
    # 06/08: top-30 global truncava a lista do vendedor em silêncio; e agrupar
    # só por cliente fundia receita de dois vendedores numa linha só)
    clientes_rows = consultar(f"""
        SELECT cliente_nome, coalesce(vendedor_nome,'(sem vendedor)') AS v,
               count(*) AS n, round(sum(valor_total)::numeric,2) AS r,
               max(uf) AS uf
        FROM silver.faturamento
        WHERE ano = %s AND mes = %s AND {FILTRO}
        GROUP BY 1, 2""", [ano, mes])
    anterior = []
    if mes_ant:
        anterior = consultar(f"""
            SELECT coalesce(vendedor_nome,'(sem vendedor)') AS v,
                   data_emissao AS d, round(sum(valor_total)::numeric,2) AS r
            FROM silver.faturamento
            WHERE ano = %s AND mes = %s AND {FILTRO}
            GROUP BY 1, 2""", [ano_ant, mes_ant])

    # -- agrega (consultar devolve lista de DICTS — chave = alias do SELECT) --
    total_r = round(sum(_f(x["r"]) for x in atual), 2)
    total_n = sum(int(x["n"]) for x in atual)
    prova_r, prova_n = _f(prova[0]["r"]), int(prova[0]["n"] or 0)
    if abs(total_r - prova_r) > 0.01 or total_n != prova_n:
        print(f"ABORTADO: agregacao nao bate com a prova independente "
              f"({total_r} x {prova_r} | {total_n} x {prova_n})")
        return 1

    por_dia = {}
    for x in atual:
        k = str(x["d"])[:10]
        pd_ = por_dia.setdefault(k, {"r": 0.0, "n": 0})
        pd_["r"] = round(pd_["r"] + _f(x["r"]), 2)
        pd_["n"] += int(x["n"])

    # metas em MELHOR ESFORÇO (auditoria 06/08): planilha ausente/trancada/sem
    # o mês não pode matar a receita — a página lida com "sem meta cadastrada"
    try:
        metas = metas_do_mes(ano, mes)
        if not metas:
            print("AVISO: nenhuma meta achada p/ o mês na Metas_Vendedores.xlsx")
    except Exception as e:
        metas = {}
        print(f"AVISO: metas indisponíveis ({type(e).__name__}: {e}) — publico sem metas")
    # de-para banco -> nome da meta/app pelo 1º nome, já com o merge de
    # carteiras do app (fonte única em util_comum — mesmo critério da sombra)
    meta_por_1o = {m.split(" ")[0].lower(): m for m in metas}

    def _nome_app(v):
        return meta_por_1o.get(primeiro_nome_vendedor(v), str(v))

    vend = {}
    for x in atual:
        vd = vend.setdefault(_nome_app(x["v"]), {"receita": 0.0, "notas": 0})
        vd["receita"] = round(vd["receita"] + _f(x["r"]), 2)
        vd["notas"] += int(x["n"])

    # clientes do mês: identidade = norm_cliente (funde variantes '| SN' etc.)
    # POR VENDEDOR unificado — receita de cada vendedor fica separada
    cli = {}
    for x in clientes_rows:
        chave = (norm_cliente(x["cliente_nome"]), _nome_app(x["v"]))
        c = cli.setdefault(chave, {"nome": "", "notas": 0, "receita": 0.0, "uf": ""})
        nome_raw = str(x["cliente_nome"] or "").strip()
        # exibe a variante mais curta (sem sufixo de regime tributário)
        if nome_raw and (not c["nome"] or len(nome_raw) < len(c["nome"])):
            c["nome"] = nome_raw
        c["notas"] += int(x["n"])
        c["receita"] = round(c["receita"] + _f(x["r"]), 2)
        c["uf"] = c["uf"] or str(x["uf"] or "")
    # contagens SEM dupla contagem: distinto global e por vendedor
    total_clientes = len({norm for norm, _ in cli})
    for (norm, vnome) in cli:
        vd = vend.setdefault(vnome, {"receita": 0.0, "notas": 0})
        vd.setdefault("_norms", set()).add(norm)
    for vd in vend.values():
        vd["clientes"] = len(vd.pop("_norms", set()))
    # prova cruzada extra: mesma tabela, GROUP BY diferente → receitas batem
    soma_cli = round(sum(c["receita"] for c in cli.values()), 2)
    if abs(soma_cli - total_r) > 0.05:
        print(f"ABORTADO: receita por cliente ({soma_cli}) diverge do total ({total_r})")
        return 1

    ant_total, ant_ate_dia = 0.0, 0.0
    ant_vend = {}
    for x in anterior:
        r = _f(x["r"])
        ant_total = round(ant_total + r, 2)
        av = ant_vend.setdefault(_nome_app(x["v"]), {"total": 0.0, "ate_dia": 0.0})
        av["total"] = round(av["total"] + r, 2)
        try:
            if int(str(x["d"])[8:10]) <= dia:
                ant_ate_dia = round(ant_ate_dia + r, 2)
                av["ate_dia"] = round(av["ate_dia"] + r, 2)
        except ValueError:
            pass

    vendedores = []
    for nome in sorted(set(vend) | set(metas), key=lambda x: -vend.get(x, {}).get("receita", 0)):
        m = metas.get(nome, {})
        av = ant_vend.get(nome, {})
        vendedores.append({
            "nome": nome,
            "receita": vend.get(nome, {}).get("receita", 0.0),
            "notas": vend.get(nome, {}).get("notas", 0),
            "clientes": vend.get(nome, {}).get("clientes", 0),
            "meta": m.get("meta"), "meta_desbloqueio": m.get("desbloqueio"),
            "anterior_total": av.get("total", 0.0),
            "anterior_ate_dia": av.get("ate_dia", 0.0),
        })

    import calendar
    saida = {
        "gerado_em": f"{datetime.now():%Y-%m-%d %H:%M}",
        "mes": f"{ano}-{mes:02d}",
        "mes_nome": f"{MESES_PT[mes]}/{ano}",
        "dia": dia, "dias_no_mes": calendar.monthrange(ano, mes)[1],
        "total": {"receita": total_r, "notas": total_n,
                  "clientes": total_clientes},
        "anterior": {"mes_nome": f"{MESES_PT[mes_ant]}/{ano_ant}" if mes_ant else None,
                     "receita_total": ant_total, "receita_ate_dia": ant_ate_dia},
        "por_dia": [{"d": k, **v} for k, v in sorted(por_dia.items())],
        "vendedores": vendedores,
        # TODOS os clientes do mês (a página corta o top-30 só na visão do
        # gestor; o vendedor vê a lista COMPLETA dele)
        "top_clientes": sorted(
            [{"nome": c["nome"], "vendedor": vnome, "notas": c["notas"],
              "receita": c["receita"], "uf": c["uf"]}
             for (_, vnome), c in cli.items()],
            key=lambda x: -x["receita"])[:1500],
    }
    json.dump(saida, open(SAIDA, "w", encoding="utf-8"),
              ensure_ascii=False, separators=(",", ":"))
    kb = os.path.getsize(SAIDA) / 1024
    if len(cli) > 1500:
        print(f"AVISO: {len(cli)} clientes no mês — lista cortada em 1500 no json")
    print(f"mes ao vivo: {saida['mes_nome']} dia {dia} | receita R$ {total_r:,.2f} "
          f"| {total_n} notas | {saida['total']['clientes']} clientes "
          f"| vendedores: {[v['nome'].split(' ')[0] for v in vendedores]} "
          f"| metas casadas: {sum(1 for v in vendedores if v['meta'])} "
          f"| json {kb:.0f} KB")
    return 0


if __name__ == "__main__":
    sys.exit(coletar())
