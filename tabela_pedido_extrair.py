# -*- coding: utf-8 -*-
"""Extrai o modelo de dados p/ a tabela de pedidos.
ESTRUTURA (categorias, ordem, múltiplos, barras) vem da TABELA FOB antiga;
PREÇO/PROMO/ESTOQUE/CUSTO vêm do arquivo-mestre 'Novos valores distribuicao'
(aba do mês: col F=SKU, O=FOB tabela, Y=PROMO FOB, P=FOB+IPI, I=Estoque, L=Custo TRADE)."""
import sys, json, re, os, shutil
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

BASE = r"C:\Users\leoda\OneDrive\Área de Trabalho\Automação Dados\Automação Distribuição"
ARQ = BASE + r"\TABELA FOB AGOSTO 2026 xlsx.xlsx"
OUT = BASE + r"\_modelo_tabela.json"
NOVOS = (r"C:\Users\leoda\OneDrive\Área de Trabalho\Dcorp\ATUALIZAR MES A MES"
         r"\Tabela de Preço\Novos valores distribuicao (1).xlsx")
ABA_MES = "Ago-26"

def _abrivel(caminho, apelido):
    """Arquivo aberto no Excel/OneDrive trava a leitura direta — copia p/ temp."""
    try:
        with open(caminho, "rb"):
            return caminho
    except PermissionError:
        destino = os.path.join(os.environ.get("TEMP", "."), apelido)
        shutil.copy2(caminho, destino)
        print(f"(fonte travada — lendo cópia: {apelido})")
        return destino

ARQ = _abrivel(ARQ, "_fob_antiga.xlsx")
wbv = openpyxl.load_workbook(ARQ, data_only=True)    # valores
wbf = openpyxl.load_workbook(ARQ, data_only=False)   # fórmulas

fob_v, fob_f = wbv["FOB"], wbf["FOB"]
p2 = wbv["Planilha2"]
promo_ws = wbv["PROMO"]
sug = wbv["VALOR SUGERIDO"]

# ---- Planilha2: fonte da verdade de preços (desc -> a prazo, codigo) ----
precos = {}   # sku -> preco a prazo
nomes_p2 = {}
for r in range(1, p2.max_row + 1):
    cod = p2.cell(r, 4).value
    val = p2.cell(r, 2).value
    nome = p2.cell(r, 1).value
    if cod and isinstance(val, (int, float)):
        precos[str(cod).strip()] = float(val)
        nomes_p2[str(cod).strip()] = str(nome or "").strip()

# ---- PROMO: sku -> valor (>0) ----
promos_raw = {}
for r in range(2, promo_ws.max_row + 1):
    sku = promo_ws.cell(r, 1).value
    val = promo_ws.cell(r, 3).value
    if sku and isinstance(val, (int, float)) and val > 0:
        promos_raw[str(sku).strip()] = float(val)

# ---- FONTE NOVA (mestre de preços): sobrepõe preço/promo/estoque/custo ----
# cópia p/ temp: o OneDrive às vezes trava leitura direta do arquivo
_tmp_novos = os.path.join(os.environ.get("TEMP", "."), "_novos_valores_dist.xlsx")
shutil.copy2(NOVOS, _tmp_novos)  # sempre via cópia: o mestre vive aberto no Excel
wn = openpyxl.load_workbook(_tmp_novos, data_only=True)[ABA_MES]
novos = {}   # sku -> {fob, promo, ipi, estoque, custo}
for r in range(28, wn.max_row + 1):
    sku = wn.cell(r, 6).value      # F
    fob = wn.cell(r, 15).value     # O
    if not sku or not isinstance(fob, (int, float)) or fob <= 0:
        continue
    sku = str(sku).strip()
    p_cipi = wn.cell(r, 16).value  # P = FOB+IPI
    ipi = None
    if isinstance(p_cipi, (int, float)) and p_cipi >= fob > 0:
        ipi = round(p_cipi / fob - 1, 4)
    y = wn.cell(r, 25).value       # Y = PROMO FOB
    promo = float(y) if isinstance(y, (int, float)) and 0 < y < fob - 0.01 else None
    est = wn.cell(r, 9).value      # I
    custo = wn.cell(r, 12).value   # L = Custo TRADE
    novos[sku] = {"fob": round(float(fob), 2), "promo": round(promo, 2) if promo else None,
                  "ipi": ipi, "estoque": est if isinstance(est, (int, float)) else None,
                  "custo": round(float(custo), 2) if isinstance(custo, (int, float)) else None,
                  "promo_bruta": float(y) if isinstance(y, (int, float)) and y > 0 else None}

# preços da fonte nova SUBSTITUEM os da Planilha2 antiga
atualizados, mantidos_antigos = [], []
for sku in list(precos.keys()):
    if sku in novos:
        if abs(precos[sku] - novos[sku]["fob"]) > 0.01:
            atualizados.append((sku, round(precos[sku], 2), novos[sku]["fob"]))
        precos[sku] = novos[sku]["fob"]
    else:
        mantidos_antigos.append(sku)

# promos: SÓ da fonte nova (a aba PROMO antiga está vencida)
promos_raw = {sku: d["promo"] for sku, d in novos.items() if d["promo"]}

# promo REAL = valor < preço de tabela (tolerância 1 centavo)
promos_reais = {}
promos_lixo = {}
for sku, v in promos_raw.items():
    tab = precos.get(sku)
    if tab is None:
        promos_lixo[sku] = {"promo": v, "tabela": None, "motivo": "sem preço de tabela"}
    elif v < tab - 0.01:
        promos_reais[sku] = {"promo": round(v, 2), "tabela": round(tab, 2),
                             "desc_pct": round((1 - v / tab) * 100, 1)}
    else:
        promos_lixo[sku] = {"promo": round(v, 2), "tabela": round(tab, 2),
                            "motivo": "promo >= tabela (desconto zero ou negativo)"}

# ---- FOB: categorias, produtos, IPI, barras, sem-estoque, validações ----
produtos = []
categoria = None
sem_estoque = []
for r in range(16, 195):
    b = fob_v.cell(r, 2).value
    d = fob_v.cell(r, 4).value
    if b is not None and str(b).strip().upper().startswith(("CÓDIGO", "CODIGO")):
        categoria = str(fob_v.cell(r, 4).value or "").strip()
        continue
    if not b or not d:
        continue
    if not isinstance(fob_v.cell(r, 7).value, (int, float, type(None))):
        # linha-cabeçalho fora do padrão (ex.: col C='CÓDIGO ' com espaço)
        categoria = str(d or "").strip() or categoria
        continue
    sku = str(b).strip()
    nome = str(d).strip()
    ipi = fob_v.cell(r, 7).value or 0
    barras = fob_v.cell(r, 3).value
    fill = fob_v.cell(r, 4).fill
    cor = None
    try:
        if fill and fill.fgColor is not None:
            cor = str(fill.fgColor.rgb or fill.fgColor.theme)
    except Exception:
        pass
    vermelho = (cor == "FFFF0000") or "SEM ESTOQUE" in nome.upper() or "ESGOTADO" in nome.upper()
    hidden = fob_v.row_dimensions[r].hidden if r in fob_v.row_dimensions else False
    preco = precos.get(sku)
    item = {"linha_orig": r, "sku": sku, "nome": nome, "barras": str(barras or "").strip(),
            "categoria": categoria, "ipi": float(ipi or 0),
            "preco_tabela": round(preco, 2) if preco is not None else None,
            "sem_estoque": bool(vermelho), "linha_oculta": bool(hidden)}
    if vermelho:
        sem_estoque.append(sku)
    produtos.append(item)

# ---- validações de quantidade (múltiplos) por linha ----
mult_por_linha = {}
for dv in fob_f["FOB"].data_validations.dataValidation if hasattr(fob_f["FOB"], "data_validations") else []:
    pass
# openpyxl: acesso correto
fobf_ws = wbf["FOB"]
for dv in fobf_ws.data_validations.dataValidation:
    if dv.type != "list" or not dv.formula1:
        continue
    seq = [s.strip().strip('"') for s in str(dv.formula1).strip('"').split(",")]
    nums = []
    for s in seq:
        try:
            nums.append(int(float(s)))
        except Exception:
            pass
    # passo REAL = degrau modal da lista (auditoria 2026-07-21: a versão anterior
    # devolvia sempre o 1º valor — Cortantes viravam "múltiplos de 4" sendo caixa de 2)
    passo = None
    nn = sorted(set(n for n in nums if n > 0))
    if len(nn) >= 2:
        difs = [nn[i+1] - nn[i] for i in range(len(nn)-1)]
        freq = {}
        for d in difs:
            freq[d] = freq.get(d, 0) + 1
        passo = max(freq, key=freq.get)
    elif len(nn) == 1:
        passo = nn[0]
    lista = sorted(set(nums)) if nums else []
    if lista and lista[0] != 0:
        lista = [0] + lista
    for rng in dv.sqref.ranges:
        for row in range(rng.min_row, rng.max_row + 1):
            if rng.min_col <= 5 <= rng.max_col:
                mult_por_linha[row] = {"passo": passo, "max": nn[-1] if nn else None,
                                       "lista": lista}

# Regra por categoria (decisão Leonardo 2026-07-21):
#  - equipamentos/lâminas/adaptadores/tesouras/estojos/secadores: o nº é QTD MÍNIMA
#    (acima dela, qualquer quantidade vale — 5 máquinas pode, mínimo 4);
#  - demais (acessórios): CAIXA FECHADA — múltiplos exatos.
_MIN_KEYS = ("maquina", "máquina", "lamina", "lâmina", "adaptador", "tesoura",
             "estojo", "secador")

def _regra_da_categoria(cat):
    c = str(cat or "").lower()
    return "min" if any(k in c for k in _MIN_KEYS) else "caixa"

for item in produtos:
    m = mult_por_linha.get(item["linha_orig"])
    item["multiplo"] = m["passo"] if m else None
    item["qtd_max"] = m["max"] if m else None
    item["qtd_lista"] = m["lista"] if m else None  # lista ORIGINAL, verbatim
    _lista_pos = [v for v in (item["qtd_lista"] or []) if v > 0]
    item["qtd_min"] = _lista_pos[0] if _lista_pos else (item["multiplo"] or 1)
    item["regra_qtd"] = _regra_da_categoria(item["categoria"])

# ---- quantidade mínima por modelo (linhas ocultas 212-222: D=modelo, E=qtd) ----
minimos = []
for r in range(212, 223):
    modelo = fob_v.cell(r, 4).value
    qtd = fob_v.cell(r, 5).value
    if modelo and isinstance(qtd, (int, float)):
        par = [str(modelo).strip(), int(qtd)]
        if par not in minimos:
            minimos.append(par)

# ---- fallback de preço: 3 SKUs com referência quebrada usam o valor da FOB F ----
FALLBACK_F = {"L1PCPA-100": 104, "H1DMBA-100": 116, "L1C75A-100": 139}
for sku, linha in FALLBACK_F.items():
    if sku in precos:
        continue
    if sku in novos:
        precos[sku] = novos[sku]["fob"]  # fonte nova ganha do valor antigo da FOB
        continue
    # blindagem p/ meses futuros: só usa a linha se o SKU dela for o esperado
    # (linhas deslocam quando a FOB muda — melhor faltar preço do que pegar errado)
    if str(fob_v.cell(linha, 2).value or "").strip() != sku:
        print(f"AVISO: fallback de {sku} ignorado — linha {linha} agora tem outro SKU")
        continue
    v = fob_v.cell(linha, 6).value
    if isinstance(v, (int, float)):
        precos[sku] = float(v)
for it in produtos:
    if it["preco_tabela"] is None and it["sku"] in precos:
        it["preco_tabela"] = round(precos[it["sku"]], 2)
    nv = novos.get(it["sku"])
    if nv:
        it["preco_tabela"] = nv["fob"]
        if nv["ipi"] is not None:
            it["ipi"] = nv["ipi"]          # IPI da fonte nova (P/O − 1)
        it["estoque"] = nv["estoque"]
        if nv["estoque"] is not None:
            # o estoque do MESTRE manda nos 2 sentidos: zera -> sai do pedido;
            # tem estoque -> volta (a marca vermelha manual da FOB antiga envelhece)
            it["sem_estoque"] = nv["estoque"] <= 0
    if not it["sem_estoque"]:
        # item à venda não pode carregar sufixo herdado do formulário antigo
        for suf in ("(SEM ESTOQUE)", "( SEM ESTOQUE )", "(ESGOTADO)"):
            it["nome"] = it["nome"].replace(suf, "")
        it["nome"] = " ".join(it["nome"].split())

# ---- troca de variante -100/-200: o formulário antigo lista uma variante que
# morreu (sem estoque ou sem preço no mestre) enquanto a irmã tem estoque ----
_ja_listados = {p["sku"] for p in produtos}
trocas_variante = []
for it in produtos:
    sku = it["sku"]
    precisa = it["sem_estoque"] or sku not in novos
    if not precisa:
        continue
    for a, b in (("-100", "-200"), ("-200", "-100")):
        if not sku.endswith(a):
            continue
        alt = sku[: -len(a)] + b
        nv = novos.get(alt)
        if nv and (nv["estoque"] or 0) > 0 and alt not in _ja_listados:
            trocas_variante.append([sku, alt, nv["fob"], nv["estoque"]])
            it["sku"] = alt
            it["preco_tabela"] = nv["fob"]
            if nv["ipi"] is not None:
                it["ipi"] = nv["ipi"]
            it["estoque"] = nv["estoque"]
            it["sem_estoque"] = False
            it["barras"] = ""  # barras da variante nova não consta na FOB antiga
            _ja_listados.add(alt)
        break

# ---- VALOR SUGERIDO (só Propetz): sku -> sugerido, minimo ----
sugeridos = {}
for r in range(2, sug.max_row + 1):
    cod = sug.cell(r, 1).value
    f = sug.cell(r, 6).value
    g = sug.cell(r, 7).value
    if cod and isinstance(f, (int, float)) and str(cod).strip() in precos:
        sugeridos[str(cod).strip()] = {"sugerido": round(float(f), 2),
                                       "minimo": round(float(g), 2) if isinstance(g, (int, float)) else None}

# ---- custos (Base Mãe via abc_valor.json) p/ conferência de margem ----
try:
    abc = json.load(open(BASE + r"\abc_valor.json", encoding="utf-8"))
    custos = abc.get("custo_unitario", {})
except Exception:
    custos = {}
abaixo_custo = []
for it in produtos:
    nv = novos.get(it["sku"]) or {}
    c = nv.get("custo") or custos.get(it["sku"])   # Custo TRADE do mestre > cmv Base Mãe
    if not c or not it["preco_tabela"]:
        continue
    vigente = (promos_reais.get(it["sku"]) or {}).get("promo") or it["preco_tabela"]
    if vigente < c:
        abaixo_custo.append({"sku": it["sku"], "nome": it["nome"], "custo": round(c, 2),
                             "tabela": it["preco_tabela"], "vigente": round(vigente, 2)})

modelo = {"vigencia": "Agosto/2026", "produtos": produtos,
          "promos_reais": promos_reais, "promos_descartadas": promos_lixo,
          "minimos": minimos, "sugeridos": sugeridos,
          "sem_estoque": sem_estoque, "abaixo_custo": abaixo_custo,
          "trocas_variante": trocas_variante, "precos_antigos_mantidos": mantidos_antigos}
json.dump(modelo, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

cats = []
for p in produtos:
    if p["categoria"] not in cats:
        cats.append(p["categoria"])
print(f"produtos: {len(produtos)} | categorias: {len(cats)} | promos REAIS: {len(promos_reais)} "
      f"| promos descartadas: {len(promos_lixo)} | sem estoque: {len(sem_estoque)} "
      f"| sugeridos: {len(sugeridos)} | abaixo do custo: {len(abaixo_custo)}")
print("\nPROMOS REAIS:")
for sku, p in sorted(promos_reais.items(), key=lambda x: -x[1]["desc_pct"]):
    nome = next((i["nome"] for i in produtos if i["sku"] == sku), nomes_p2.get(sku, "?"))
    est = " [SEM ESTOQUE]" if sku in sem_estoque else ""
    print(f"  {sku}: {p['tabela']} -> {p['promo']} (-{p['desc_pct']}%){est} {nome[:45]}")
print("\nABAIXO DO CUSTO (preço vigente < Custo TRADE):")
for a in abaixo_custo:
    print(f"  {a['sku']}: vigente {a['vigente']} < custo {a['custo']} — {a['nome'][:45]}")

print(f"\nPREÇOS ATUALIZADOS pela fonte nova: {len(atualizados)}")
for sku, antigo, novo_v in atualizados[:15]:
    print(f"  {sku}: {antigo} -> {novo_v}")
if len(atualizados) > 15:
    print(f"  ... e mais {len(atualizados) - 15}")
print(f"\nTROCAS DE VARIANTE (listada morta -> irmã com estoque): {len(trocas_variante)}")
for velho, novo_sku, fob, est in trocas_variante:
    print(f"  {velho} -> {novo_sku} (fob {fob}, estoque {est})")
print(f"SEM preço na fonte nova (mantido o antigo — CONFERIR): {len(mantidos_antigos)}")
print("  ", mantidos_antigos)
_skus_fob = {p["sku"] for p in produtos}
extras = [(s, d) for s, d in novos.items()
          if s not in _skus_fob and (d["estoque"] or 0) > 0]
print(f"NA FONTE NOVA COM ESTOQUE, MAS FORA DO FORMULÁRIO: {len(extras)}")
for s, d in extras[:12]:
    print(f"  {s}: fob {d['fob']} | estoque {d['estoque']}")
_promo_vencida = [s for s, d in novos.items() if d["promo_bruta"] and not d["promo"]]
print(f"PROMO VENCIDA/IGUAL na fonte nova (ignorada): {len(_promo_vencida)} -> {_promo_vencida[:10]}")
print("\nSEM MÚLTIPLO DEFINIDO:", sum(1 for p in produtos if not p["multiplo"]))
print("SEM PREÇO NA PLANILHA2:", [p["sku"] for p in produtos if p["preco_tabela"] is None])
print("MÍNIMOS:", minimos)
