"""Contraprovas independentes: funções reais via AST, arquivos temporários e API falsa.

Não importa app.py, não lê cadastro/estado real e não faz rede. As expectativas
descrevem o comportamento seguro; falham quando um defeito é reproduzido.
"""
import ast
import base64
import copy
from contextlib import nullcontext
from datetime import datetime, date
import io
import hashlib
import hmac
import json
import os
from pathlib import Path
import tempfile
import threading
import types
import unittest
from unittest.mock import Mock, patch
import uuid
import secrets

import openpyxl
import pandas as pd
import yaml
from util_comum import parse_label_ym, normalize_vendor
import teste_etapa1_deploy as deploy_fixture


APP_PATH = Path(os.environ.get('PROPETZ_AUDIT_APP', Path(__file__).with_name('app.py')))
TREE = ast.parse(APP_PATH.read_text(encoding='utf-8-sig'))
FUNCTIONS = [copy.deepcopy(n) for n in TREE.body if isinstance(n, ast.FunctionDef)]
for node in FUNCTIONS:
    node.decorator_list = []
CODE = compile(ast.Module(body=FUNCTIONS, type_ignores=[]), '<app-functions-only>', 'exec')


class Rerun(Exception):
    pass


class FakeUI:
    def __init__(self):
        self.session_state = {'authenticated': True, 'username': 'synthetic-admin', 'role': 'admin'}
        self.cache_data = types.SimpleNamespace(clear=Mock())
        self.uploaded = io.BytesIO(b'synthetic-upload')
        self.upload_key = None
        self.messages = []

    def __getattr__(self, name):
        if name in ('form', 'spinner', 'expander'):
            return lambda *a, **kw: nullcontext()
        if name == 'columns':
            return lambda spec, **kw: [nullcontext() for _ in range(spec if isinstance(spec, int) else len(spec))]
        if name in ('button', 'form_submit_button'):
            return lambda *a, **kw: False
        if name == 'text_input':
            return lambda *a, **kw: ''
        if name in ('selectbox', 'radio'):
            return lambda label, options, **kw: options[0] if len(options) and kw.get('index', 0) is not None else None
        if name == 'file_uploader':
            def uploader(*a, **kw):
                widget_key = kw.get('key', 'implicit-widget')
                if self.upload_key is None:
                    self.upload_key = widget_key
                return self.uploaded if widget_key == self.upload_key else None
            return uploader
        if name == 'rerun':
            def rerun():
                raise Rerun()
            return rerun
        return lambda *a, **kw: self.messages.append((name, a))


class CrosscuttingAudit(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory(prefix='propetz-security-synthetic-')
        self.addCleanup(self.temp.cleanup)
        self.folder = Path(self.temp.name)
        self.ui = FakeUI()
        self.ns = {'__file__': str(self.folder / 'app.py'), 'os': os, 'json': json,
                   'base64': base64, 'datetime': datetime, 'date': date, 'pd': pd, 'yaml': yaml,
                   'io': io, 'uuid': uuid, 'openpyxl': openpyxl,
                   'hashlib': hashlib, 'hmac': hmac, 'secrets': secrets,
                   '_SCRYPT_N': 16384, '_SCRYPT_R': 8, '_SCRYPT_P': 1,
                   '_parse_label_ym': parse_label_ym, 'normalize_vendor': normalize_vendor,
                   'st': self.ui, '_GH_WRITE_LOCK': threading.Lock(), '_STATE_RAW_CACHE': {},
                   '_GH_STATE_BRANCH': 'state', 'USERS_FILE': str(self.folder / 'users.yaml')}
        exec(CODE, self.ns)
        self.ns['_gh_token'] = lambda: None
        self.ns['_refresh_session_access'] = lambda: None
        self.ns['_session_expired'] = lambda: False
        self.ns['load_silver_distribuicao'] = lambda: {}

    def valid_upload(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'IA'
        for col, value in enumerate(['Cliente sintético', 'SP', 'C1', 'Carteira fictícia', 'Ativo', 100, 200], 4):
            sheet.cell(4, col, value)
        sheet.cell(3, 9, 'jan/26')
        sheet.cell(3, 10, 'fev/26')
        data = io.BytesIO()
        workbook.save(data)
        workbook.close()
        return data

    def test_invalid_upload_preserves_source_and_never_publishes(self):
        target = self.folder / 'Relatorio Distribuidores Mensal.xlsx'
        target.write_bytes(b'original-synthetic-source')
        put = Mock(return_value=True)
        self.ns['_gh_put_file'] = put
        try:
            self.ns['_handle_planilha_upload'](io.BytesIO(b'this-is-not-an-xlsx'))
        except (ValueError, OSError):
            pass
        self.assertEqual(target.read_bytes(), b'original-synthetic-source',
                         'Upload inválido substituiu a planilha válida antes de validar.')
        put.assert_not_called()

    def test_admin_upload_is_not_reprocessed_on_immediate_rerun(self):
        self.ns['load_users'] = lambda: {'users': {'synthetic-admin': {
            'name': 'Admin de teste', 'role': 'admin', 'password': 'synthetic'}}}
        self.ns['_load_access_log'] = lambda: []
        handle = Mock(return_value=True)
        self.ns['_handle_planilha_upload'] = handle
        for _ in range(2):
            try:
                self.ns['page_admin'](['Carteira fictícia'])
            except Rerun:
                pass
        self.assertEqual(handle.call_count, 1,
                         'Mesmo arquivo do uploader provocou novo processamento/PUT após rerun.')

    def test_local_write_failure_is_never_reported_as_persisted(self):
        # Caminho pai inexistente simula falha de armazenamento sem tocar disco real.
        target = self.folder / 'does-not-exist' / 'state.json'
        _, ok = self.ns['_gh_mutate_json']('synthetic.json', str(target),
            lambda state: {'values': state['values'] + ['synthetic']}, {'values': []})
        self.assertFalse(ok, 'Adaptador declarou sucesso embora nenhum arquivo exista.')
        self.assertFalse(target.exists())

    def test_unreadable_local_json_is_preserved(self):
        target = self.folder / 'state.json'
        target.write_bytes(b'{invalid-synthetic')
        _, ok = self.ns['_gh_mutate_json']('synthetic.json', str(target),
            lambda state: {'values': ['synthetic']}, {'values': []})
        self.assertFalse(ok)
        self.assertEqual(target.read_bytes(), b'{invalid-synthetic')

    def test_atomic_local_replace_failure_preserves_previous_bytes(self):
        target = self.folder / 'state.json'
        target.write_bytes(b'{"values":["original"]}')
        with patch.object(os, 'replace', side_effect=PermissionError('synthetic disk failure')):
            _, ok = self.ns['_gh_mutate_json']('synthetic.json', str(target),
                lambda state: {'values': ['synthetic']}, {'values': []})
        self.assertFalse(ok)
        self.assertEqual(target.read_bytes(), b'{"values":["original"]}')
        self.assertFalse(list(self.folder.glob('*.tmp')))

    def test_failed_inactivation_never_approves_request(self):
        state = {'requests': [{'client_id': 'C1', 'status': 'pendente'}]}
        self.ns['inactivate_clients'] = lambda ids: False
        self.ns['INACTIVE_REQUESTS_FILE'] = str(self.folder / 'requests.json')
        def mutate(remote, path, apply, default):
            state.update(apply(copy.deepcopy(state)))
            return copy.deepcopy(state), True
        self.ns['_gh_mutate_json'] = mutate
        result = self.ns['decide_inactivation_request']('C1', True, 'Admin fictício')
        self.assertFalse(result)
        self.assertEqual(state['requests'][0]['status'], 'pendente',
                         'Pedido sumiu da fila como aprovado apesar da inativação falhar.')

    def test_direct_approval_stops_when_inactivation_fails(self):
        self.ns['inactivate_clients'] = lambda ids: False
        mutate = Mock()
        self.ns['_gh_mutate_json'] = mutate
        self.ns['INACTIVE_REQUESTS_FILE'] = str(self.folder / 'requests.json')
        self.assertFalse(self.ns['add_inactivation_request']('C1', 'Sintético', 'Carteira',
                         motivo='Outro', observacao='Sintético', direct_approve=True))
        mutate.assert_not_called()

    def test_valid_upload_uses_real_parser_before_remote_and_replaces_atomically(self):
        target = self.folder / 'Relatorio Distribuidores Mensal.xlsx'
        target.write_bytes(b'original-synthetic-source')
        uploaded = self.valid_upload()
        put = Mock(return_value=True)
        self.ns['_gh_token'] = lambda: 'synthetic-token'
        self.ns['_gh_put_file'] = put
        self.assertTrue(self.ns['_handle_planilha_upload'](uploaded))
        self.assertEqual(target.read_bytes(), uploaded.getvalue())
        self.assertEqual(put.call_count, 1)
        self.assertFalse(list(self.folder.glob('*.tmp')))

    def test_valid_upload_remote_failure_preserves_source_and_reports_failure(self):
        target = self.folder / 'Relatorio Distribuidores Mensal.xlsx'
        target.write_bytes(b'original-synthetic-source')
        self.ns['_gh_token'] = lambda: 'synthetic-token'
        self.ns['_gh_put_file'] = Mock(return_value=False)
        with self.assertRaisesRegex(ValueError, 'confirmar'):
            self.ns['_handle_planilha_upload'](self.valid_upload())
        self.assertEqual(target.read_bytes(), b'original-synthetic-source')
        self.assertFalse(list(self.folder.glob('*.tmp')))

    def test_empty_ia_workbook_is_rejected_before_any_write(self):
        workbook = openpyxl.Workbook()
        workbook.active.title = 'IA'
        uploaded = io.BytesIO()
        workbook.save(uploaded)
        workbook.close()
        put = Mock(return_value=True)
        self.ns['_gh_put_file'] = put
        with self.assertRaises(ValueError):
            self.ns['_handle_planilha_upload'](uploaded)
        put.assert_not_called()
        self.assertFalse((self.folder / 'Relatorio Distribuidores Mensal.xlsx').exists())

    def test_user_remote_failure_preserves_local_registry_and_raises(self):
        target = Path(self.ns['USERS_FILE'])
        target.write_text('users: {}\n', encoding='utf-8')
        self.ns['_gh_token'] = lambda: 'synthetic-token'
        self.ns['_gh_put_file'] = Mock(return_value=False)
        self.ns['_gh_get_file'] = Mock(return_value=(target.read_bytes(), 'synthetic-sha'))
        self.ns['_gh_put_file_status'] = Mock(return_value=(False, 503))
        users = self.ns['load_users']()
        users['users']['synthetic-user'] = {'name': 'Novo', 'role': 'vendedor'}
        with self.assertRaisesRegex(ValueError, 'confirmar'):
            self.ns['save_users'](users)
        self.assertEqual(target.read_text(encoding='utf-8'), 'users: {}\n')
        self.assertFalse(list(self.folder.glob('*.tmp')))

    def test_user_remote_confirmation_precedes_local_replacement(self):
        target = Path(self.ns['USERS_FILE'])
        target.write_text('users: {}\n', encoding='utf-8')
        self.ns['_gh_token'] = lambda: 'synthetic-token'
        def confirm(*args, **kwargs):
            self.assertEqual(target.read_text(encoding='utf-8'), 'users: {}\n')
            return True
        self.ns['_gh_put_file'] = Mock(side_effect=confirm)
        self.ns['_gh_get_file'] = Mock(return_value=(target.read_bytes(), 'synthetic-sha'))
        def confirm_status(*args, **kwargs):
            self.assertTrue(confirm())
            self.assertNotIn(b'_source_digest', args[1])
            return True, 200
        self.ns['_gh_put_file_status'] = Mock(side_effect=confirm_status)
        users = self.ns['load_users']()
        users['users']['synthetic-user'] = {'name': 'Novo', 'role': 'vendedor'}
        self.assertTrue(self.ns['save_users'](users))
        self.assertIn('synthetic-user', yaml.safe_load(target.read_text(encoding='utf-8'))['users'])
        self.assertNotIn('_source_digest', yaml.safe_load(target.read_text(encoding='utf-8')))

    def test_stale_user_snapshot_cannot_rollback_another_admin_change(self):
        target = Path(self.ns['USERS_FILE'])
        target.write_text(yaml.safe_dump({'users': {'synthetic-user': {
            'name': 'Pessoa', 'role': 'vendedor', 'vendor_filter': 'Carteira A'}}}), encoding='utf-8')
        remote = {'bytes': target.read_bytes(), 'sha': 'synthetic-revision-1'}
        first, second = self.ns['load_users'](), self.ns['load_users']()
        self.ns['_gh_token'] = lambda: 'synthetic-token'
        self.ns['_gh_get_file'] = lambda *a, **kw: (remote['bytes'], remote['sha'])
        def put_legacy(path, content, *args, **kwargs):
            remote.update(bytes=content, sha='synthetic-revision-2')
            return True
        def put_status(path, content, message, branch, sha, token):
            if sha != remote['sha']:
                return False, 409
            put_legacy(path, content)
            return True, 200
        self.ns['_gh_put_file'] = put_legacy
        self.ns['_gh_put_file_status'] = put_status
        first['users']['synthetic-user']['vendor_filter'] = 'Carteira B'
        self.ns['save_users'](first)
        second['users']['synthetic-user']['name'] = 'Nome corrigido'
        with self.assertRaisesRegex(ValueError, 'cadastro mudou'):
            self.ns['save_users'](second)
        self.assertEqual(yaml.safe_load(remote['bytes'])['users']['synthetic-user']['vendor_filter'], 'Carteira B')

    def test_user_put_conflict_never_retries_blindly(self):
        target = Path(self.ns['USERS_FILE'])
        target.write_text('users: {}\n', encoding='utf-8')
        users = self.ns['load_users']()
        users['users']['synthetic-user'] = {'name': 'Pessoa', 'role': 'admin'}
        self.ns['_gh_token'] = lambda: 'synthetic-token'
        self.ns['_gh_get_file'] = Mock(return_value=(target.read_bytes(), 'synthetic-sha'))
        put = Mock(return_value=(False, 409))
        self.ns['_gh_put_file_status'] = put
        with self.assertRaises(ValueError):
            self.ns['save_users'](users)
        self.assertEqual(put.call_count, 1)
        self.assertEqual(target.read_text(encoding='utf-8'), 'users: {}\n')

    def test_legacy_login_stays_valid_when_migration_cannot_publish(self):
        target = Path(self.ns['USERS_FILE'])
        legacy = hashlib.sha256(b'SyntheticLegacyPassword!').hexdigest()
        target.write_text(yaml.safe_dump({'users': {'synthetic-user': {
            'name': 'Pessoa', 'role': 'admin', 'password': legacy}}}), encoding='utf-8')
        self.ns['_gh_token'] = lambda: 'synthetic-token'
        self.ns['_gh_get_file'] = Mock(return_value=(target.read_bytes(), 'synthetic-sha'))
        self.ns['_gh_put_file_status'] = Mock(return_value=(False, 503))
        self.assertIsNotNone(self.ns['verify_login']('synthetic-user', 'SyntheticLegacyPassword!'))
        self.assertEqual(yaml.safe_load(target.read_bytes())['users']['synthetic-user']['password'], legacy)

    def test_admin_demotion_is_rechecked_before_registry_mutation(self):
        save = Mock(return_value=True)
        self.ns['save_users'] = save
        def demote():
            self.ui.session_state['role'] = 'vendedor'
            return None
        self.ns['_refresh_session_access'] = demote
        self.assertFalse(self.ns['_save_users_from_admin']({'users': {}}, 'Salvo'))
        save.assert_not_called()

    def test_rejected_upload_does_not_announce_success_or_rerun(self):
        self.ns['_handle_planilha_upload'] = Mock(side_effect=ValueError('Falha remota sintética'))
        self.ns['_render_planilha_upload']('Upload fictício', 'audit_upload')
        self.assertTrue(any(kind == 'error' for kind, _ in self.ui.messages))
        self.assertFalse(any(kind == 'success' for kind, _ in self.ui.messages))
        self.assertNotIn('_planilha_upload_generation', self.ui.session_state)

    def test_unreadable_remote_does_not_put_default(self):
        self.ns['_gh_token'] = lambda: 'synthetic-token'
        self.ns['_gh_get_file'] = lambda *a, **kw: (b'{invalid', 'synthetic-sha')
        put = Mock(return_value=(True, 200))
        self.ns['_gh_put_file_status'] = put
        result, ok = self.ns['_gh_mutate_json']('synthetic.json', str(self.folder / 'state.json'),
            lambda state: {'values': ['synthetic']}, {'values': []})
        self.assertFalse(ok)
        put.assert_not_called()

    def test_conflict_retry_keeps_concurrent_changes(self):
        states = [(b'{"values":["first"]}', 'sha-1'),
                  (b'{"values":["first","concurrent"]}', 'sha-2')]
        self.ns['_gh_token'] = lambda: 'synthetic-token'
        self.ns['_gh_get_file'] = Mock(side_effect=states)
        put = Mock(side_effect=[(False, 409), (True, 200)])
        self.ns['_gh_put_file_status'] = put
        result, ok = self.ns['_gh_mutate_json']('synthetic.json', str(self.folder / 'state.json'),
            lambda state: {'values': state['values'] + ['ours']}, {'values': []})
        self.assertTrue(ok)
        self.assertEqual(result['values'], ['first', 'concurrent', 'ours'])
        self.assertEqual(put.call_count, 2)


@unittest.skipUnless(deploy_fixture.GIT and deploy_fixture.POWERSHELL,
                     'Git e PowerShell necessários; somente repositórios temporários locais.')
class DeployMergeAudit(unittest.TestCase):
    def test_disallowed_file_added_only_by_merge_resolution_blocks_push(self):
        fixture = deploy_fixture.SafeDeployTests('test_no_changes_confirms_existing_sha')
        fixture.setUp()
        self.addCleanup(fixture.tearDown)
        fixture.git(fixture.local, 'checkout', '-b', 'synthetic-feature')
        fixture.write('app.py', 'feature = True\n')
        fixture.git(fixture.local, 'add', 'app.py')
        fixture.git(fixture.local, 'commit', '-m', 'Synthetic feature')
        fixture.git(fixture.local, 'checkout', 'main')
        fixture.write('.gitignore', 'CREDENCIAIS-LOCAL.md\nBackups/\nsynthetic-ignore/\n')
        fixture.git(fixture.local, 'add', '.gitignore')
        fixture.git(fixture.local, 'commit', '-m', 'Synthetic main change')
        fixture.git(fixture.local, 'merge', '--no-commit', '--no-ff', 'synthetic-feature')
        # Nenhum dos commits pais modifica este arquivo. Ele entra somente no merge.
        fixture.write('internal-synthetic.txt', 'synthetic local-only content\n')
        fixture.git(fixture.local, 'add', 'internal-synthetic.txt')
        fixture.git(fixture.local, 'commit', '-m', 'Synthetic merge resolution')
        result = fixture.deploy(expect_ok=False)
        self.assertIn('fora da lista', result.stdout)
        fixture.remote_unchanged()


if __name__ == '__main__':
    unittest.main(verbosity=2)
